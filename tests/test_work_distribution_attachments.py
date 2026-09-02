"""Tests for Work Distribution's per-employee email attachments
(Milestone 57, 2026-08-27): each flagged employee's own Doctor List (RGD
Coverage) or BM Monthly Trend (Manager Work Allocation) is attached to the
recipient's email as a real file -- one attachment per flagged employee
per reason, never one combined file per recipient, and the HTML/text body
itself is unchanged.

Uses the REAL pipelines (process_work_distribution_report/
process_manager_work_allocation_report/process_rbm_report) to produce real
findings, then build_notification_batch() to produce real drafts -- no
second, test-only calculation path. Hierarchy lookups are faked (a plain
in-memory dict), everything else runs the actual code.
"""

import io

import openpyxl
import pytest

import app.manager_work_allocation_rbm_service as rbm_svc
import app.manager_work_allocation_service as abm_svc
import app.work_distribution_notification_service as wdns
import app.work_distribution_service as wds


# --- Shared in-memory DB + fake hierarchy --------------------------------

def _in_memory_session_factory():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from database.connection import Base

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False)


# key: normalized employee_name -> hierarchy row. Deliberately simple --
# every BM reports to the SAME ABM/RBM pair, and vacancy/DOJ/SM/AGM/GM are
# out of scope for this feature.
_HIERARCHY_BY_NAME = {
    "rahul sharma": {
        "employee_code": "BM01", "employee_name": "Rahul Sharma", "email": "rahul.bm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
        "abm_code": "ABM01", "abm_name": "Anita ABM", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "second bm": {
        "employee_code": "BM02", "employee_name": "Second BM", "email": "second.bm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
        "abm_code": "ABM01", "abm_name": "Anita ABM", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "third bm": {
        "employee_code": "BM03", "employee_name": "Third BM", "email": "third.bm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
        "abm_code": "ABM01", "abm_name": "Anita ABM", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "fourth bm": {
        "employee_code": "BM04", "employee_name": "Fourth BM", "email": "fourth.bm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
        "abm_code": "ABM01", "abm_name": "Anita ABM", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "healthy bm": {
        "employee_code": "BM05", "employee_name": "Healthy BM", "email": "healthy.bm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
        "abm_code": "ABM01", "abm_name": "Anita ABM", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "anita abm": {
        "employee_code": "ABM01", "employee_name": "Anita ABM", "email": "anita.abm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1", "rbm_code": "RBM01", "rbm_name": "Rohit RBM",
    },
    "rohit rbm": {
        "employee_code": "RBM01", "employee_name": "Rohit RBM", "email": "rohit.rbm@example.com",
        "division": "Onyx", "source_sheet": "Sheet1",
    },
}
_HIERARCHY_BY_CODE = {row["employee_code"]: row for row in _HIERARCHY_BY_NAME.values()}


def _fake_find_by_employee_name(name):
    row = _HIERARCHY_BY_NAME.get((name or "").strip().lower())
    return [row] if row else []


def _fake_find_by_employee_code(code):
    return _HIERARCHY_BY_CODE.get(code)


def _fake_find_by_designation(division, source_sheet, designation):
    return None  # no SM/AGM/GM in these fixtures -- chain just ends there


@pytest.fixture(autouse=True)
def _wire_fakes(monkeypatch):
    factory = _in_memory_session_factory()
    monkeypatch.setattr("database.connection._ConfigSession", factory)

    for mod in (wdns,):
        monkeypatch.setattr(mod, "find_by_employee_name", _fake_find_by_employee_name)
        monkeypatch.setattr(mod, "find_by_employee_code", _fake_find_by_employee_code)
        monkeypatch.setattr(mod, "find_by_designation", _fake_find_by_designation)

    for mod in (wds, abm_svc, rbm_svc):
        monkeypatch.setattr(mod, "load_doj_by_code", lambda: {})
        monkeypatch.setattr(mod, "load_doj_by_name", lambda: {})
    # RGD Coverage resolves its own finding's display name via
    # find_by_employee_code(code) -- must resolve to the SAME fake
    # hierarchy row (by code) so build_notification_batch's own
    # find_by_employee_name(name) lookup then succeeds.
    monkeypatch.setattr(wds, "find_by_employee_code", _fake_find_by_employee_code)


# --- RGD Coverage fixtures -------------------------------------------------

def _rgd_doctor(bm_code, doctor_code, bm_visits=0, category="B-RGD"):
    return {
        "division": "Onyx", "doctor_code": doctor_code, "doctor_name": f"Dr. {doctor_code}",
        "speciality": "General", "category": category, "abm_rgd": "",
        "city": "City", "hq": "HQ", "region": "Region",
        "bm_code": bm_code, "abm_code": "", "bm_name": "", "abm_name": "",
        "bm_visit_count": bm_visits, "abm_visit_count": 0, "period_label": "Aug-2026",
    }


def _mwa_record(manager_code, manager_name, manager_designation, bm_code, bm_name, month, joint_days):
    return {
        "division": "Onyx", "emp_code": manager_code, "emp_name": manager_name,
        "emp_designation": manager_designation, "team_emp_code": bm_code, "team_emp_name": bm_name,
        "team_emp_designation": "BM", "month": month, "joint_days": joint_days,
    }


# --- 1. RGD Coverage: one flagged BM -> correct Doctor List attached ------

def test_flagged_rgd_bm_gets_correct_doctor_list_attached():
    doctors = [_rgd_doctor("BM01", "D1", bm_visits=0), _rgd_doctor("BM01", "D2", bm_visits=0)]
    wds.process_work_distribution_report(doctors)

    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")

    assert len(rbm_draft["attachments"]) == 1
    filename, file_bytes = rbm_draft["attachments"][0]
    assert filename == "Rahul Sharma - RGD Coverage.xlsx"

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Doctor Code", "Doctor Name", "Division", "City", "Visit Count", "Status"]
    doctor_codes = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert sorted(doctor_codes) == ["D1", "D2"]


# --- 2/3. Manager Work Allocation: flagged ABM / RBM -> correct BM Monthly
#          Trend attached -------------------------------------------------

def test_flagged_mwa_abm_gets_correct_bm_monthly_trend_attached():
    records = [
        _mwa_record("ABM01", "Anita ABM", "ABM", "BM01", "Rahul Sharma", "Aug-2026", 1),
    ]
    abm_svc.process_manager_work_allocation_report(records)

    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")

    attachment = next(
        (fn, b) for fn, b in rbm_draft["attachments"] if fn == "Anita ABM - Manager Work Allocation.xlsx"
    )
    filename, file_bytes = attachment
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "BM Name"
    assert header[-1] == "Status"  # ABM engine's own status_heading
    bm_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert bm_names == ["Rahul Sharma"]


def test_rbm_monthly_trend_attachment_uses_covered_heading_not_status():
    """The RBM engine's own artifact uses "Covered" as its status-column
    heading (Yes/No), distinct from the ABM engine's "Status" (Pass/Fail)
    -- see ui.work_distribution_employee_details_page's own
    BM_DETAIL_HEADINGS convention this mirrors. Exercises
    _monthly_trend_attachment() directly against a real
    get_employee_bm_monthly_history() shape, decoupled from the RBM
    engine's own tiered auto-flag floor (a separate business rule, not
    what this test is about)."""
    records = [
        _mwa_record("RBM01", "Rohit RBM", "RBM", "BM01", "Rahul Sharma", "Aug-2026", 0),
    ]
    rbm_svc.process_rbm_report(records)

    history = rbm_svc.get_employee_bm_monthly_history("Rohit RBM")
    filename, file_bytes = wdns._monthly_trend_attachment(history, "Rohit RBM", "Covered")
    assert filename == "Rohit RBM - Manager Work Allocation.xlsx"

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "BM Name"
    assert header[-1] == "Covered"  # RBM engine's own status_heading, not "Status"
    assert ws.cell(row=2, column=1).value == "Rahul Sharma"


# --- 4. Four flagged BMs under the same senior -> four separate,
#        employee-specific attachments on ONE email ------------------------

def test_four_flagged_bms_same_senior_get_four_separate_attachments():
    doctors = []
    for code in ("BM01", "BM02", "BM03", "BM04"):
        doctors.append(_rgd_doctor(code, f"D-{code}", bm_visits=0))
    wds.process_work_distribution_report(doctors)

    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")

    assert rbm_draft["employee_count"] == 4
    assert len(rbm_draft["attachments"]) == 4
    filenames = sorted(fn for fn, _ in rbm_draft["attachments"])
    assert filenames == [
        "Fourth BM - RGD Coverage.xlsx",
        "Rahul Sharma - RGD Coverage.xlsx",
        "Second BM - RGD Coverage.xlsx",
        "Third BM - RGD Coverage.xlsx",
    ]
    # never merged into one combined file
    assert len({fn for fn, _ in rbm_draft["attachments"]}) == 4


# --- 5. Filenames contain employee name + flag reason ---------------------

def test_attachment_filenames_contain_employee_name_and_reason():
    doctors = [_rgd_doctor("BM01", "D1", bm_visits=0)]
    wds.process_work_distribution_report(doctors)
    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")
    filename, _ = rbm_draft["attachments"][0]
    assert "Rahul Sharma" in filename
    assert "RGD Coverage" in filename
    assert filename.endswith(".xlsx")


# --- 6. Existing HTML email body remains unchanged -------------------------

def test_email_body_unaffected_by_attachments():
    doctors = [_rgd_doctor("BM01", "D1", bm_visits=0)]
    wds.process_work_distribution_report(doctors)
    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")

    assert "Rahul Sharma" in rbm_draft["body"]
    assert "Dear Rohit RBM" in rbm_draft["body"]
    # the body/text_body are pure HTML/text -- no reference to the binary
    # attachment mechanism leaks into the rendered content
    assert "attachment" not in rbm_draft["body"].lower()
    assert "attachment" not in rbm_draft["text_body"].lower()


# --- 7. Unflagged (Healthy) employees get no attachment --------------------

def test_healthy_employee_gets_no_attachment():
    doctors = [
        _rgd_doctor("BM01", "D1", bm_visits=0),  # flagged (0 calls)
        _rgd_doctor("BM05", "D2", bm_visits=200),  # Healthy BM -- plenty of calls
    ]
    wds.process_work_distribution_report(doctors)
    findings = {f["employee_code"]: f for f in wds.get_all_findings()}
    assert findings["BM05"]["status"] == "Healthy"

    drafts = wdns.build_notification_batch()
    rbm_draft = next(d for d in drafts if d["recipient_email"] == "rohit.rbm@example.com")
    names_in_draft = {e["employee_name"] for e in rbm_draft["employees"]}
    assert "Healthy BM" not in names_in_draft
    filenames = [fn for fn, _ in rbm_draft["attachments"]]
    assert not any("Healthy BM" in fn for fn in filenames)


# --- 8. Existing notification behavior does not regress --------------------

def test_recipient_routing_and_subject_unaffected():
    doctors = [_rgd_doctor("BM01", "D1", bm_visits=0)]
    wds.process_work_distribution_report(doctors)
    drafts = wdns.build_notification_batch()

    recipients = {d["recipient_email"] for d in drafts}
    # RGD BM chain is [ABM, RBM] -- both must still be notified
    assert recipients == {"anita.abm@example.com", "rohit.rbm@example.com"}
    for d in drafts:
        assert d["subject"] == "Saffron Automation - Work Distribution Review Required (1 Employee)"
        assert d["status"] == "Draft"
