"""Tests for the Review System's centralized schema validator
(app.review_validation.validate_review_file) and readiness gate
(app.review_upload_service.is_review_analysis_ready).

Each required-columns list is exercised at least once; the full matrix
(one test per schema x every failure mode) would be a lot of near-identical
tests for no extra coverage, since every schema goes through the exact
same _validate against a plain column list -- the failure-mode tests below
use one representative schema (opus_annual_targets) and a second
(coverage_avg_calls_onyx) to also prove independent-slot-identity, per the
spec's explicit requirement.
"""

import os

import pandas as pd
import pytest

from app.review_schemas import (
    ANNUAL_TARGETS_COLUMNS,
    AVG_AND_CALLS_COLUMNS,
    AVG_AND_CALLS_OPTIONAL_COLUMNS,
    LAST_YEAR_PRIMARY_SALES_COLUMNS,
    LAST_YEAR_PRIMARY_SALES_OPTIONAL_COLUMNS,
    PRIMARY_SALES_COLUMNS,
    REVIEW_FILE_SCHEMAS,
    REVIEW_FILE_SLOTS,
    SECONDARY_SALES_COLUMNS,
    VISITS_AND_SUPPORT_COLUMNS,
    TOTAL_REQUIRED_SLOTS,
)
from app.review_upload_service import (
    get_all_slot_states,
    is_review_analysis_ready,
    remove_review_file,
    upload_review_file,
)
from app.review_validation import _normalize_header_lenient, validate_review_file
from database.connection import get_session, init_db
from database.models import ReviewFileSlot

init_db()


@pytest.fixture(autouse=True)
def _isolated_review_storage(tmp_path, monkeypatch):
    """CRITICAL data-safety fixture. Tests below call upload_review_file()/
    remove_review_file() against the SAME 12 fixed slot_ids the real
    running app uses -- without this, a test run while real files are
    uploaded overwrites them (upload_review_file copies over the same
    slot_id-named path) and then DELETES them in cleanup (remove_review_file).
    This is not hypothetical: it happened for real on 2026-08-18, wiping a
    user's just-uploaded files/validation state.

    Redirects physical storage to a per-test temp directory, and snapshots
    + restores the real ReviewFileSlot database rows around every test --
    so no test in this file can ever touch the user's real Review System
    state, regardless of what slot_ids it uses internally.
    """
    import app.review_upload_service as review_upload_service

    # A SEPARATE subdirectory, not bare tmp_path -- tests write their own
    # source fixture files directly into tmp_path via _write_excel(), and
    # sharing one directory let _clear_stale_copies() glob-delete a
    # not-yet-copied source file out from under upload_review_file().
    storage_dir = tmp_path / "_review_uploads_storage"
    storage_dir.mkdir()
    monkeypatch.setattr(review_upload_service, "REVIEW_UPLOADS_DIR", storage_dir)

    session = get_session()
    try:
        columns = [c.name for c in ReviewFileSlot.__table__.columns if c.name != "id"]
        snapshot = [{name: getattr(row, name) for name in columns} for row in session.query(ReviewFileSlot).all()]
    finally:
        session.close()

    yield

    session = get_session()
    try:
        session.query(ReviewFileSlot).delete()
        for row_data in snapshot:
            session.add(ReviewFileSlot(**row_data))
        session.commit()
    finally:
        session.close()


def _write_excel(tmp_path, filename: str, columns: list, n_rows: int = 3):
    path = tmp_path / filename
    data = {col: [f"v{i}" for i in range(n_rows)] for col in columns}
    pd.DataFrame(data).to_excel(path, index=False)
    return str(path)


def _full_valid_columns_for_slot(slot):
    """A column list guaranteed to satisfy `slot` completely -- literal
    required_columns plus one synthesized member per month_families/
    fiscal_year_families requirement (2026-08-18 semantic-validation
    refactor: for a slot like Secondary Sales, required_columns alone is
    no longer a complete schema). Used by tests that build a fixture
    generically across every slot rather than hardcoding real column text."""
    columns = list(slot.required_columns)
    for family in slot.month_families:
        for month in family.months:
            columns.append(f"{family.prefix}{month}'26{family.suffix}")
    for family in slot.fiscal_year_families:
        columns.append(f"{family.prefix}25-26 {family.suffix_options[0]}")
    return columns


# --- Schema/slot registry sanity --------------------------------------------

def test_exactly_12_slots():
    assert TOTAL_REQUIRED_SLOTS == 12
    assert len(REVIEW_FILE_SLOTS) == 12
    assert len(REVIEW_FILE_SCHEMAS) == 12


def test_slot_ids_are_stable_internal_identifiers():
    ids = {slot.slot_id for slot in REVIEW_FILE_SLOTS}
    assert ids == {
        "opus_annual_targets",
        "opus_primary_sales",
        "opus_last_year_primary_sales",
        "opus_secondary_sales_onyx",
        "opus_secondary_sales_guardians",
        "opus_secondary_sales_xandra",
        "coverage_avg_calls_onyx",
        "coverage_avg_calls_guardians",
        "coverage_avg_calls_xandra",
        "coverage_visits_support_onyx",
        "coverage_visits_support_guardians",
        "coverage_visits_support_xandra",
    }


def test_primary_sales_and_last_year_primary_sales_are_separate_slots():
    # No longer identical schemas (2026-08-18: Last Year Primary Sales
    # dropped 4 columns as optional) -- but Primary Sales' own schema is
    # exactly Last Year's plus those 4, and they remain independent slots.
    assert "opus_primary_sales" != "opus_last_year_primary_sales"
    assert set(REVIEW_FILE_SCHEMAS["opus_last_year_primary_sales"]) == set(
        REVIEW_FILE_SCHEMAS["opus_primary_sales"]
    ) - set(LAST_YEAR_PRIMARY_SALES_OPTIONAL_COLUMNS)


# --- Valid files, one per distinct schema -----------------------------------

@pytest.mark.parametrize(
    "slot_id,columns",
    [
        ("opus_annual_targets", ANNUAL_TARGETS_COLUMNS),
        ("opus_primary_sales", PRIMARY_SALES_COLUMNS),
        ("opus_last_year_primary_sales", PRIMARY_SALES_COLUMNS),
        ("coverage_avg_calls_onyx", AVG_AND_CALLS_COLUMNS),
        ("coverage_visits_support_onyx", VISITS_AND_SUPPORT_COLUMNS),
    ],
)
def test_valid_file_for_each_distinct_schema(tmp_path, slot_id, columns):
    path = _write_excel(tmp_path, f"{slot_id}.xlsx", columns, n_rows=5)
    result = validate_review_file(slot_id, path)
    assert result["valid"] is True
    assert result["missing_columns"] == []
    assert result["unexpected_columns"] == []
    assert result["duplicate_columns"] == []
    assert result["wrong_order"] is False
    assert result["row_count"] == 5
    assert result["column_count"] == len(columns)


def test_valid_file_for_secondary_sales(tmp_path):
    # Secondary Sales' schema is no longer a plain literal column list (see
    # app/review_schemas.py's module docstring) -- covered by its own
    # dedicated section below, using _full_valid_columns_for_slot here just
    # to prove the generic synthesized-fixture helper itself is complete.
    from app.review_schemas import get_slot_def

    columns = _full_valid_columns_for_slot(get_slot_def("opus_secondary_sales_onyx"))
    path = _write_excel(tmp_path, "secondary_sales.xlsx", columns, n_rows=5)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]
    assert result["missing_columns"] == []


# --- Failure modes -----------------------------------------------------------

def test_missing_column(tmp_path):
    columns = [c for c in ANNUAL_TARGETS_COLUMNS if c != "DIFF"]
    path = _write_excel(tmp_path, "missing.xlsx", columns)
    result = validate_review_file("opus_annual_targets", path)
    assert result["valid"] is False
    assert "DIFF" in result["missing_columns"]


def test_unexpected_column_is_recognized_but_does_not_invalidate(tmp_path):
    # 2026-08-18 semantic-validation refactor: a harmless extra column is
    # not evidence this is the wrong report -- only a genuinely missing
    # required column (or a duplicate) is. unexpected_columns is still
    # reported, informationally.
    columns = list(ANNUAL_TARGETS_COLUMNS) + ["EXTRA COLUMN"]
    path = _write_excel(tmp_path, "extra.xlsx", columns)
    result = validate_review_file("opus_annual_targets", path)
    assert result["valid"] is True, result["errors"]
    assert "EXTRA COLUMN" in result["unexpected_columns"]
    assert result["missing_columns"] == []


def test_wrong_order_is_recognized_but_does_not_invalidate(tmp_path):
    # 2026-08-18 semantic-validation refactor: column order is never a
    # validity requirement -- two real sources uploading the same report
    # through different processes can legitimately order columns
    # differently. wrong_order is still reported, informationally.
    columns = list(ANNUAL_TARGETS_COLUMNS)
    columns[0], columns[1] = columns[1], columns[0]  # swap REGION / HQ
    path = _write_excel(tmp_path, "reordered.xlsx", columns)
    result = validate_review_file("opus_annual_targets", path)
    assert result["valid"] is True, result["errors"]
    assert result["wrong_order"] is True
    assert result["missing_columns"] == []
    assert result["unexpected_columns"] == []


def test_genuinely_wrong_report_still_invalidates_despite_extra_columns(tmp_path):
    # The safety rule extra/order tolerance must not undermine: a file
    # that is missing most required columns (i.e. is NOT this report) is
    # still rejected, even though it also carries unrelated extra columns.
    columns = ["Some Other Column", "Totally Unrelated"]
    path = _write_excel(tmp_path, "wrong_report.xlsx", columns)
    result = validate_review_file("opus_annual_targets", path)
    assert result["valid"] is False
    assert len(result["missing_columns"]) == len(ANNUAL_TARGETS_COLUMNS)


def test_duplicate_column(tmp_path):
    path = tmp_path / "dup.xlsx"
    columns = list(ANNUAL_TARGETS_COLUMNS) + ["REGION"]  # duplicate header on write
    df = pd.DataFrame([["v"] * len(columns) for _ in range(3)], columns=columns)
    df.to_excel(path, index=False)
    result = validate_review_file("opus_annual_targets", str(path))
    assert result["valid"] is False
    assert "REGION" in result["duplicate_columns"]


def test_empty_file(tmp_path):
    path = tmp_path / "empty.xlsx"
    pd.DataFrame().to_excel(path, index=False)
    result = validate_review_file("opus_annual_targets", str(path))
    assert result["valid"] is False
    assert result["errors"]


def test_wrong_file_type(tmp_path):
    path = tmp_path / "not_excel.txt"
    path.write_text("just some text", encoding="utf-8")
    result = validate_review_file("opus_annual_targets", str(path))
    assert result["valid"] is False


def test_whitespace_is_normalized_but_not_case_over_fragile():
    # _normalize_header (reused from app.excel_validation) lowercases and
    # collapses whitespace -- confirm the validator relies on it rather than
    # an exact-string comparison, without asserting excel_validation's own
    # behavior a second time.
    from app.excel_validation import _normalize_header
    assert _normalize_header("  Region  ") == _normalize_header("Region")
    assert _normalize_header("REGION") == _normalize_header("region")


# --- Identical schema / different slots stay independent --------------------

def test_uploading_onyx_does_not_mark_guardians_or_xandra_uploaded(tmp_path):
    from app.review_schemas import get_slot_def

    columns = _full_valid_columns_for_slot(get_slot_def("opus_secondary_sales_onyx"))
    path = _write_excel(tmp_path, "secondary_sales.xlsx", columns)
    # Baseline BEFORE uploading, not an assumed-empty state: the real
    # database this fixture snapshots/restores may legitimately already
    # have Guardians/Xandra uploaded (the user's own real data) -- the
    # actual claim under test is that uploading Onyx changes NOTHING about
    # its sibling slots, whatever their starting state.
    before_guardians = get_all_slot_states()["opus_secondary_sales_guardians"]
    before_xandra = get_all_slot_states()["opus_secondary_sales_xandra"]
    try:
        upload_review_file("opus_secondary_sales_onyx", path)
        states = get_all_slot_states()
        assert states["opus_secondary_sales_onyx"]["uploaded"] is True
        assert states["opus_secondary_sales_onyx"]["valid"] is True
        assert states["opus_secondary_sales_guardians"] == before_guardians
        assert states["opus_secondary_sales_xandra"] == before_xandra
    finally:
        remove_review_file("opus_secondary_sales_onyx")


# --- Global readiness matrix --------------------------------------------------

def test_readiness_matrix(tmp_path):
    uploaded_slots = []
    try:
        # 0/12
        for slot in REVIEW_FILE_SLOTS:
            remove_review_file(slot.slot_id)
        assert is_review_analysis_ready() is False

        # 11/12: upload valid files to all but one slot
        for slot in REVIEW_FILE_SLOTS[:-1]:
            path = _write_excel(tmp_path, f"{slot.slot_id}.xlsx", _full_valid_columns_for_slot(slot))
            upload_review_file(slot.slot_id, path)
            uploaded_slots.append(slot.slot_id)
        assert is_review_analysis_ready() is False

        # 12/12, but the last one is invalid (missing a column)
        last_slot = REVIEW_FILE_SLOTS[-1]
        bad_columns = _full_valid_columns_for_slot(last_slot)[:-1]
        bad_path = _write_excel(tmp_path, f"{last_slot.slot_id}_bad.xlsx", bad_columns)
        upload_review_file(last_slot.slot_id, bad_path)
        uploaded_slots.append(last_slot.slot_id)
        assert is_review_analysis_ready() is False

        # 12/12, all valid
        good_path = _write_excel(tmp_path, f"{last_slot.slot_id}_good.xlsx", _full_valid_columns_for_slot(last_slot))
        upload_review_file(last_slot.slot_id, good_path)
        assert is_review_analysis_ready() is True
    finally:
        for slot_id in uploaded_slots:
            remove_review_file(slot_id)


# --- Optional-column exceptions (2026-08-18, per real-file verification) ----

def test_last_year_primary_sales_accepts_file_missing_the_four_optional_columns(tmp_path):
    path = _write_excel(tmp_path, "last_year.xlsx", LAST_YEAR_PRIMARY_SALES_COLUMNS)
    result = validate_review_file("opus_last_year_primary_sales", path)
    assert result["valid"] is True
    assert result["missing_columns"] == []


def test_last_year_primary_sales_also_accepts_the_four_columns_if_present(tmp_path):
    # Optional means never checked either way -- present is fine too.
    path = _write_excel(tmp_path, "last_year_full.xlsx", PRIMARY_SALES_COLUMNS)
    result = validate_review_file("opus_last_year_primary_sales", path)
    assert result["valid"] is True


def test_primary_sales_itself_is_unaffected_by_the_last_year_exception(tmp_path):
    # Primary Sales must still require all four -- only Last Year is relaxed.
    columns = [c for c in PRIMARY_SALES_COLUMNS if c not in LAST_YEAR_PRIMARY_SALES_OPTIONAL_COLUMNS]
    path = _write_excel(tmp_path, "primary_sales_missing_four.xlsx", columns)
    result = validate_review_file("opus_primary_sales", path)
    assert result["valid"] is False
    assert set(result["missing_columns"]) == set(LAST_YEAR_PRIMARY_SALES_OPTIONAL_COLUMNS)


def test_avg_calls_accepts_file_missing_zone(tmp_path):
    path = _write_excel(tmp_path, "avg_calls_no_zone.xlsx", AVG_AND_CALLS_COLUMNS)
    result = validate_review_file("coverage_avg_calls_onyx", path)
    assert result["valid"] is True
    assert result["missing_columns"] == []


def test_avg_calls_accepts_file_with_zone_present(tmp_path):
    # Mirrors the real Guardians file, which already carries Zone.
    columns = ["Division", "Zone"] + [c for c in AVG_AND_CALLS_COLUMNS if c != "Division"]
    path = _write_excel(tmp_path, "avg_calls_with_zone.xlsx", columns)
    result = validate_review_file("coverage_avg_calls_guardians", path)
    assert result["valid"] is True
    assert result["unexpected_columns"] == []


def test_avg_calls_optional_columns_constant_is_exactly_zone():
    assert AVG_AND_CALLS_OPTIONAL_COLUMNS == ["Zone"]


# --- Lenient header matching (month/year/punctuation variations) ------------

def test_lenient_matching_accepts_harmless_month_year_format_variations(tmp_path):
    # Same logical Apr/May/Jun/Jul 2026 columns, four different real-world
    # spellings for each -- apostrophe, dot, underscore, plain space.
    varied_columns = [
        "Division", "Region", "Employee Name", "Employee Code", "Desig.",
        "Reporting HQ", "Parameters", "Apr'26", "May.2026", "Jun_2026", "Jul 2026",
    ]
    path = _write_excel(tmp_path, "avg_calls_varied_dates.xlsx", varied_columns)
    result = validate_review_file("coverage_avg_calls_onyx", path)
    assert result["valid"] is True, result["errors"]


def test_lenient_matching_still_rejects_a_genuinely_different_year(tmp_path):
    # A stale/wrong-year file (2022 instead of 2026) must NOT be accepted --
    # leniency narrows format noise, it does not widen which periods match.
    wrong_year_columns = [
        "Division", "Region", "Employee Name", "Employee Code", "Desig.",
        "Reporting HQ", "Parameters", "Apr-2022", "May-2022", "Jun-2022", "Jul-2022",
    ]
    path = _write_excel(tmp_path, "avg_calls_wrong_year.xlsx", wrong_year_columns)
    result = validate_review_file("coverage_avg_calls_onyx", path)
    assert result["valid"] is False
    assert len(result["missing_columns"]) == 4


def test_lenient_matching_does_not_collapse_prefixed_or_suffixed_columns(tmp_path):
    # "BM Visit Count Apr-2026" must stay distinct from a bare "Apr-26" --
    # only the month/year fragment canonicalizes, not the whole header.
    result = validate_review_file(
        "coverage_visits_support_onyx",
        _write_excel(tmp_path, "visits.xlsx", VISITS_AND_SUPPORT_COLUMNS),
    )
    assert result["valid"] is True
    # And a file missing the "BM Visit Count " prefix (bare month/year
    # instead) is correctly still missing that specific required column.
    swapped = list(VISITS_AND_SUPPORT_COLUMNS)
    swapped[swapped.index("BM Visit Count Apr-2026")] = "Apr-2026 (again)"
    bad_result = validate_review_file(
        "coverage_visits_support_onyx",
        _write_excel(tmp_path, "visits_bad.xlsx", swapped),
    )
    assert bad_result["valid"] is False
    assert "BM Visit Count Apr-2026" in bad_result["missing_columns"]


def test_lenient_normalization_never_collides_within_a_schema():
    """Safety check the whole leniency mechanism depends on: no two
    DISTINCT required/optional columns in any single slot's schema may
    ever normalize to the same value -- if they did, the validator could
    no longer tell them apart (a real "missing X" could silently vanish
    behind an unrelated column named Y)."""
    for slot in REVIEW_FILE_SLOTS:
        names = list(slot.required_columns) + list(slot.optional_columns)
        normalized = [_normalize_header_lenient(n) for n in names]
        assert len(set(normalized)) == len(set(names)), (
            f"Lenient normalization collision in {slot.slot_id!r}: "
            f"{len(names)} distinct column names normalized to only {len(set(normalized))} distinct values"
        )


# --- Secondary Sales semantic period-family validation (2026-08-18) ---------
# Onyx/Guardians/Xandra each maintain their own workbook independently, and
# real-world month/fiscal-year LABELS drift out of sync with each other and
# with the calendar (e.g. one real file labels its rolling Apr-Jul columns
# "Apr'25 Pri" / "May'26 Pri" / "Jun'26 Pri" / "Jul'26 Pri" -- mixed years in
# the same row). These synthesize fixtures shaped like the real files rather
# than depending on review_uploads/ being present (see app/review_schemas.py's
# module docstring for the full mechanism).

def _secondary_sales_columns(month_years=("25", "25", "25", "25"), fy_range="25-26", fy_sup_word="Sup", closing_suffix=""):
    # The monthly Sup column is always literally "Sup" in every real file
    # observed (Onyx/Guardians/Xandra alike) -- only the F.Y-total's third
    # member varies ("Sup" vs "Support"), hence `fy_sup_word` applies there
    # only, not to the monthly triple.
    months = ["Apr", "May", "Jun", "Jul"]
    columns = ["Region", "HQ", "ABM Name", "No Of BM"]
    for month, year in zip(months, month_years):
        columns += [f"{month}'{year} Pri", f"{month}'{year} Sec", f"{month}'{year} Sup"]
    columns += [f"F.Y {fy_range} Prim", f"F.Y {fy_range} Sec", f"F.Y {fy_range} {fy_sup_word}"]
    for month, year in zip(months, month_years):
        columns.append(f"{month}'{year}{closing_suffix}")
    return columns


def test_secondary_sales_accepts_consistent_year_labels(tmp_path):
    # Shaped like the real Xandra file: every column labeled '26.
    columns = _secondary_sales_columns(month_years=("26", "26", "26", "26"), fy_range="26-27")
    path = _write_excel(tmp_path, "xandra_shaped.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_xandra", path)
    assert result["valid"] is True, result["errors"]


def test_secondary_sales_accepts_mixed_year_labels_within_one_file(tmp_path):
    # Shaped like the real Guardians file: Apr'25 alongside May/Jun/Jul'26
    # in the same row -- the YEAR is never checked, only the calendar month.
    columns = _secondary_sales_columns(month_years=("25", "26", "26", "26"), fy_range="25-26")
    path = _write_excel(tmp_path, "guardians_shaped.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_guardians", path)
    assert result["valid"] is True, result["errors"]


def test_secondary_sales_accepts_closing_suffix_variant(tmp_path):
    # Real Onyx/Xandra files append " Closing" to the trailing 4 columns;
    # Guardians doesn't -- both must validate (allow_trailing_closing).
    columns = _secondary_sales_columns(closing_suffix=" Closing")
    path = _write_excel(tmp_path, "onyx_shaped.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]


def test_secondary_sales_accepts_sup_vs_support_spelling(tmp_path):
    # The real Onyx/Xandra F.Y-total column spells it "Support"; Guardians
    # spells it "Sup" -- FiscalYearFamily.suffix_options recognizes both.
    columns = _secondary_sales_columns(fy_sup_word="Support")
    path = _write_excel(tmp_path, "support_spelling.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]


def test_secondary_sales_fy_range_26_27_accepted_same_as_25_26(tmp_path):
    # The specific fiscal-year range is a wildcard -- whichever period the
    # file is actually reporting for.
    columns = _secondary_sales_columns(fy_range="26-27")
    path = _write_excel(tmp_path, "fy_26_27.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]


def test_secondary_sales_tolerates_extra_earlier_months(tmp_path):
    # Additional legitimate history (e.g. Feb) beyond the required Apr-Jul
    # is MORE DATA, not stale data -- must not be flagged unexpected.
    columns = _secondary_sales_columns()
    columns[4:4] = ["Feb'25 Pri", "Feb'25 Sec", "Feb'25 Sup"]
    path = _write_excel(tmp_path, "extra_months.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]
    assert result["missing_columns"] == []
    assert not any(c.startswith("Feb") for c in result["unexpected_columns"])


def test_secondary_sales_rejects_when_a_required_month_is_truly_absent(tmp_path):
    columns = [c for c in _secondary_sales_columns() if not c.startswith("Jul'")]
    path = _write_excel(tmp_path, "missing_jul.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is False
    assert any("Jul" in m for m in result["missing_columns"])


def test_secondary_sales_reordered_columns_still_valid(tmp_path):
    columns = _secondary_sales_columns()
    columns[0], columns[1] = columns[1], columns[0]  # swap Region/HQ
    path = _write_excel(tmp_path, "reordered.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]
    assert result["wrong_order"] is True  # still reported, informationally


def test_secondary_sales_extra_unrelated_columns_still_valid(tmp_path):
    # Shaped like the real Xandra file's trailing QTR/percentage columns.
    columns = _secondary_sales_columns() + ["Division", "QTR-1 PRI", "QTR-1 SEC", "%"]
    path = _write_excel(tmp_path, "with_extras.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is True, result["errors"]
    assert set(result["unexpected_columns"]) == {"Division", "QTR-1 PRI", "QTR-1 SEC", "%"}


def test_secondary_sales_genuinely_missing_fy_total_still_invalidates(tmp_path):
    # Safety rule: tolerance for extras/order/year-drift must not extend to
    # a genuinely absent required field.
    columns = [c for c in _secondary_sales_columns() if not c.startswith("F.Y")]
    path = _write_excel(tmp_path, "no_fy_total.xlsx", columns)
    result = validate_review_file("opus_secondary_sales_onyx", path)
    assert result["valid"] is False
    assert any("F.Y" in m for m in result["missing_columns"])


def test_month_family_matches_excel_datetime_header(tmp_path):
    """A bare closing-family header shaped like a date (e.g. "Apr'26") may
    be read back by Excel/pandas as an actual datetime object rather than
    the text label a user sees -- must still be recognized as that month."""
    import openpyxl
    from datetime import datetime as dt

    columns = _secondary_sales_columns()
    trailing_start = len(columns) - 4  # the 4 bare closing-family columns
    month_nums = {"Apr": 4, "May": 5, "Jun": 6, "Jul": 7}

    path = tmp_path / "datetime_header.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, col in enumerate(columns):
        if i >= trailing_start:
            ws.cell(row=1, column=i + 1, value=dt(2026, month_nums[col.split("'")[0]], 1))
        else:
            ws.cell(row=1, column=i + 1, value=col)
    for r in range(2, 5):
        for c in range(1, len(columns) + 1):
            ws.cell(row=r, column=c, value=f"v{r}")
    wb.save(path)

    result = validate_review_file("opus_secondary_sales_onyx", str(path))
    assert result["valid"] is True, result["errors"]


# --- Real uploaded files (verification-run source of truth, 2026-08-18) -----
# Skips cleanly if the local review_uploads/ directory isn't present (e.g. a
# clean checkout, or CI) -- these tests exercise whatever the user's own
# machine currently has uploaded, not a fixture shipped with the repo.

_REAL_UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "review_uploads")


def _real_file_states():
    if not os.path.isdir(_REAL_UPLOADS_DIR):
        return {}
    states = {}
    for slot in REVIEW_FILE_SLOTS:
        for ext in (".xlsx", ".xls", ".xlsm", ".csv"):
            candidate = os.path.join(_REAL_UPLOADS_DIR, f"{slot.slot_id}{ext}")
            if os.path.isfile(candidate):
                states[slot.slot_id] = candidate
                break
    return states


@pytest.mark.skipif(not _real_file_states(), reason="no real review_uploads/ files present on this machine")
def test_real_last_year_primary_sales_file_now_valid():
    real = _real_file_states()
    if "opus_last_year_primary_sales" not in real:
        pytest.skip("opus_last_year_primary_sales not currently uploaded")
    result = validate_review_file("opus_last_year_primary_sales", real["opus_last_year_primary_sales"])
    assert result["valid"] is True, result["errors"]


@pytest.mark.skipif(not _real_file_states(), reason="no real review_uploads/ files present on this machine")
def test_real_avg_calls_files_now_valid():
    real = _real_file_states()
    for slot_id in ("coverage_avg_calls_onyx", "coverage_avg_calls_guardians", "coverage_avg_calls_xandra"):
        if slot_id not in real:
            continue
        result = validate_review_file(slot_id, real[slot_id])
        assert result["valid"] is True, f"{slot_id}: {result['errors']}"


@pytest.mark.skipif(not _real_file_states(), reason="no real review_uploads/ files present on this machine")
def test_real_every_currently_uploaded_slot_is_valid():
    """2026-08-18 semantic-validation refactor regression: every real file
    currently in review_uploads/ must validate -- whichever subset of the
    12 slots happens to be uploaded on this machine (annual targets,
    primary sales, all 3 secondary sales sources, all 3 avg & calls, all
    3 visits & support -- see the user's exact regression checklist)."""
    real = _real_file_states()
    for slot_id, path in real.items():
        result = validate_review_file(slot_id, path)
        assert result["valid"] is True, f"{slot_id}: {result['errors']}"
