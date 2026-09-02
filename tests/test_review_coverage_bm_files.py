"""Tests for Coverage Summary's per-BM file generation
(app.review_coverage_service.generate_coverage_summary_bm_files) -- the
"one BM = one Coverage Summary file" step of the new automated-email
workflow (see app.review_coverage_notification_service). Synthetic
fixtures only (no real review_uploads/ files needed), monkeypatching
_load_avg_calls/_load_visits_support the same way
test_roster_falls_back_to_everything_when_no_hq_distribution_file (in
test_review_coverage_service.py) monkeypatches get_valid_hqs_for_division
-- both are the module's own private loaders, safe to swap out for a
fixed DataFrame.
"""

import openpyxl
import pandas as pd
import pytest

import app.review_coverage_service as coverage_service
from app.review_coverage_service import ROW_LABELS, generate_coverage_summary_bm_files


def _avg_df_for(bms: list[tuple[str, str]], division="Xandra", hq="Guntur") -> pd.DataFrame:
    """`bms`: list of (emp_code, name). One 3-row block per BM (Field
    Visit Days / # of Doctor Visits / Doctor Call Average), mirroring
    test_review_coverage_service.py's own _avg_df shape but varying name
    per BM (that file's helper hardcodes one name for all rows)."""
    rows = []
    for code, name in bms:
        for param, value in (("Field Visit Days", 20.0), ("# of Doctor Visits", 150.0), ("Doctor Call Average", 7.5)):
            rows.append({
                "Division": division, "Region": "R", "Employee Name": name, "Employee Code": code,
                "Desig.": "BM", "Reporting HQ": hq, "Parameters": param, "Apr-2026": value,
            })
    return pd.DataFrame(rows)


_APR = pd.Timestamp("2026-04-01")
_VISIT_COL = "BM Visit Count Apr-2026"


def _empty_vs_df() -> pd.DataFrame:
    return pd.DataFrame(columns=["BM code", "Dr. Code", "Category", _APR, _VISIT_COL])


def _apply_fixture(monkeypatch, bms: list[tuple[str, str]], division="Xandra", tmp_path=None):
    monkeypatch.setattr(coverage_service, "_load_avg_calls", lambda div: _avg_df_for(bms, division=div))
    monkeypatch.setattr(coverage_service, "_load_visits_support", lambda div: _empty_vs_df())
    monkeypatch.setattr(coverage_service, "coverage_prerequisites_ready", lambda div: (True, []))
    monkeypatch.setattr(coverage_service, "get_valid_hqs_for_division", lambda div: None)
    if tmp_path is not None:
        monkeypatch.setattr(coverage_service, "REVIEW_UPLOADS_DIR", tmp_path)


def _read_block_identity(xlsx_path) -> tuple:
    """(emp_code, name) pairs actually present in the workbook's data rows
    -- reads back what was written, not what we think we wrote."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    codes_and_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        # headers: Division, Region Name, HQ, Emp Code, Name, Designation, No, Parameters, APR, MAY, JUN
        codes_and_names.add((row[3], row[4]))
    return codes_and_names


def test_bm_wise_file_isolation(monkeypatch, tmp_path):
    """Every generated attachment contains EXACTLY one BM's data -- no
    other BM's rows leak into it."""
    bms = [("E1", "Alice BM"), ("E2", "Bob BM"), ("E3", "Carol BM")]
    _apply_fixture(monkeypatch, bms, tmp_path=tmp_path)

    result = generate_coverage_summary_bm_files("Xandra")
    assert result["success"] is True, result["errors"]
    assert len(result["files"]) == 3

    for f in result["files"]:
        identity = _read_block_identity(f["file_path"])
        assert identity == {(f["emp_code"], f["name"])}
        # Exactly ROW_LABELS worth of data rows (one BM's own 9-row block).
        wb = openpyxl.load_workbook(f["file_path"])
        data_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == len(ROW_LABELS)


def test_filename_format(monkeypatch, tmp_path):
    bms = [("E1", "Kayakakula Mallikarjuna")]
    _apply_fixture(monkeypatch, bms, tmp_path=tmp_path)

    result = generate_coverage_summary_bm_files("Xandra")
    assert result["success"] is True
    [f] = result["files"]
    from pathlib import Path
    assert Path(f["file_path"]).name == "Coverage Summary - Kayakakula Mallikarjuna.xlsx"


def test_invalid_filename_characters_stripped(monkeypatch, tmp_path):
    bms = [("E1", 'Bad/Name:For<>Files?"*')]
    _apply_fixture(monkeypatch, bms, tmp_path=tmp_path)

    result = generate_coverage_summary_bm_files("Xandra")
    assert result["success"] is True, result["errors"]
    [f] = result["files"]
    from pathlib import Path
    filename = Path(f["file_path"]).name
    for bad_char in '/:<>?"*':
        assert bad_char not in filename
    assert filename.startswith("Coverage Summary - ")
    assert filename.endswith(".xlsx")


def test_duplicate_display_name_disambiguated_not_overwritten(monkeypatch, tmp_path):
    """Two different Employee Codes sharing the same display Name must
    produce TWO distinct files, never one silently overwriting the
    other's data."""
    bms = [("E1", "Same Name"), ("E2", "Same Name")]
    _apply_fixture(monkeypatch, bms, tmp_path=tmp_path)

    result = generate_coverage_summary_bm_files("Xandra")
    assert result["success"] is True, result["errors"]
    assert len(result["files"]) == 2
    file_paths = {f["file_path"] for f in result["files"]}
    assert len(file_paths) == 2  # genuinely two different files on disk
    for f in result["files"]:
        identity = _read_block_identity(f["file_path"])
        assert identity == {(f["emp_code"], f["name"])}


def test_full_replace_removes_stale_files_from_previous_run(monkeypatch, tmp_path):
    _apply_fixture(monkeypatch, [("E1", "Alice BM"), ("E2", "Bob BM")], tmp_path=tmp_path)
    first = generate_coverage_summary_bm_files("Xandra")
    assert len(first["files"]) == 2

    # Second run: roster shrinks to one BM -- Bob's stale file must be gone.
    _apply_fixture(monkeypatch, [("E1", "Alice BM")], tmp_path=tmp_path)
    second = generate_coverage_summary_bm_files("Xandra")
    assert len(second["files"]) == 1

    from app.review_coverage_service import _bm_files_output_dir
    remaining = list(_bm_files_output_dir("Xandra").glob("*.xlsx"))
    assert len(remaining) == 1


def test_bm_with_no_data_never_appears(monkeypatch, tmp_path):
    """A BM with zero Coverage Summary data (absent from the Avg & Calls
    roster entirely) simply never appears in the roster -- nothing to
    generate, nothing to email, no error."""
    _apply_fixture(monkeypatch, [("E1", "Alice BM")], tmp_path=tmp_path)
    result = generate_coverage_summary_bm_files("Xandra")
    assert {f["emp_code"] for f in result["files"]} == {"E1"}


def test_unknown_division_reports_error_not_exception(tmp_path):
    result = generate_coverage_summary_bm_files("Nonexistent")
    assert result["success"] is False
    assert result["files"] == []
    assert result["errors"]


def test_prerequisites_not_ready_reports_error(monkeypatch):
    monkeypatch.setattr(coverage_service, "coverage_prerequisites_ready", lambda div: (False, ["some_slot"]))
    result = generate_coverage_summary_bm_files("Xandra")
    assert result["success"] is False
    assert result["files"] == []
    assert "some_slot" in result["errors"][0]


def test_combined_workbook_unaffected_by_per_bm_split(monkeypatch, tmp_path):
    """Regression: generate_coverage_summary()'s own combined-workbook
    output is untouched by the new per-BM function existing -- same
    bm_count, same single file_path, both generated independently."""
    from app.review_coverage_service import generate_coverage_summary

    bms = [("E1", "Alice BM"), ("E2", "Bob BM")]
    _apply_fixture(monkeypatch, bms, tmp_path=tmp_path)

    combined = generate_coverage_summary("Xandra")
    assert combined["success"] is True
    assert combined["bm_count"] == 2
    wb = openpyxl.load_workbook(combined["file_path"])
    data_rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2 * len(ROW_LABELS)  # both BMs in ONE sheet, as before
