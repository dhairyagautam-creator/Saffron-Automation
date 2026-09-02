"""Tests for RGD Visit and Support (app.review_rgd_service) -- the
deliberately simplest report in Review System: a filtered, unmodified
replica of the raw Visits & Support upload. Covers the spec's own
acceptance criteria directly (2026-08-19): no calculation, exact 20-column
order, raw value preservation (blanks stay blank, "-" stays "-", no
rounding), Category != "General" filtering only.
"""

import openpyxl
import pandas as pd
import pytest

from app.review_rgd_service import (
    DIVISIONS,
    OUTPUT_HEADERS,
    _format_cell,
    _load_rgd_rows,
    _raw,
    _support_columns_by_month,
    _visit_columns_by_month,
    generate_rgd_summary,
)

# --- Column contract -----------------------------------------------------

def test_output_headers_are_the_exact_20_columns_in_spec_order():
    assert OUTPUT_HEADERS == (
        "Region", "HQ", "BM Code", "BM Name", "Dr Code", "Dr Name", "Town", "B-RGD/A-RGD", "Speciality",
        "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26", "Jul-26",
        "Apr", "May", "Jun", "Jul", "Aug",
    )
    assert len(OUTPUT_HEADERS) == 20


# --- Raw value preservation ------------------------------------------------

def test_raw_preserves_numbers_exactly_no_rounding():
    assert _raw(12118.26) == 12118.26


def test_raw_preserves_literal_dash_string():
    assert _raw("-") == "-"


def test_raw_blank_cell_becomes_empty_string_not_zero():
    assert _raw(None) == ""
    assert _raw(float("nan")) == ""
    assert _raw(None) != 0
    assert _raw(float("nan")) != 0


def test_raw_never_alters_a_code_string():
    assert _raw("'00068196") == "'00068196"


def test_format_cell_does_not_round():
    assert _format_cell(12118.26) == "12118.26"


def test_format_cell_strips_trailing_dot_zero_from_whole_number_floats_only():
    # Cosmetic preview-only normalization (pandas upcasts a column with any
    # blank to float64) -- the underlying raw value is untouched elsewhere.
    assert _format_cell(4.0) == "4"
    assert _format_cell(4.5) == "4.5"


def test_format_cell_preserves_literal_dash():
    assert _format_cell("-") == "-"


def test_format_cell_blank_stays_blank():
    assert _format_cell("") == ""


# --- Month column lookup (per-division variability) ------------------------

def test_support_columns_located_by_month_number_not_fixed_position():
    df = pd.DataFrame({
        "X": [1],
        pd.Timestamp("2026-04-01"): [100],
        pd.Timestamp("2026-02-01"): [200],
    })
    result = _support_columns_by_month(df)
    assert result[4] == pd.Timestamp("2026-04-01")
    assert result[2] == pd.Timestamp("2026-02-01")
    assert 7 not in result  # July genuinely absent -- never fabricated


def test_visit_columns_located_by_name_per_month():
    df = pd.DataFrame({"BM Visit Count Apr-2026": [1], "BM Visit Count Jul-2026": [2], "Other": [3]})
    result = _visit_columns_by_month(df)
    assert result[4] == "BM Visit Count Apr-2026"
    assert result[7] == "BM Visit Count Jul-2026"
    assert 8 not in result  # August genuinely absent -- never fabricated
    assert 5 not in result


# --- Filtering: Category != "General" only, nothing else -------------------

def _fake_visits_support_df():
    return pd.DataFrame([
        {"Division": "Xandra", "Region": "R1", "HQ": "H1", "BM code": "B1", "BM Name": "BM One",
         "Dr. Code": "D1", "Dr. Name": "Doc One", "Town": "T1", "Category": "General", "Speciality": "GP",
         pd.Timestamp("2026-04-01"): 100.0, "BM Visit Count Apr-2026": 2},
        {"Division": "Xandra", "Region": "R1", "HQ": "H1", "BM code": "B1", "BM Name": "BM One",
         "Dr. Code": "D2", "Dr. Name": "Doc Two", "Town": "T1", "Category": "B-RGD", "Speciality": "ORTHO",
         pd.Timestamp("2026-04-01"): 200.0, "BM Visit Count Apr-2026": "-"},
        {"Division": "Xandra", "Region": "R2", "HQ": "H2", "BM code": "B2", "BM Name": "BM Two",
         "Dr. Code": "D3", "Dr. Name": "Doc Three", "Town": "T2", "Category": "B-RGD/A-RGD", "Speciality": "CP",
         pd.Timestamp("2026-04-01"): None, "BM Visit Count Apr-2026": None},
    ])


def test_load_rgd_rows_excludes_general_category(monkeypatch):
    import app.review_rgd_service as rgd_service
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: _fake_visits_support_df())

    rows = _load_rgd_rows("Xandra")

    assert len(rows) == 2  # D1 (General) excluded, D2 and D3 kept
    assert all(row["dr_code"] != "D1" for row in rows)
    assert {row["dr_code"] for row in rows} == {"D2", "D3"}


def test_load_rgd_rows_preserves_category_value_as_brgd_argd_column(monkeypatch):
    import app.review_rgd_service as rgd_service
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: _fake_visits_support_df())

    rows = _load_rgd_rows("Xandra")

    by_dr = {row["dr_code"]: row for row in rows}
    assert by_dr["D2"]["category"] == "B-RGD"
    assert by_dr["D3"]["category"] == "B-RGD/A-RGD"


def test_load_rgd_rows_preserves_dash_and_blank_and_number(monkeypatch):
    import app.review_rgd_service as rgd_service
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: _fake_visits_support_df())

    rows = _load_rgd_rows("Xandra")
    by_dr = {row["dr_code"]: row for row in rows}

    assert by_dr["D2"]["support_apr"] == 200.0  # exact number, not rounded
    assert by_dr["D2"]["visit_apr"] == "-"       # literal dash preserved
    assert by_dr["D3"]["support_apr"] == ""      # blank stays blank, not 0
    assert by_dr["D3"]["visit_apr"] == ""


def test_load_rgd_rows_leaves_nonexistent_months_blank_not_fabricated(monkeypatch):
    """The spec asks for Jul-26 support and Aug visits, which don't exist
    in this fake (or any real) source file -- must be blank, never a
    fabricated 0 or invented value."""
    import app.review_rgd_service as rgd_service
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: _fake_visits_support_df())

    rows = _load_rgd_rows("Xandra")
    for row in rows:
        assert row["support_jul"] == ""
        assert row["visit_aug"] == ""


def test_no_bm_code_column_filter_needed_every_row_already_owns_one(monkeypatch):
    """Confirms the whole-division dataset naturally satisfies "BM Code =
    selected BM Code" for every row simultaneously -- rows from different
    BMs (B1 and B2) coexist correctly, each keeping its own BM Code."""
    import app.review_rgd_service as rgd_service
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: _fake_visits_support_df())

    rows = _load_rgd_rows("Xandra")
    bm_codes = {row["bm_code"] for row in rows}
    assert bm_codes == {"B1", "B2"}


# --- No derivation: row count in equals row count out (minus General) ------

def test_no_aggregation_row_count_matches_filtered_source_row_count(monkeypatch):
    import app.review_rgd_service as rgd_service
    df = _fake_visits_support_df()
    monkeypatch.setattr(rgd_service, "_load_visits_support", lambda division: df)

    rows = _load_rgd_rows("Xandra")
    expected = len(df[df["Category"] != "General"])
    assert len(rows) == expected  # one output row per non-General source row, no grouping/summing


# --- Real-file validation (skips cleanly if review_uploads/ isn't present) --

import os

_REAL_UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "review_uploads")
_REQUIRES_REAL_FILES = pytest.mark.skipif(
    not os.path.isdir(_REAL_UPLOADS_DIR), reason="no real review_uploads/ files present on this machine"
)


@_REQUIRES_REAL_FILES
@pytest.mark.parametrize("division", DIVISIONS)
def test_generate_rgd_summary_end_to_end(division):
    from app.review_rgd_service import rgd_prerequisites_ready

    ready, missing = rgd_prerequisites_ready(division)
    if not ready:
        pytest.skip(f"{division} source files not fully uploaded/valid: {missing}")

    result = generate_rgd_summary(division)
    assert result["success"] is True, result["errors"]
    assert result["row_count"] > 0
    assert os.path.isfile(result["file_path"])

    wb = openpyxl.load_workbook(result["file_path"])
    ws = wb.active
    assert ws.title == "RGD VISIT AND SUPPORT"
    assert ws.max_column == 20
    headers = [ws.cell(row=1, column=c).value for c in range(1, 21)]
    assert tuple(headers) == OUTPUT_HEADERS
    assert ws.max_row == result["row_count"] + 1  # +1 header row -- no derived summary rows added


@_REQUIRES_REAL_FILES
def test_no_general_category_rows_in_generated_output():
    from app.review_rgd_service import rgd_prerequisites_ready

    ready, missing = rgd_prerequisites_ready("Xandra")
    if not ready:
        pytest.skip(f"Xandra source files not fully uploaded/valid: {missing}")

    rows = _load_rgd_rows("Xandra")
    assert all(row["category"] != "General" for row in rows)
    assert all(row["category"] in ("B-RGD", "B-RGD/A-RGD") for row in rows)
