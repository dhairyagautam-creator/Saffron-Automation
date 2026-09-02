"""Tests for app.review_export_service ("Export Entire File") --
synthetic fixtures only, no real review_uploads/ dependency, per this
project's established convention of never generating a fake "final
review file" just to demonstrate a feature (see
tests/test_hq_distribution_service.py's own docstring).
"""

import openpyxl
import pytest
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.review_export_service import EXPORT_REPORTS, ExportReport, _copy_worksheet, export_all_divisions


# --- _copy_worksheet: pure formatting-preservation ---------------------------

def _make_source_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SRC"
    ws["A1"] = "Hello"
    ws["A1"].font = Font(name="Calibri", size=10, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="D9E1F2")
    ws["A1"].border = Border(left=Side(style="thin"), right=Side(style="thin"))
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].number_format = "0.00"
    ws["B1"] = 5
    ws["C1"] = "=B1*2"
    ws.column_dimensions["A"].width = 25
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    return ws


def test_copy_worksheet_preserves_values_and_formulas():
    src_ws = _make_source_sheet()
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    dst_ws = dst_wb.create_sheet("DST")

    _copy_worksheet(src_ws, dst_ws)

    assert dst_ws["A1"].value == "Hello"
    assert dst_ws["B1"].value == 5
    assert dst_ws["C1"].value == "=B1*2"


def test_copy_worksheet_preserves_formatting():
    src_ws = _make_source_sheet()
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    dst_ws = dst_wb.create_sheet("DST")

    _copy_worksheet(src_ws, dst_ws)

    assert dst_ws["A1"].font.bold is True
    assert dst_ws["A1"].font.name == "Calibri"
    assert dst_ws["A1"].fill.fgColor.rgb == "00D9E1F2"
    assert dst_ws["A1"].border.left.style == "thin"
    assert dst_ws["A1"].alignment.horizontal == "center"
    assert dst_ws["A1"].number_format == "0.00"


def test_copy_worksheet_preserves_dimensions_and_freeze_panes():
    src_ws = _make_source_sheet()
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    dst_ws = dst_wb.create_sheet("DST")

    _copy_worksheet(src_ws, dst_ws)

    assert dst_ws.column_dimensions["A"].width == 25
    assert dst_ws.row_dimensions[1].height == 30
    assert dst_ws.freeze_panes == "A2"


def test_copy_worksheet_style_memoization_never_merges_different_styles():
    """Regression guard for the cell._style-keyed memoization added
    2026-08-20 (performance fix): two cells with genuinely DIFFERENT
    formatting must never end up sharing a cached, incorrectly-merged
    style just because they were processed close together."""
    src_wb = openpyxl.Workbook()
    src_ws = src_wb.active
    src_ws["A1"] = "bold-red"
    src_ws["A1"].font = Font(bold=True, color="FF0000")
    src_ws["A2"] = "plain-blue"
    src_ws["A2"].font = Font(bold=False, color="0000FF")
    src_ws["A3"] = "bold-red-again"
    src_ws["A3"].font = Font(bold=True, color="FF0000")

    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    dst_ws = dst_wb.create_sheet("DST")
    _copy_worksheet(src_ws, dst_ws)

    assert dst_ws["A1"].font.bold is True and dst_ws["A1"].font.color.rgb == "00FF0000"
    assert dst_ws["A2"].font.bold is False and dst_ws["A2"].font.color.rgb == "000000FF"
    assert dst_ws["A3"].font.bold is True and dst_ws["A3"].font.color.rgb == "00FF0000"


def test_copy_worksheet_survives_a_real_save_and_reload(tmp_path):
    # Formula strings and styles must round-trip through an actual save --
    # not just live correctly in-memory before the workbook is written.
    src_ws = _make_source_sheet()
    dst_wb = openpyxl.Workbook()
    dst_wb.remove(dst_wb.active)
    dst_ws = dst_wb.create_sheet("DST")
    _copy_worksheet(src_ws, dst_ws)

    out_path = tmp_path / "combined.xlsx"
    dst_wb.save(out_path)

    reopened = openpyxl.load_workbook(out_path)
    assert reopened.sheetnames == ["DST"]
    assert reopened["DST"]["C1"].value == "=B1*2"
    assert reopened["DST"]["A1"].font.bold is True


# --- export_all_divisions: registry-driven, division-isolated ---------------

def _write_tiny_workbook(path, cell_text) -> None:
    wb = openpyxl.Workbook()
    wb.active["A1"] = cell_text
    wb.save(path)


def _fake_report(name, tmp_path, *, pre_existing=(), division_errors=None):
    """A stand-in ExportReport -- path_fn checks a fixed per-division path;
    generate_fn (only called when that path doesn't already exist) writes
    one real, tiny .xlsx so _copy_worksheet has something real to copy.
    `pre_existing`: divisions whose file is written up front, so path_fn
    finds it immediately and generate_fn must never be called for them."""
    calls = []

    def path_fn(division):
        return tmp_path / f"{division}_{name}.xlsx"

    for division in pre_existing:
        _write_tiny_workbook(path_fn(division), f"{division}-{name}-PREEXISTING")

    def generate_fn(division, report_progress=None):
        calls.append(division)
        if division_errors and division in division_errors:
            return {"success": False, "file_path": None, "errors": [division_errors[division]]}
        path = path_fn(division)
        _write_tiny_workbook(path, f"{division}-{name}-GENERATED")
        return {"success": True, "file_path": str(path), "errors": []}

    generate_fn.calls = calls
    return ExportReport(name=name, path_fn=path_fn, generate_fn=generate_fn)


def test_export_all_divisions_isolates_each_divisions_data(monkeypatch, tmp_path):
    import app.review_export_service as export_service

    monkeypatch.setattr(export_service, "_generated_output_dir", lambda: tmp_path)
    opus_report = _fake_report("OPUS SUMMARY", tmp_path)
    coverage_report = _fake_report("COVERAGE SUMMARY", tmp_path)
    monkeypatch.setattr(export_service, "EXPORT_REPORTS", (opus_report, coverage_report))
    monkeypatch.setattr(export_service, "DIVISIONS", ("Xandra", "Onyx"))

    result = export_all_divisions()

    assert result["success"] is True
    assert set(result["files"]) == {"Xandra", "Onyx"}
    for division, path in result["files"].items():
        assert path is not None
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["OPUS SUMMARY", "COVERAGE SUMMARY"]
        assert wb["OPUS SUMMARY"]["A1"].value == f"{division}-OPUS SUMMARY-GENERATED"
        assert wb["COVERAGE SUMMARY"]["A1"].value == f"{division}-COVERAGE SUMMARY-GENERATED"


def test_export_all_divisions_reuses_existing_file_without_regenerating(monkeypatch, tmp_path):
    """The exact behavior the user hit (2026-08-19): if a division's
    report is already generated on disk, Export Entire File must copy it
    directly -- never call generate_fn again just to produce the same
    file. This is the whole point of preferring path_fn over always
    regenerating (see this module's docstring)."""
    import app.review_export_service as export_service

    monkeypatch.setattr(export_service, "_generated_output_dir", lambda: tmp_path)
    opus_report = _fake_report("OPUS SUMMARY", tmp_path, pre_existing=("Xandra",))
    coverage_report = _fake_report("COVERAGE SUMMARY", tmp_path, pre_existing=("Xandra",))
    monkeypatch.setattr(export_service, "EXPORT_REPORTS", (opus_report, coverage_report))
    monkeypatch.setattr(export_service, "DIVISIONS", ("Xandra",))

    result = export_all_divisions()

    assert opus_report.generate_fn.calls == []  # never called -- file already existed
    assert coverage_report.generate_fn.calls == []
    wb = openpyxl.load_workbook(result["files"]["Xandra"])
    assert wb["OPUS SUMMARY"]["A1"].value == "Xandra-OPUS SUMMARY-PREEXISTING"
    assert wb["COVERAGE SUMMARY"]["A1"].value == "Xandra-COVERAGE SUMMARY-PREEXISTING"


def test_export_all_divisions_generates_only_missing_divisions(monkeypatch, tmp_path):
    import app.review_export_service as export_service

    monkeypatch.setattr(export_service, "_generated_output_dir", lambda: tmp_path)
    opus_report = _fake_report("OPUS SUMMARY", tmp_path, pre_existing=("Xandra",))
    monkeypatch.setattr(export_service, "EXPORT_REPORTS", (opus_report,))
    monkeypatch.setattr(export_service, "DIVISIONS", ("Xandra", "Onyx"))

    export_all_divisions()

    assert opus_report.generate_fn.calls == ["Onyx"]  # Xandra already existed, Onyx didn't


def test_export_all_divisions_omits_failed_report_sheet_not_a_fake_one(monkeypatch, tmp_path):
    import app.review_export_service as export_service

    monkeypatch.setattr(export_service, "_generated_output_dir", lambda: tmp_path)
    opus_report = _fake_report("OPUS SUMMARY", tmp_path)
    coverage_report = _fake_report("COVERAGE SUMMARY", tmp_path, division_errors={"Guardians": "source files not uploaded"})
    monkeypatch.setattr(export_service, "EXPORT_REPORTS", (opus_report, coverage_report))
    monkeypatch.setattr(export_service, "DIVISIONS", ("Guardians",))

    result = export_all_divisions()

    wb = openpyxl.load_workbook(result["files"]["Guardians"])
    assert wb.sheetnames == ["OPUS SUMMARY"]  # Coverage Summary omitted, not faked
    assert result["errors"]["Guardians"] == ["COVERAGE SUMMARY: source files not uploaded"]


def test_export_all_divisions_skips_division_entirely_when_every_report_fails(monkeypatch, tmp_path):
    import app.review_export_service as export_service

    monkeypatch.setattr(export_service, "_generated_output_dir", lambda: tmp_path)
    opus_report = _fake_report("OPUS SUMMARY", tmp_path, division_errors={"Guardians": "no mapping"})
    coverage_report = _fake_report("COVERAGE SUMMARY", tmp_path, division_errors={"Guardians": "no source files"})
    monkeypatch.setattr(export_service, "EXPORT_REPORTS", (opus_report, coverage_report))
    monkeypatch.setattr(export_service, "DIVISIONS", ("Guardians",))

    result = export_all_divisions()

    assert result["success"] is False
    assert result["files"]["Guardians"] is None


def test_export_reports_registry_has_all_three_current_reports():
    # Pins the current registry contents -- RGD Visit and Support
    # (2026-08-19) proved this registry's extensibility: appending one
    # ExportReport here was the entire change. A future fourth report
    # follows the same pattern.
    assert [r.name for r in EXPORT_REPORTS] == ["OPUS SUMMARY", "COVERAGE SUMMARY", "RGD VISIT AND SUPPORT"]
