"""Coverage Summary generation -- the BM-wise coverage report built from
the Review System's Avg. & Calls and Visits & Support uploads, for one
division (Xandra/Onyx/Guardians) at a time. Architecture deliberately
mirrors app/review_opus_service.py (same division-scoping, same HQ
Distribution applicability filter, same literal-value-vs-formula split,
same JSON preview sidecar) -- see that module for the shared reasoning;
this docstring only covers what's specific to Coverage Summary.

Source files (division-specific uploads, like Secondary Sales -- the
Division boundary is the FILE itself, never a column filter on a shared
file):
  - Avg. & Calls (coverage_avg_calls_{division}) -- one row per
    (BM, Parameter), Parameters in {"Field Visit Days",
    "# of Doctor Visits", "Doctor Call Average"}. Source of the BM roster
    itself (every distinct Employee Code is one BM block) and rows 1-3.
  - Visits & Support (coverage_visits_support_{division}) -- one row per
    (BM, Doctor). Source of rows 4-9. Its 5 monthly SUPPORT-VALUE columns
    (Feb/Mar/Apr/May/Jun 2026) are read back by pandas/openpyxl as actual
    `datetime` objects, not text -- same phenomenon
    app.review_validation._stringify_header_value exists for -- so they're
    located by `.month`, never by a literal column-name string.

BM identity, verified against the real files (2026-08-19): Avg & Calls'
"Employee Code" and Visits & Support's "BM code" are the SAME code scheme
for real employees (224/224 Xandra Employee Codes found verbatim as BM
codes) -- Visits & Support additionally carries extra V-/A-prefixed codes
(vacant BM slots) that never appear in Avg & Calls and are therefore never
part of the Coverage Summary roster (the roster is built from Avg & Calls,
per spec). Matched by code, never by name (names are supporting/display
fields only, per explicit instruction).

Applicability: reuses app.hq_distribution_service.get_valid_hqs_for_division
exactly like Opus does -- a BM whose "Reporting HQ" isn't in the uploaded
HQ Distribution file for this division is NOT APPLICABLE (excluded from
the roster entirely, never generated, never flagged red) rather than
"missing" (falls back to including every BM when no HQ Distribution file
is uploaded yet, same as Opus).

RGD category: the real files only ever contain "General", "B-RGD", or
"B-RGD/A-RGD" (verified across all three divisions) -- no bare "A-RGD"
value exists anywhere. "RGD" therefore means Category != "General", not a
literal string match against "BRGD"/"ARGD".

Two DIFFERENT doctor populations per BM+month (per explicit spec
distinction, not to be confused with each other):
  - Total Rxrs / Total Dr Support / Total RGD Rxrs / Total RGD Support all
    key off the SAME "non-blank support value for this month" population
    (RGD variants additionally filtered to Category != General).
  - RGD Missed uses a WIDER population: every RGD-category doctor assigned
    to this BM, regardless of whether their support value is blank that
    month -- then counts how many of THOSE have a blank BM Visit Count for
    the month. Implementing these as the same population would silently
    change RGD Missed's meaning.

Months: Apr/May/Jun only -- Jul is deliberately not generated, because the
real Visits & Support files have a "BM Visit Count Jul-2026" column but NO
July monthly support-value column at all (verified structurally, not
assumed), so Total Rxrs/Support/etc genuinely cannot be computed for July
from current source data. COVERAGE_REPORT_MONTHS is still a plain tuple
callers iterate, exactly like OPUS_REPORT_MONTHS -- extending to July once
that source data exists is adding one entry here, not a redesign.
"""

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import REVIEW_UPLOADS_DIR
from app.hq_distribution_service import get_valid_hqs_for_division
from app.review_upload_service import get_slot_state
from app.review_validation import _select_sheet

DIVISIONS = ("Xandra", "Onyx", "Guardians")

COVERAGE_REPORT_MONTHS = ("APR", "MAY", "JUN")
MONTH_TO_NUM = {"APR": 4, "MAY": 5, "JUN": 6, "JUL": 7}
MONTH_TO_AVG_CALLS_COLUMN = {"APR": "Apr-2026", "MAY": "May-2026", "JUN": "Jun-2026", "JUL": "Jul-2026"}
MONTH_TO_BM_VISIT_COUNT_COLUMN = {
    "APR": "BM Visit Count Apr-2026", "MAY": "BM Visit Count May-2026",
    "JUN": "BM Visit Count Jun-2026", "JUL": "BM Visit Count Jul-2026",
}

ROW_LABELS = (
    "Field Visit Days", "Total Calls", "Call Average", "Total Rxrs", "Total RGD Rxrs",
    "Total Dr Support", "Total RGD Support", "% RGD Support", "RGD Missed",
)
# Rows copied/rounded straight from Avg & Calls -- no cross-row dependency.
_AVG_CALLS_ROWS = {"Field Visit Days": "Field Visit Days", "Total Calls": "# of Doctor Visits", "Call Average": "Doctor Call Average"}
PERCENT_ROWS = {"% RGD Support"}


def _norm(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return " ".join(str(value).strip().upper().split())


def _generated_output_dir() -> Path:
    out = REVIEW_UPLOADS_DIR / "generated_reports"
    out.mkdir(parents=True, exist_ok=True)
    return out


def generated_coverage_summary_path(division: str) -> Path:
    return _generated_output_dir() / f"coverage_summary_{division.strip().lower()}.xlsx"


def generated_coverage_preview_path(division: str) -> Path:
    return _generated_output_dir() / f"coverage_summary_{division.strip().lower()}.preview.json"


def _avg_calls_slot_id(division: str) -> str:
    return f"coverage_avg_calls_{division.strip().lower()}"


def _visits_support_slot_id(division: str) -> str:
    return f"coverage_visits_support_{division.strip().lower()}"


def coverage_prerequisites_ready(division: str) -> tuple:
    """(ready: bool, missing: [str])."""
    missing = []
    for slot_id in (_avg_calls_slot_id(division), _visits_support_slot_id(division)):
        state = get_slot_state(slot_id)
        if not (state["uploaded"] and state["valid"]):
            missing.append(slot_id)
    return (len(missing) == 0, missing)


# --- Source-file loading -----------------------------------------------------

def _load_avg_calls(division: str) -> pd.DataFrame:
    state = get_slot_state(_avg_calls_slot_id(division))
    return pd.read_excel(state["file_path"], sheet_name=0)


def _load_visits_support(division: str) -> pd.DataFrame:
    slot_id = _visits_support_slot_id(division)
    state = get_slot_state(slot_id)
    path = state["file_path"]
    extension = Path(path).suffix.lower()
    from app.review_schemas import get_slot_def
    slot = get_slot_def(slot_id)
    sheet = _select_sheet(path, extension, slot)
    df = pd.read_excel(path, sheet_name=sheet)
    # "BM code" (Xandra/Guardians) vs "BM Code" (Onyx) -- same column,
    # different casing per division's real file (verified 2026-08-19).
    for col in df.columns:
        if isinstance(col, str) and col.strip().lower() == "bm code":
            df = df.rename(columns={col: "BM code"})
            break
    return df


def _support_value_columns_by_month(df: pd.DataFrame) -> dict:
    """{month_code: column_key} for the datetime-typed monthly support-value
    columns -- located by `.month`, never by a literal name string (see
    module docstring)."""
    num_to_code = {v: k for k, v in MONTH_TO_NUM.items()}
    by_month = {}
    for col in df.columns:
        if isinstance(col, (datetime, date)):
            code = num_to_code.get(col.month)
            if code:
                by_month[code] = col
    return by_month


_HQ_SPELLING_ALIASES = {
    # Avg & Calls' "Reporting HQ" spelling -> HQ Distribution/Annual Targets
    # canonical spelling. Not a guess -- these are the exact same pairs
    # app/review_opus_mapping.py already hand-verified for Xandra's Opus
    # reference structure (Chapra/Purnia/Gadhinglaj/Nashik Pool), reused here.
    "CHAPRA": "CHHAPRA",
    "PURNIA": "PURNEA",
    "GADHINGLAJ": "GADHINGLAJ/KUDAL",
    "NASIK": "NASHIK POOL",
}


def _hq_is_applicable(hq_name: str, valid_hqs: set) -> bool:
    """True if `hq_name` (Avg & Calls' "Reporting HQ") matches the HQ
    Distribution file's valid-HQ set for this division -- tried as written,
    with " POOL" appended, and through _HQ_SPELLING_ALIASES, confirmed
    equivalent (2026-08-19): Avg & Calls' "Reporting HQ" field consistently
    drops the "Pool" suffix HQ Distribution/Annual Targets carry for the
    same aggregate territory (e.g. Avg & Calls says "Ahmedabad", HQ
    Distribution says "Ahmedabad Pool" -- same location). Deliberately NOT
    the reverse (stripping "Pool" off an already-suffixed name) -- every
    case observed was bare-name Avg & Calls vs Pool-suffixed HQ
    Distribution, never the other way."""
    hq_norm = _norm(hq_name)
    if hq_norm in valid_hqs or f"{hq_norm} POOL" in valid_hqs:
        return True
    alias = _HQ_SPELLING_ALIASES.get(hq_norm)
    return alias is not None and alias in valid_hqs


def _build_bm_roster(division: str, avg_df: pd.DataFrame) -> pd.DataFrame:
    """One row per distinct BM (Employee Code), identity columns only --
    filtered to HQs applicable to `division` per the uploaded HQ
    Distribution file (falls back to including every BM if that file
    hasn't been uploaded, same as app.review_opus_service._filter_applicable_blocks)."""
    roster = avg_df[["Division", "Region", "Employee Name", "Employee Code", "Desig.", "Reporting HQ"]].drop_duplicates(
        subset=["Employee Code"]
    ).reset_index(drop=True)

    valid_hqs = get_valid_hqs_for_division(division)
    if valid_hqs is not None:
        roster = roster[roster["Reporting HQ"].apply(lambda hq: _hq_is_applicable(hq, valid_hqs))]
    return roster


# --- Per-BM calculation -------------------------------------------------------

@dataclass
class ComputedBmBlock:
    division: str
    region: str
    hq: str
    emp_code: str
    name: str
    designation: str
    rows: dict = field(default_factory=dict)  # {row_label: {month_code: value}}


def _compute_bm_block(bm_row, avg_sub: pd.DataFrame, bm_docs: pd.DataFrame,
                       support_cols: dict, months: tuple) -> ComputedBmBlock:
    """`avg_sub`/`bm_docs` are this BM's ALREADY-filtered rows (Avg & Calls
    / Visits & Support respectively) -- the caller groups both DataFrames
    by employee/BM code ONCE via pandas .groupby() before the roster loop
    and passes each BM's own slice in, rather than this function
    re-scanning the full division-wide frame per BM (measured 2026-08-20:
    189 per-BM `vs_df[vs_df["BM code"] == emp_code]` scans over a
    33,346-row frame cost 1.55s for Xandra; grouping once costs 0.04s --
    same rows in, same rows out, only how they're located changes)."""
    emp_code = bm_row["Employee Code"]

    rows = {}
    for label, source_param in _AVG_CALLS_ROWS.items():
        param_row = avg_sub[avg_sub["Parameters"] == source_param]
        if param_row.empty:
            rows[label] = {m: None for m in months}
            continue
        r = param_row.iloc[0]
        values = {m: r.get(MONTH_TO_AVG_CALLS_COLUMN[m]) for m in months}
        if label == "Call Average":
            values = {m: (round(float(v)) if v is not None and not pd.isna(v) else None) for m, v in values.items()}
        else:
            values = {m: (None if v is None or pd.isna(v) else float(v)) for m, v in values.items()}
        rows[label] = values

    rgd_docs = bm_docs[bm_docs["Category"].apply(_norm) != "GENERAL"]

    total_rxrs, total_dr_support = {}, {}
    total_rgd_rxrs, total_rgd_support, pct_rgd_support = {}, {}, {}
    rgd_missed = {}

    for m in months:
        support_col = support_cols.get(m)
        if support_col is None:
            total_rxrs[m] = total_dr_support[m] = total_rgd_rxrs[m] = total_rgd_support[m] = pct_rgd_support[m] = rgd_missed[m] = None
            continue

        non_blank = bm_docs[support_col].notna()
        total_rxrs[m] = int(non_blank.sum())
        total_dr_support[m] = float(bm_docs.loc[non_blank, support_col].sum()) / 100_000

        rgd_non_blank = rgd_docs[support_col].notna()
        total_rgd_rxrs[m] = int(rgd_non_blank.sum())
        total_rgd_support[m] = float(rgd_docs.loc[rgd_non_blank, support_col].sum()) / 100_000

        pct_rgd_support[m] = (total_rgd_support[m] / total_dr_support[m]) if total_dr_support[m] else None

        visit_count_col = MONTH_TO_BM_VISIT_COUNT_COLUMN[m]
        if visit_count_col in rgd_docs.columns:
            rgd_missed[m] = int(rgd_docs[visit_count_col].isna().sum())
        else:
            rgd_missed[m] = None

    rows["Total Rxrs"] = total_rxrs
    rows["Total RGD Rxrs"] = total_rgd_rxrs
    rows["Total Dr Support"] = total_dr_support
    rows["Total RGD Support"] = total_rgd_support
    rows["% RGD Support"] = pct_rgd_support
    rows["RGD Missed"] = rgd_missed

    return ComputedBmBlock(
        division=bm_row["Division"], region=bm_row["Region"], hq=bm_row["Reporting HQ"],
        emp_code=emp_code, name=bm_row["Employee Name"], designation=bm_row["Desig."], rows=rows,
    )


# --- Workbook writing ---------------------------------------------------------

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True)
_BODY_FONT = Font(name="Calibri", size=10, bold=False)
_CENTER = Alignment(horizontal="center")


def _write_workbook(computed_blocks: list, months: tuple, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COVERAGE SUMMARY"

    headers = ("Division", "Region Name", "HQ", "Emp Code", "Name", "Designation", "No", "Parameters") + tuple(months)
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 23.25
    ws.freeze_panes = "A2"
    for col, width in zip("ABCDEFGH", (12, 18, 16, 10, 22, 12, 5, 20)):
        ws.column_dimensions[col].width = width

    row = 2
    for block in computed_blocks:
        for no, label in enumerate(ROW_LABELS, start=1):
            ws.cell(row=row, column=1, value=block.division).font = _BODY_FONT
            ws.cell(row=row, column=2, value=block.region).font = _BODY_FONT
            ws.cell(row=row, column=3, value=block.hq).font = _BODY_FONT
            ws.cell(row=row, column=4, value=block.emp_code).font = _BODY_FONT
            ws.cell(row=row, column=5, value=block.name).font = _BODY_FONT
            ws.cell(row=row, column=6, value=block.designation).font = _BODY_FONT
            ws.cell(row=row, column=7, value=no).font = _BODY_FONT
            ws.cell(row=row, column=8, value=label).font = _BODY_FONT

            is_percent = label in PERCENT_ROWS
            number_format = "0%" if is_percent else "0.00"
            values = block.rows[label]
            for col_idx, m in enumerate(months, start=9):
                cell = ws.cell(row=row, column=col_idx, value=values.get(m))
                cell.number_format = number_format
                cell.font = _BODY_FONT
                cell.alignment = _CENTER
                cell.border = _BORDER

            for col_idx in range(1, 9):
                ws.cell(row=row, column=col_idx).border = _BORDER
            row += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _format_cell(value, is_percent: bool) -> str:
    if value is None:
        return ""
    if is_percent:
        return f"{value * 100:.0f}%"
    if isinstance(value, int):
        return str(value)
    return f"{value:.2f}"


def _build_preview_rows(computed_blocks: list, months: tuple) -> list:
    rows = []
    for block in computed_blocks:
        for no, label in enumerate(ROW_LABELS, start=1):
            is_percent = label in PERCENT_ROWS
            values = block.rows[label]
            rows.append({
                "division": block.division, "region": block.region, "hq": block.hq,
                "emp_code": block.emp_code, "name": block.name, "designation": block.designation,
                "no": str(no), "particulars": label,
                "months": [_format_cell(values.get(m), is_percent) for m in months],
            })
    return rows


# --- Shared computation step (used by both the combined workbook and the
# per-BM Coverage Summary email workflow) ----------------------------------

def _compute_coverage_blocks(division: str, report_progress=None) -> list:
    """The exact roster-load + per-BM calculation pipeline, extracted so
    generate_coverage_summary() (one combined workbook) and
    generate_coverage_summary_bm_files() (one workbook per BM, for
    app/review_coverage_notification_service.py) call the SAME
    computation and can never disagree about a BM's own numbers -- pure
    extraction, no behavior change to either caller. May raise; callers
    are responsible for the "never raises externally" contract (see each
    function's own docstring)."""
    if report_progress:
        report_progress(10, "Loading Avg. & Calls...")
    avg_df = _load_avg_calls(division)

    if report_progress:
        report_progress(30, "Loading Visits & Support...")
    vs_df = _load_visits_support(division)
    support_cols = _support_value_columns_by_month(vs_df)

    if report_progress:
        report_progress(50, "Building BM roster...")
    roster = _build_bm_roster(division, avg_df)

    if report_progress:
        report_progress(60, "Calculating...")
    avg_groups = dict(list(avg_df.groupby("Employee Code")))
    vs_groups = dict(list(vs_df.groupby("BM code")))
    empty_avg = avg_df.iloc[0:0]
    empty_vs = vs_df.iloc[0:0]
    return [
        _compute_bm_block(
            bm_row,
            avg_groups.get(bm_row["Employee Code"], empty_avg),
            vs_groups.get(bm_row["Employee Code"], empty_vs),
            support_cols, COVERAGE_REPORT_MONTHS,
        )
        for _, bm_row in roster.iterrows()
    ]


# --- Top-level entry point -----------------------------------------------------

def generate_coverage_summary(division: str, report_progress=None) -> dict:
    """Generates the Coverage Summary workbook for `division`. Never
    raises for expected failure modes -- reports them in the returned
    dict, same contract as app.review_opus_service.generate_opus_summary.

    Returns:
        {
            "success": bool,
            "division": str,
            "file_path": str | None,
            "generated_at": datetime | None,
            "bm_count": int,
            "errors": [str],
        }
    """
    if report_progress:
        report_progress(0, f"Checking {division} prerequisites...")

    if division not in DIVISIONS:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "bm_count": 0, "errors": [f"Unknown division {division!r}."]}

    ready, missing = coverage_prerequisites_ready(division)
    if not ready:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "bm_count": 0, "errors": [f"Required source file(s) not uploaded/valid yet: {', '.join(missing)}"]}

    try:
        computed = _compute_coverage_blocks(division, report_progress)

        if report_progress:
            report_progress(90, "Writing workbook...")
        out_path = generated_coverage_summary_path(division)
        _write_workbook(computed, COVERAGE_REPORT_MONTHS, out_path)

        preview_path = generated_coverage_preview_path(division)
        with open(preview_path, "w", encoding="utf-8") as f:
            json.dump({
                "columns": ["Division", "Region Name", "HQ", "Emp Code", "Name", "Designation", "No", "Parameters"] + list(COVERAGE_REPORT_MONTHS),
                "rows": _build_preview_rows(computed, COVERAGE_REPORT_MONTHS),
            }, f)

    except Exception as exc:
        logger.exception(f"Coverage Summary generation failed for {division}")
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "bm_count": 0, "errors": [f"Generation failed: {exc!r}"]}

    if report_progress:
        report_progress(100, "Done.")

    logger.info(f"Coverage Summary generated for {division}: {len(computed)} BM blocks -> {out_path}")
    return {
        "success": True, "division": division, "file_path": str(out_path),
        "generated_at": datetime.now(), "bm_count": len(computed), "errors": [],
    }


# --- Per-BM Coverage Summary files (for app/review_coverage_notification_service.py) --

_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _safe_filename_component(name: str) -> str:
    """Strips characters Windows filenames can't contain and collapses
    whitespace -- preserves spaces/hyphens (unlike
    app.table_export_service.default_export_filename, which strips EVERY
    non-alphanumeric character and would mangle the required
    "Coverage Summary - <BM Name>.xlsx" format). "" (never invented text)
    if nothing safe survives."""
    cleaned = _INVALID_FILENAME_CHARS.sub("", name or "")
    return " ".join(cleaned.split())


def _bm_files_output_dir(division: str) -> Path:
    out = _generated_output_dir() / "coverage_summary_bm" / division.strip().lower()
    out.mkdir(parents=True, exist_ok=True)
    return out


def generate_coverage_summary_bm_files(division: str, report_progress=None) -> dict:
    """Generates ONE Coverage Summary .xlsx per BM for `division` -- the
    same calculations/formatting as generate_coverage_summary()'s single
    combined workbook (both call the exact same
    _compute_coverage_blocks()), just written one BM at a time via the
    SAME _write_workbook() with a single-element block list, per
    app/review_coverage_notification_service.py's "one BM = one file"
    email-workflow rule. Never touches or replaces the existing combined
    workbook (generated_coverage_summary_path) -- these are additional,
    separate files under their own output folder.

    Full-replace per division: every existing file under this division's
    own per-BM output folder is deleted before regenerating, same
    "an upload/run is a complete snapshot" convention as
    generate_coverage_summary()'s own single combined file -- never a
    stale file from a previous roster left behind.

    A BM whose sanitized display name collides with an earlier BM's in
    the SAME run (two different Employee Codes, same Name) is
    disambiguated by appending that BM's own Employee Code in
    parentheses, logged as a warning -- never silently overwriting one
    BM's file with another's (BM identity is Employee Code, not name --
    see this module's own docstring).

    Returns:
        {
            "success": bool,
            "division": str,
            "files": [{"emp_code": str, "name": str, "file_path": str}, ...],
            "errors": [str],
        }
    """
    if report_progress:
        report_progress(0, f"Checking {division} prerequisites...")

    if division not in DIVISIONS:
        return {"success": False, "division": division, "files": [], "errors": [f"Unknown division {division!r}."]}

    ready, missing = coverage_prerequisites_ready(division)
    if not ready:
        return {"success": False, "division": division, "files": [],
                "errors": [f"Required source file(s) not uploaded/valid yet: {', '.join(missing)}"]}

    try:
        computed = _compute_coverage_blocks(division, report_progress)

        if report_progress:
            report_progress(90, "Writing per-BM workbooks...")
        out_dir = _bm_files_output_dir(division)
        for old_file in out_dir.glob("*.xlsx"):
            old_file.unlink()

        files = []
        used_filenames: dict[str, str] = {}  # sanitized filename stem -> emp_code that claimed it
        for block in computed:
            stem = _safe_filename_component(block.name) or block.emp_code
            claimed_by = used_filenames.get(stem)
            if claimed_by is not None and claimed_by != block.emp_code:
                logger.warning(
                    f"Coverage Summary BM files ({division}): two BMs share the display name "
                    f"{block.name!r} ({claimed_by!r} and {block.emp_code!r}) -- disambiguating "
                    f"{block.emp_code!r}'s filename with its own Employee Code."
                )
                stem = f"{stem} ({block.emp_code})"
            used_filenames[stem] = block.emp_code

            out_path = out_dir / f"Coverage Summary - {stem}.xlsx"
            _write_workbook([block], COVERAGE_REPORT_MONTHS, out_path)
            files.append({"emp_code": block.emp_code, "name": block.name, "file_path": str(out_path)})

    except Exception as exc:
        logger.exception(f"Coverage Summary per-BM file generation failed for {division}")
        return {"success": False, "division": division, "files": [], "errors": [f"Generation failed: {exc!r}"]}

    if report_progress:
        report_progress(100, "Done.")

    logger.info(f"Coverage Summary per-BM files generated for {division}: {len(files)} file(s) -> {out_dir}")
    return {"success": True, "division": division, "files": files, "errors": []}
