"""Opus Summary generation -- the business logic that reproduces the
manually-created Opus Summary workbook from the Review System's verified
uploads, for one division (Xandra/Onyx/Guardians) at a time.

Source files (all read from app.review_upload_service's stored paths, never
re-validated here -- Review System's Uploads page already guarantees they're
present and schema-valid before generation is offered):
  - Annual Targets (opus_annual_targets) -- one sheet PER DIVISION, named
    exactly "XANDRA"/"ONYX"/"GUARDIANS" (discovered 2026-08-19; NOT a single
    shared sheet). Selecting the wrong sheet silently mixes another
    division's targets in, so this module always reads by that exact
    division-name sheet lookup, never by position or by
    review_validation._select_sheet's generic best-match heuristic (that
    heuristic exists for schema-compliance checking, a different concern,
    and ties across identically-shaped sheets).
  - Primary Sales / Last Year Primary Sales (shared across divisions,
    filtered by their own Division column).
  - Secondary Sales -- already division-specific (opus_secondary_sales_xandra
    etc.), no further Division filter applied.

Region/HQ structure: NOT derived from Annual Targets. See
app/review_opus_mapping.py's module docstring -- the manual reference
workbook (Xandra's) is the authoritative source for which (Region, HQ)
blocks exist and their order, for ALL THREE divisions alike; Annual Targets
is only ever a per-division VALUE lookup keyed by the HQ name(s) recorded
in that mapping. A block whose `annual_targets_keys` is None for a given
division is UNRESOLVED under that division and must never have a TARGET
(or anything derived from it, including No of BM) computed -- see
_compute_hq_block. Onyx and Guardians resolve far fewer of the 165 blocks
than Xandra (verified 2026-08-19: 75 and 29 respectively) because they are
genuinely smaller divisions that don't operate in every HQ Xandra does --
not a bug.

No of BM: computed fresh from each division's OWN Annual Targets rows at
generation time (summed across annual_targets_keys, same as TARGET) --
never copied from the reference workbook, which only reflects Xandra's own
BM allocation (verified: Ahmedabad Pool is 6 BM under Xandra, 5 under Onyx,
3 under Guardians -- the same HQ, three different real allocations).

Precision: every calculation is done in full float64 precision (pandas'
native dtype) and never rounded in Python -- only the generated workbook's
NUMBER FORMAT ('0.00' / '0%') controls what's displayed, exactly matching
the reference workbook (verified: its CUMULATIVE column sums to a value
that only reconciles against full-precision monthly figures, not against
the rounded-for-display ones).

Formulas vs literal values: the 7 rows that only ever aggregate ROWS
ALREADY IN THIS SHEET (CUMULATIVE columns; %ACH; %GR; YPM; TOTAL CN; DUAL
NET ACH; DUAL NET ACH %; VALUE GAIN/DEFICIT) are written as live Excel
formulas, exactly like the reference workbook (which itself uses e.g.
"=F3/F2" for %ACH) -- edit a cell in Excel and everything downstream
recalculates. The remaining rows (TARGET, PRIMARY, LY PRIMARY, SECONDARY,
SALABLE CN, EXPIRY CN, DUAL INCREMENT MIN ELIGIBLE TGT) read averaged/summed
values out of OTHER workbooks entirely, so there is nothing in-sheet for a
formula to reference -- those are written as literal computed numbers.
"""

import json
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import REVIEW_UPLOADS_DIR
from app.hq_distribution_service import get_valid_hqs_for_division
from app.review_opus_mapping import OPUS_HQ_BLOCKS_BY_DIVISION
from app.review_schemas import MonthFamily
from app.review_upload_service import get_slot_state
from app.review_validation import _match_month_in_family, _select_sheet

DIVISIONS = ("Xandra", "Onyx", "Guardians")

# The reporting window every currently-uploaded source file actually covers
# (all of Annual Targets, Primary Sales, Last Year Primary Sales, and
# Secondary Sales carry Apr-Jul 2026 data only) -- "the reporting period
# must be handled dynamically" per spec, so every calculation/writer
# function below takes `months` as a parameter rather than hardcoding 4;
# this is the one place the current window is pinned. Extending the period
# once more months of source data exist is: add the month codes here.
OPUS_REPORT_MONTHS = ("APR", "MAY", "JUN", "JUL")

MONTH_TO_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}
MONTH_TO_ANNUAL_TARGETS_COLUMN = {
    "JAN": "JAN TGT", "FEB": "FEB TGT", "MAR": "MAR TGT", "APR": "APRIL TGT",
    "MAY": "MAY TGT", "JUN": "JUN TGT", "JUL": "JULY TGT", "AUG": "AUG TGT",
    "SEP": "SEP TGT", "OCT": "OCT TGT", "NOV": "NOV TGT", "DEC": "DEC TGT",
}
# DUAL INCREMENT MIN ELIGIBLE TGT's fixed Apr-Sep window (spec-mandated,
# independent of OPUS_REPORT_MONTHS).
DUAL_INCREMENT_BASE_MONTHS = (4, 5, 6, 7, 8, 9)

ROW_LABELS = (
    "TARGET", "PRIMARY", "LY PRIMARY", "% ACH (Normal)", "% GR", "YPM (PRIMARY)",
    "SECONDARY", "SALABLE CN", "EXPIRY CN", "TOTAL CN",
    "DUAL INCREMENT MIN ELIGIBLE TGT", "DUAL INCREMENT NET ACH",
    "DUAL INCREMENT NET ACH %", "DUAL INCREMENT VALUE GAIN / DEFICIT",
)
# Rows computed purely from OTHER ROWS ALREADY IN THIS SHEET -- written as
# live Excel formulas. Every other row reads external source data and is
# written as a literal value (see module docstring).
FORMULA_ROWS = {
    "% ACH (Normal)", "% GR", "YPM (PRIMARY)", "TOTAL CN",
    "DUAL INCREMENT NET ACH", "DUAL INCREMENT NET ACH %",
    "DUAL INCREMENT VALUE GAIN / DEFICIT",
}

_SECONDARY_SALES_SEC = MonthFamily(suffix=" Sec", months=("Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"))


def _norm(value) -> str:
    """Data-value normalization (Region/HQ/Division cell contents) --
    case/whitespace only, deliberately simpler than
    app.review_validation._normalize_header_lenient (which also
    canonicalizes month/year fragments -- not relevant to plain identifier
    text like an HQ name)."""
    return " ".join(str(value).strip().upper().split()) if value is not None else ""


def _filter_applicable_blocks(division: str, hq_blocks: tuple) -> list:
    """The subset of `hq_blocks` that actually operate under `division`,
    per the uploaded HQ Distribution master file (app.hq_distribution_service)
    -- NOT APPLICABLE blocks are dropped here, before any Annual Targets
    lookup ever runs on them, so they never appear in the generated
    workbook at all (per explicit instruction, 2026-08-19: "It simply
    should not be part of that division's analysis" -- not a red/unresolved
    row, not present).

    If the HQ Distribution file hasn't been uploaded (or isn't valid) yet,
    every block is treated as applicable -- today's pre-existing behavior,
    unchanged, so generation still works before that file exists. Once
    it's uploaded, this is the ONLY place "applicable" is decided; whatever
    survives this filter still goes through the existing resolved/
    unresolved (DATA MISSING) machinery in _compute_hq_block exactly as
    before -- the two concepts are deliberately kept separate rather than
    collapsed into one condition.

    Matched by HQ NAME ALONE, never (region, hq) -- see
    get_valid_hqs_for_division's docstring for why (region-naming
    conventions are simply incompatible across these two independently
    authored files; HQ names are not). A block checks BOTH its own display
    name (block.hq) and, when already resolved against this division's
    Annual Targets sheet, that sheet's own HQ spelling
    (block.annual_targets_keys) -- the two occasionally differ (e.g. the
    reference workbook's "Rajahmundry" vs Annual Targets/HQ Distribution's
    shared "RAJAHMUNDARY"), and this file consistently uses the Annual
    Targets spelling, not the reference workbook's."""
    valid_hqs = get_valid_hqs_for_division(division)
    if valid_hqs is None:
        return list(hq_blocks)

    applicable = []
    for b in hq_blocks:
        candidate_names = {_norm(b.hq)}
        if b.annual_targets_keys is not None:
            candidate_names.update(_norm(hq) for _region, hq in b.annual_targets_keys)
        if candidate_names & valid_hqs:
            applicable.append(b)
    return applicable


def _generated_output_dir() -> Path:
    out = REVIEW_UPLOADS_DIR / "generated_reports"
    out.mkdir(parents=True, exist_ok=True)
    return out


def generated_opus_summary_path(division: str) -> Path:
    return _generated_output_dir() / f"opus_summary_{division.strip().lower()}.xlsx"


def generated_opus_preview_path(division: str) -> Path:
    """A fast-to-read sidecar next to the generated .xlsx -- the FULL grid
    (every HQ block, every row, every month, real display-formatted
    values), built once at generation time from the same computed data the
    workbook itself was written from. Exists because openpyxl can't read
    back a formula cell's VALUE from a file no spreadsheet program has ever
    opened (there is no cached result to read -- only the formula string),
    so File Preview's embedded full-workbook view reads this instead of
    re-opening the .xlsx, no Excel/LibreOffice round-trip required."""
    return _generated_output_dir() / f"opus_summary_{division.strip().lower()}.preview.json"


REQUIRED_SLOTS_FOR_OPUS = ("opus_annual_targets", "opus_primary_sales", "opus_last_year_primary_sales")


def _secondary_sales_slot_id(division: str) -> str:
    return f"opus_secondary_sales_{division.strip().lower()}"


def opus_prerequisites_ready(division: str) -> tuple:
    """(ready: bool, missing: [str]) -- every source file Opus generation
    for `division` needs, uploaded AND valid (Review System's own gate,
    reused rather than re-derived -- see
    app.review_upload_service.get_slot_state)."""
    missing = []
    for slot_id in REQUIRED_SLOTS_FOR_OPUS + (_secondary_sales_slot_id(division),):
        state = get_slot_state(slot_id)
        if not (state["uploaded"] and state["valid"]):
            missing.append(slot_id)
    return (len(missing) == 0, missing)


# --- Source-file loading -----------------------------------------------------

def _load_annual_targets(division: str) -> dict:
    """{normalized HQ: [row dict, ...]} for the division's OWN sheet in
    Annual Targets -- a list per HQ (not a single row) because a handful of
    HQ names appear as 2-3 duplicate/near-duplicate rows in the real file;
    every caller (TARGET and No of BM alike) sums across the whole list
    (see app.review_opus_mapping's module docstring -- verified against
    Xandra's real reference numbers, not assumed)."""
    state = get_slot_state("opus_annual_targets")
    wb = openpyxl.load_workbook(state["file_path"], data_only=True, read_only=True)
    sheet_name = division.strip().upper()
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Annual Targets has no {sheet_name!r} sheet (found: {wb.sheetnames}) -- "
            f"cannot generate {division} without division-specific target data."
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    by_hq: dict = {}
    for row in rows[1:]:
        record = dict(zip(header, row))
        hq = record.get("HQ")
        if hq is None:
            continue
        by_hq.setdefault(_norm(hq), []).append(record)
    wb.close()
    return by_hq


# {slot_id: (file_path, mtime, raw DataFrame)} -- caches the RAW,
# unfiltered read only; one entry per slot_id (never grows beyond the
# handful of slot_ids that exist), keyed by the file's own mtime so a
# fresh upload (new mtime) is a cache miss automatically -- no explicit
# invalidation logic needed. Measured 2026-08-20: opus_primary_sales
# takes 35s to read, opus_last_year_primary_sales takes 92s, and both are
# genuinely division-agnostic (the Division-column filter that makes each
# division's numbers different happens AFTER this read, in
# _load_primary_sales_lookups below, completely unchanged) -- reading
# them once per 3-division export instead of once per division removes
# ~254s of pure duplicate I/O with zero change to any filter/aggregation
# logic downstream of this cache.
_SOURCE_DF_CACHE: dict = {}


def _load_primary_sales_full(slot_id: str) -> pd.DataFrame:
    state = get_slot_state(slot_id)
    file_path = state["file_path"]
    mtime = os.path.getmtime(file_path)
    cached = _SOURCE_DF_CACHE.get(slot_id)
    if cached is not None and cached[0] == file_path and cached[1] == mtime:
        return cached[2]
    df = pd.read_excel(file_path, sheet_name=0)
    _SOURCE_DF_CACHE[slot_id] = (file_path, mtime, df)
    return df


def _load_primary_sales_lookups(slot_id: str, division: str) -> dict:
    """Aggregated NetSales lookups from a Primary-Sales-shaped file (used
    for both opus_primary_sales and opus_last_year_primary_sales), filtered
    to `division` up front. Returns:
        {
            "primary": {(hq_norm, month_num): SUM of NetSales},
            "cn_gst": {(hq_norm, month_num): SUM of NetSales, sign-flipped},   # SALABLE CN
            "cn_exp": {(hq_norm, month_num): SUM of NetSales, sign-flipped},   # EXPIRY CN
            "apr_sep_sum": {hq_norm: SUM of NetSales over months 4-9},
        }

    SUM, not average -- despite the master spec literally saying
    "average(NetSales)" throughout, this file is LINE-ITEM level (dozens of
    individual invoice-line rows per HQ/month, each a small fraction of the
    true total), so a literal mean() gives a meaningless per-line figure.
    Verified against the reference workbook's exact Guntur numbers
    (2026-08-19): SUM(NetSales) for Xandra/Guntur/April reproduces
    2.2074823 to full float precision; mean() gives 0.038. Every other row
    that reads this file was checked the same way and also needs SUM.

    CN sign flip -- a credit note's NetSales is stored NEGATIVE in the
    source (it reduces sales), but SALABLE CN / EXPIRY CN are reported as
    POSITIVE figures in the reference workbook: SUM(NetSales) where
    Daybook Name=CNGST for Xandra/Guntur/April is -0.5340741, and the
    reference's SALABLE CN there is +0.5340741 -- negated to match.

    Grouped once via pandas rather than re-filtering per HQ/month -- these
    files run 80k-220k rows, and generation covers 100+ HQs x 4 months."""
    df = _load_primary_sales_full(slot_id)
    df = df[df["Division"].apply(_norm) == _norm(division)].copy()
    df["_hq_norm"] = df["HQ"].apply(_norm)

    primary = df.groupby(["_hq_norm", "Month"])["NetSales"].sum().to_dict()

    cn_gst_df = df[df["Daybook Name"].apply(_norm) == "CNGST"]
    cn_gst = {k: -v for k, v in cn_gst_df.groupby(["_hq_norm", "Month"])["NetSales"].sum().to_dict().items()}

    cn_exp_df = df[df["Daybook Name"].apply(_norm) == "CNEXP"]
    cn_exp = {k: -v for k, v in cn_exp_df.groupby(["_hq_norm", "Month"])["NetSales"].sum().to_dict().items()}

    apr_sep_df = df[df["Month"].isin(DUAL_INCREMENT_BASE_MONTHS)]
    apr_sep_sum = apr_sep_df.groupby("_hq_norm")["NetSales"].sum().to_dict()

    return {"primary": primary, "cn_gst": cn_gst, "cn_exp": cn_exp, "apr_sep_sum": apr_sep_sum}


def _load_secondary_sales(division: str) -> dict:
    """{(region_norm, hq_norm): {"APR": sec_value, ...}} read DIRECTLY off
    the division's own Secondary Sales file (never averaged/summed -- see
    the master spec's SECONDARY section) -- (region, hq) keyed, not HQ
    alone, since this file legitimately has the same HQ name under more
    than one region (e.g. "Mumbai Pool" under both Mumbai - 1 and
    Mumbai - 2)."""
    slot_id = _secondary_sales_slot_id(division)
    state = get_slot_state(slot_id)
    path = state["file_path"]
    extension = Path(path).suffix.lower()
    from app.review_schemas import get_slot_def
    slot = get_slot_def(slot_id)
    sheet = _select_sheet(path, extension, slot)

    df = pd.read_excel(path, sheet_name=sheet)
    columns = list(df.columns)

    # For each OPUS_REPORT_MONTHS code, find this file's actual "<Month>'<any
    # year> Sec" column -- reuses the exact same year-agnostic family
    # matcher app.review_validation already validates Secondary Sales
    # uploads with, rather than a second hand-rolled month/year parser.
    sec_column_by_month = {}
    for code in OPUS_REPORT_MONTHS:
        target_month_num = f"{MONTH_TO_NUM[code]:02d}"
        for col in columns:
            if _match_month_in_family(str(col), _SECONDARY_SALES_SEC) == target_month_num:
                sec_column_by_month[code] = col
                break

    result: dict = {}
    for _, row in df.iterrows():
        region, hq = row.get("Region"), row.get("HQ")
        if hq is None:
            continue
        key = (_norm(region), _norm(hq))
        values = {code: row.get(col) for code, col in sec_column_by_month.items()}
        result[key] = values
    return result


# --- Per-HQ calculation -------------------------------------------------------

@dataclass
class ComputedHqBlock:
    region: str
    hq: str
    unresolved: bool
    no_of_bm: int | None = None  # None only when unresolved -- nothing to compute it from
    unresolved_reason: str | None = None
    # Only the 7 "source" rows are stored -- the 7 formula rows are derived
    # at write time (see FORMULA_ROWS) and, for validation/tests, by
    # recompute_formula_rows() below using the identical formulas.
    source_rows: dict = field(default_factory=dict)  # {row_label: {month_code: float}}


def _compute_hq_block(block, at_lookup: dict, primary_lookups: dict, ly_lookups: dict,
                       secondary_lookup: dict, months: tuple) -> ComputedHqBlock:
    if block.annual_targets_keys is None:
        return ComputedHqBlock(
            region=block.region, hq=block.hq, unresolved=True,
            unresolved_reason=(
                "No Region/HQ mapping to this division's Annual Targets sheet for this "
                "block -- see app/review_opus_mapping.py's module docstring."
            ),
        )

    at_hq_norm = _norm(block.annual_targets_keys[0][1])
    at_rows = at_lookup.get(at_hq_norm, [])
    if len(at_rows) != len(block.annual_targets_keys):
        logger.warning(
            f"Opus: Annual Targets row count for HQ {block.hq!r} changed since the mapping "
            f"was built (expected {len(block.annual_targets_keys)}, found {len(at_rows)}) -- "
            "summing whatever is currently in the file."
        )

    no_of_bm = sum(int(r.get("NO. OF BMs") or 0) for r in at_rows)

    target = {}
    for m in months:
        col = MONTH_TO_ANNUAL_TARGETS_COLUMN[m]
        target[m] = sum(float(r.get(col) or 0) for r in at_rows)

    hq_norm = at_hq_norm
    primary = {m: float(primary_lookups["primary"].get((hq_norm, MONTH_TO_NUM[m]), 0.0) or 0.0) for m in months}
    ly_primary = {m: float(ly_lookups["primary"].get((hq_norm, MONTH_TO_NUM[m]), 0.0) or 0.0) for m in months}
    salable_cn = {m: float(primary_lookups["cn_gst"].get((hq_norm, MONTH_TO_NUM[m]), 0.0) or 0.0) for m in months}
    expiry_cn = {m: float(primary_lookups["cn_exp"].get((hq_norm, MONTH_TO_NUM[m]), 0.0) or 0.0) for m in months}

    sec_key = (_norm(block.region), _norm(block.hq))
    sec_values = secondary_lookup.get(sec_key, {})
    secondary = {m: float(sec_values.get(m) or 0.0) for m in months}

    ly_apr_sep_sum = float(ly_lookups["apr_sep_sum"].get(hq_norm, 0.0) or 0.0)
    min_eligible = (ly_apr_sep_sum + 3.5 * no_of_bm) / 6.0
    dual_min_eligible = {m: min_eligible for m in months}

    return ComputedHqBlock(
        region=block.region, hq=block.hq, unresolved=False, no_of_bm=no_of_bm,
        source_rows={
            "TARGET": target,
            "PRIMARY": primary,
            "LY PRIMARY": ly_primary,
            "SECONDARY": secondary,
            "SALABLE CN": salable_cn,
            "EXPIRY CN": expiry_cn,
            "DUAL INCREMENT MIN ELIGIBLE TGT": dual_min_eligible,
        },
    )


def recompute_formula_rows(block: ComputedHqBlock, months: tuple) -> dict:
    """Pure-Python re-derivation of the 7 FORMULA_ROWS, using the exact same
    arithmetic the generated workbook's Excel formulas perform -- exists so
    tests can assert against known reference numbers without opening the
    generated .xlsx and forcing Excel/LibreOffice to recalculate it."""
    if block.unresolved:
        return {}
    s = block.source_rows
    n_months = len(months)

    pct_ach = {m: (s["PRIMARY"][m] / s["TARGET"][m] if s["TARGET"][m] else None) for m in months}
    pct_gr = {m: (s["PRIMARY"][m] / s["LY PRIMARY"][m] - 1 if s["LY PRIMARY"][m] else None) for m in months}
    ypm = {m: (s["PRIMARY"][m] / (block.no_of_bm * 1) if block.no_of_bm else None) for m in months}
    total_cn = {m: s["SALABLE CN"][m] + s["EXPIRY CN"][m] for m in months}
    net_ach = {m: s["PRIMARY"][m] - total_cn[m] for m in months}
    net_ach_pct = {
        m: (net_ach[m] / s["DUAL INCREMENT MIN ELIGIBLE TGT"][m] if s["DUAL INCREMENT MIN ELIGIBLE TGT"][m] else None)
        for m in months
    }
    gain_deficit = {m: net_ach[m] - s["DUAL INCREMENT MIN ELIGIBLE TGT"][m] for m in months}

    def cum(d):
        return sum(v for v in d.values() if v is not None)

    target_c, primary_c, ly_primary_c = cum(s["TARGET"]), cum(s["PRIMARY"]), cum(s["LY PRIMARY"])
    total_cn_c = cum(total_cn)
    net_ach_c = primary_c - total_cn_c
    min_eligible_c = cum(s["DUAL INCREMENT MIN ELIGIBLE TGT"])

    return {
        "% ACH (Normal)": {**pct_ach, "CUMMULATIVE": (primary_c / target_c if target_c else None)},
        "% GR": {**pct_gr, "CUMMULATIVE": (primary_c / ly_primary_c - 1 if ly_primary_c else None)},
        "YPM (PRIMARY)": {**ypm, "CUMMULATIVE": (primary_c / (block.no_of_bm * n_months) if block.no_of_bm else None)},
        "TOTAL CN": {**total_cn, "CUMMULATIVE": total_cn_c},
        "DUAL INCREMENT NET ACH": {**net_ach, "CUMMULATIVE": net_ach_c},
        "DUAL INCREMENT NET ACH %": {**net_ach_pct, "CUMMULATIVE": (net_ach_c / min_eligible_c if min_eligible_c else None)},
        "DUAL INCREMENT VALUE GAIN / DEFICIT": {**gain_deficit, "CUMMULATIVE": net_ach_c - min_eligible_c},
    }


# --- Workbook writing ---------------------------------------------------------

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True)
_BODY_FONT = Font(name="Calibri", size=10, bold=False)
_BOLD_FONT = Font(name="Calibri", size=10, bold=True)
_CENTER = Alignment(horizontal="center")
_LEFT = Alignment(horizontal="left")

HEADERS = ("Region", "HQ", "PARTCULARS", "No of BM", "NO") + OPUS_REPORT_MONTHS + ("CUMMULATIVE",)


def _write_workbook(division: str, computed_blocks: list, months: tuple, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OPUS SUMMARY"

    n_months = len(months)
    month_col_start = 6  # column F
    cumulative_col = month_col_start + n_months
    month_col_letters = [get_column_letter(month_col_start + i) for i in range(n_months)]
    cumulative_letter = get_column_letter(cumulative_col)

    headers = ("Region", "HQ", "PARTCULARS", "No of BM", "NO") + tuple(months) + ("CUMMULATIVE",)
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 23.25
    ws.freeze_panes = "A2"
    for col, width in zip("ABCDEFGHIJKL", (16, 18.57, 31.86, 8.29, 3.57, 7, 7, 7, 7, 12.71, 13.71, 9.14)):
        ws.column_dimensions[col].width = width

    row = 2
    for block in computed_blocks:
        block_start = row
        for no, label in enumerate(ROW_LABELS, start=1):
            ws.cell(row=row, column=1, value=block.region).font = _BODY_FONT
            ws.cell(row=row, column=2, value=block.hq).font = _BODY_FONT
            ws.cell(row=row, column=3, value=label).font = _BODY_FONT
            ws.cell(row=row, column=4, value=block.no_of_bm).font = _BODY_FONT
            ws.cell(row=row, column=5, value=no).font = _BODY_FONT

            is_percent = "%" in label
            number_format = "0%" if is_percent else "0.00"

            if block.unresolved:
                for c in month_col_letters + [cumulative_letter]:
                    cell = ws[f"{c}{row}"]
                    cell.value = "UNRESOLVED MAPPING"
                    cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    cell.alignment = _CENTER
                    cell.border = _BORDER
            elif label in FORMULA_ROWS:
                _write_formula_row(ws, row, label, block_start, month_col_letters, cumulative_letter, number_format)
            else:
                values = block.source_rows[label]
                for c, m in zip(month_col_letters, months):
                    cell = ws[f"{c}{row}"]
                    cell.value = values[m]
                    cell.number_format = number_format
                    cell.font = _BODY_FONT
                    cell.alignment = _CENTER
                    cell.border = _BORDER
                cum_cell = ws[f"{cumulative_letter}{row}"]
                cum_cell.value = f"=SUM({month_col_letters[0]}{row}:{month_col_letters[-1]}{row})"
                cum_cell.number_format = number_format
                cum_cell.font = _BOLD_FONT
                cum_cell.alignment = _CENTER
                cum_cell.border = _BORDER

            for col_idx in range(1, cumulative_col + 1):
                ws.cell(row=row, column=col_idx).border = _BORDER
            row += 1

        # spacer row (NO=15)
        ws.cell(row=row, column=1, value=block.region).font = _BOLD_FONT
        ws.cell(row=row, column=2, value=block.hq).font = _BOLD_FONT
        ws.cell(row=row, column=4, value=block.no_of_bm).font = _BODY_FONT
        ws.cell(row=row, column=5, value=15).font = _BODY_FONT
        for col_idx in range(1, cumulative_col + 1):
            ws.cell(row=row, column=col_idx).border = _BORDER
        row += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _write_formula_row(ws, row, label, block_start, month_col_letters, cumulative_letter, number_format) -> None:
    """Writes one of FORMULA_ROWS' cells as a live Excel formula referencing
    other rows already in this same 14-row block (block_start = the row
    TARGET is on). Row offsets from block_start mirror ROW_LABELS' order:
    0=TARGET 1=PRIMARY 2=LY PRIMARY 3=%ACH 4=%GR 5=YPM 6=SECONDARY
    7=SALABLE CN 8=EXPIRY CN 9=TOTAL CN 10=DUAL MIN ELIG TGT
    11=DUAL NET ACH 12=DUAL NET ACH% 13=VALUE GAIN/DEFICIT."""
    target_row = block_start
    primary_row = block_start + 1
    ly_primary_row = block_start + 2
    salable_cn_row = block_start + 7
    expiry_cn_row = block_start + 8
    total_cn_row = block_start + 9
    min_eligible_row = block_start + 10
    net_ach_row = block_start + 11

    def formula_for(col_letter):
        if label == "% ACH (Normal)":
            return f"={col_letter}{primary_row}/{col_letter}{target_row}"
        if label == "% GR":
            return f"={col_letter}{primary_row}/{col_letter}{ly_primary_row}-1"
        if label == "YPM (PRIMARY)":
            return f"={col_letter}{primary_row}/($D{row}*1)"
        if label == "TOTAL CN":
            return f"={col_letter}{salable_cn_row}+{col_letter}{expiry_cn_row}"
        if label == "DUAL INCREMENT NET ACH":
            return f"={col_letter}{primary_row}-{col_letter}{total_cn_row}"
        if label == "DUAL INCREMENT NET ACH %":
            return f"={col_letter}{net_ach_row}/{col_letter}{min_eligible_row}"
        if label == "DUAL INCREMENT VALUE GAIN / DEFICIT":
            return f"={col_letter}{net_ach_row}-{col_letter}{min_eligible_row}"
        raise ValueError(label)

    for c in month_col_letters:
        cell = ws[f"{c}{row}"]
        cell.value = formula_for(c)
        cell.number_format = number_format
        cell.font = _BOLD_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER

    if label == "YPM (PRIMARY)":
        cum_formula = f"={cumulative_letter}{primary_row}/($D{row}*{len(month_col_letters)})"
    else:
        cum_formula = formula_for(cumulative_letter)
    cum_cell = ws[f"{cumulative_letter}{row}"]
    cum_cell.value = cum_formula
    cum_cell.number_format = number_format
    cum_cell.font = _BOLD_FONT
    cum_cell.alignment = _CENTER
    cum_cell.border = _BORDER


def _format_cell(value, is_percent: bool) -> str:
    """The exact display string the workbook's own number_format ('0%' /
    '0.00') would show -- used only for the JSON preview grid, so File
    Preview's embedded table matches what Excel would render without
    duplicating rounding logic in the UI layer."""
    if value is None:
        return ""
    return f"{value * 100:.0f}%" if is_percent else f"{value:.2f}"


def _build_preview_rows(computed_blocks: list, months: tuple) -> list:
    """The full grid, one dict per sheet row (data + spacer rows alike, in
    the exact order _write_workbook lays them out), values already
    display-formatted -- see generated_opus_preview_path's docstring for
    why this exists instead of reading the .xlsx back."""
    rows = []
    for block in computed_blocks:
        for no, label in enumerate(ROW_LABELS, start=1):
            is_percent = "%" in label
            if block.unresolved:
                month_values = ["UNRESOLVED MAPPING"] * len(months)
                cumulative = "UNRESOLVED MAPPING"
            else:
                if label in FORMULA_ROWS:
                    row_values = recompute_formula_rows(block, months)[label]
                else:
                    row_values = dict(block.source_rows[label])
                    row_values["CUMMULATIVE"] = sum(row_values[m] for m in months)
                month_values = [_format_cell(row_values[m], is_percent) for m in months]
                cumulative = _format_cell(row_values["CUMMULATIVE"], is_percent)
            rows.append({
                "region": block.region, "hq": block.hq, "particulars": label,
                "no_of_bm": "" if block.no_of_bm is None else str(block.no_of_bm),
                "no": str(no), "months": month_values, "cumulative": cumulative,
                "kind": "unresolved" if block.unresolved else "data",
            })
        rows.append({
            "region": block.region, "hq": block.hq, "particulars": "",
            "no_of_bm": "" if block.no_of_bm is None else str(block.no_of_bm),
            "no": "15", "months": [""] * len(months), "cumulative": "",
            "kind": "spacer",
        })
    return rows


# --- Shared computation step (used by both the combined workbook and the
# cross-division Top Performers ranking) -----------------------------------

def _compute_opus_blocks(division: str, report_progress=None) -> list:
    """The exact applicable-HQ-filter + source-load + per-HQ calculation
    pipeline, extracted so generate_opus_summary() (one combined workbook)
    and app.review_top_performers_service (the cross-division corporate
    YPM ranking) call the SAME computation and can never disagree about a
    given HQ's own PRIMARY/no_of_bm/YPM numbers -- pure extraction, no
    behavior change to generate_opus_summary(). Assumes the caller has
    already confirmed OPUS_HQ_BLOCKS_BY_DIVISION.get(division) is not None
    and opus_prerequisites_ready(division) is True (see
    generate_opus_summary's own error-handling for that contract) -- this
    function itself does not re-check either. May raise; callers handle
    it."""
    hq_blocks = _filter_applicable_blocks(division, OPUS_HQ_BLOCKS_BY_DIVISION[division])

    if report_progress:
        report_progress(10, "Loading Annual Targets...")
    at_lookup = _load_annual_targets(division)

    if report_progress:
        report_progress(30, "Loading Primary Sales...")
    primary_lookups = _load_primary_sales_lookups("opus_primary_sales", division)

    if report_progress:
        report_progress(55, "Loading Last Year Primary Sales...")
    ly_lookups = _load_primary_sales_lookups("opus_last_year_primary_sales", division)

    if report_progress:
        report_progress(75, "Loading Secondary Sales...")
    secondary_lookup = _load_secondary_sales(division)

    if report_progress:
        report_progress(85, "Calculating...")
    return [
        _compute_hq_block(block, at_lookup, primary_lookups, ly_lookups, secondary_lookup, OPUS_REPORT_MONTHS)
        for block in hq_blocks
    ]


# --- Top-level entry point -----------------------------------------------------

def generate_opus_summary(division: str, report_progress=None) -> dict:
    """Generates the Opus Summary workbook for `division` and writes it to
    generated_opus_summary_path(division). Never raises for expected
    failure modes (missing mapping, missing source files) -- reports them
    in the returned dict instead, same contract as
    app.review_validation.validate_review_file.

    Returns:
        {
            "success": bool,
            "division": str,
            "file_path": str | None,
            "generated_at": datetime | None,
            "hq_count": int,
            "unresolved_hqs": [{"region": str, "hq": str, "reason": str}, ...],
            "errors": [str],
        }
    """
    if report_progress:
        report_progress(0, f"Checking {division} prerequisites...")

    if division not in DIVISIONS:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "hq_count": 0, "unresolved_hqs": [], "errors": [f"Unknown division {division!r}."]}

    hq_blocks = OPUS_HQ_BLOCKS_BY_DIVISION.get(division)
    if hq_blocks is None:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "hq_count": 0, "unresolved_hqs": [],
                "errors": [f"No Region/HQ mapping has been built for {division} yet -- "
                           "a manual reference workbook is needed first "
                           "(see app/review_opus_mapping.py)."]}

    ready, missing = opus_prerequisites_ready(division)
    if not ready:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "hq_count": 0, "unresolved_hqs": [],
                "errors": [f"Required source file(s) not uploaded/valid yet: {', '.join(missing)}"]}

    try:
        computed = _compute_opus_blocks(division, report_progress)

        if report_progress:
            report_progress(95, "Writing workbook...")
        out_path = generated_opus_summary_path(division)
        _write_workbook(division, computed, OPUS_REPORT_MONTHS, out_path)

        preview_path = generated_opus_preview_path(division)
        with open(preview_path, "w", encoding="utf-8") as f:
            json.dump({
                "columns": ["Region", "HQ", "PARTCULARS", "No of BM", "NO"] + list(OPUS_REPORT_MONTHS) + ["CUMMULATIVE"],
                "rows": _build_preview_rows(computed, OPUS_REPORT_MONTHS),
            }, f)

    except Exception as exc:
        logger.exception(f"Opus generation failed for {division}")
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "hq_count": 0, "unresolved_hqs": [], "errors": [f"Generation failed: {exc!r}"]}

    unresolved = [
        {"region": c.region, "hq": c.hq, "reason": c.unresolved_reason}
        for c in computed if c.unresolved
    ]
    if report_progress:
        report_progress(100, "Done.")

    logger.info(
        f"Opus Summary generated for {division}: {len(computed)} HQ blocks "
        f"({len(unresolved)} unresolved) -> {out_path}"
    )
    return {
        "success": True, "division": division, "file_path": str(out_path),
        "generated_at": datetime.now(), "hq_count": len(computed),
        "unresolved_hqs": unresolved, "errors": [],
    }
