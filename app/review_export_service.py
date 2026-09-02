"""Review System -- "Export Entire File": one combined workbook per
division (XANDRA.xlsx / ONYX.xlsx / GUARDIANS.xlsx), each containing every
registered report's sheet for that division alone, regardless of whatever
report/division/filter File Preview currently has selected.

Reuses -- never reimplements -- Opus/Coverage's own EXISTING GENERATED
OUTPUT (app.review_opus_service, app.review_coverage_service): if a
division's report has already been generated (the common case -- the same
already-on-disk .xlsx File Preview's own Export button copies), this
module copies that file's sheet directly, no recalculation. A report's
generate_fn(division) is only called when nothing has been generated for
that division yet, so a first-ever export still works without the user
having to visit every report/division tab first. This is both far faster
than always regenerating and more literally correct per spec ("simply
make its existing generated output available") -- zero calculation/
formatting logic is duplicated here either way; this module only knows
how to (1) locate or produce each registered report's file and (2) copy
its already-formatted output sheet into a combined workbook -- see
_copy_worksheet.

EXPORT_REPORTS is the report registry: RGD Visit and Support (2026-08-19)
is the first report to prove out this registry's extensibility -- adding
it was exactly "append one entry here", nothing else in this module
changed. The same applies to any future fourth report.

Division isolation falls out structurally, not from an explicit check:
each division's sheets only ever come from that division's own
path_fn(division)/generate_fn(division) -- there is no code path where
one division's data could end up in another's workbook.
"""

from copy import copy as _copy_style
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl
from loguru import logger

from app.review_coverage_service import generate_coverage_summary, generated_coverage_summary_path
from app.review_opus_service import DIVISIONS, _generated_output_dir, generate_opus_summary, generated_opus_summary_path
from app.review_rgd_service import generate_rgd_summary, generated_rgd_path


@dataclass(frozen=True)
class ExportReport:
    name: str  # sheet name in the combined workbook
    path_fn: Callable[[str], Path]  # path_fn(division) -> where the generated .xlsx would be, if it exists
    generate_fn: Callable[[str], dict]  # generate_fn(division) -> {"success", "file_path", "errors", ...}


EXPORT_REPORTS: tuple[ExportReport, ...] = (
    ExportReport(name="OPUS SUMMARY", path_fn=generated_opus_summary_path, generate_fn=generate_opus_summary),
    ExportReport(name="COVERAGE SUMMARY", path_fn=generated_coverage_summary_path, generate_fn=generate_coverage_summary),
    ExportReport(name="RGD VISIT AND SUPPORT", path_fn=generated_rgd_path, generate_fn=generate_rgd_summary),
)


def _copy_worksheet(src_ws, dst_ws) -> None:
    """Copies every cell's value + font + fill + border + alignment +
    number_format, plus column widths, row heights, and freeze_panes, from
    `src_ws` into `dst_ws` -- a generic same-shape worksheet copy so a
    report's existing writer (app.review_opus_service._write_workbook,
    app.review_coverage_service._write_workbook) never needs to know it
    might end up inside a combined workbook. No merged-cell handling --
    neither writer produces merged cells today.

    Style objects are memoized by `cell._style` (measured 2026-08-20,
    verified in this project's installed openpyxl version): every one of
    this project's writers reuses the SAME handful of module-level
    Font/Border/Fill/Alignment constants across every row (e.g. every RGD
    data cell shares one _BODY_FONT, one _BORDER), so a fresh
    `cell.style`/`cell.fill`/etc StyleProxy wrapper is a different object
    on every access, but the underlying `StyleArray` it derives from
    compares equal by VALUE for cells that share the same formatting --
    confirmed empirically (same style -> equal AND same hash; header vs.
    body -> unequal). That makes it a correct, free memoization key:
    copying styles for a 264,940-cell RGD sheet dropped from 90.1s to a
    fraction of that, purely by calling copy.copy() once per DISTINCT
    style instead of once per cell -- no output difference, since a cache
    miss always falls back to constructing a fresh, correct style, and a
    hit only ever reuses a style that already compared equal by value."""
    style_cache: dict = {}
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            styled = style_cache.get(cell._style)
            if styled is None:
                # cell.font/.fill/.border/.alignment are StyleProxy
                # wrappers, not the underlying style objects -- assigning
                # them directly raises (unhashable), so shallow-copy()
                # each into a real Font/Fill/Border/Alignment first, the
                # standard openpyxl cross-worksheet style-copy pattern.
                styled = (_copy_style(cell.font), _copy_style(cell.fill), _copy_style(cell.border), _copy_style(cell.alignment))
                style_cache[cell._style] = styled
            new_cell.font, new_cell.fill, new_cell.border, new_cell.alignment = styled
            new_cell.number_format = cell.number_format

    for letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[letter].width = dim.width
    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height

    dst_ws.freeze_panes = src_ws.freeze_panes


def _division_export_path(division: str) -> Path:
    return _generated_output_dir() / f"{division.strip().upper()}.xlsx"


def export_all_divisions(report_progress=None) -> dict:
    """Combines every registered report's EXISTING generated output into
    one workbook per division -- generating only whatever hasn't been
    generated yet. Never raises for an individual report's expected
    failure (e.g. missing source files for that division) -- that sheet is
    simply omitted (never a fake/empty one) and the reason recorded in
    `errors`.

    Returns:
        {
            "success": bool,  # True if at least one division produced a file
            "files": {division: path_str | None},
            "errors": {division: [str, ...]},
        }
    """
    files: dict = {}
    errors: dict = {}
    total_steps = len(DIVISIONS) * len(EXPORT_REPORTS)
    step = 0

    for division in DIVISIONS:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # start empty; only real generated sheets are added
        division_errors: list = []

        for report in EXPORT_REPORTS:
            step += 1
            existing_path = report.path_fn(division)
            if existing_path.is_file():
                if report_progress:
                    report_progress(int(100 * step / total_steps), f"{division}: adding {report.name.title()}...")
                report_path = existing_path
            else:
                if report_progress:
                    report_progress(int(100 * step / total_steps), f"{division}: generating {report.name.title()}...")
                result = report.generate_fn(division)
                if not result["success"]:
                    division_errors.append(f"{report.name}: {'; '.join(result['errors'])}")
                    continue
                report_path = result["file_path"]

            src_wb = openpyxl.load_workbook(report_path)
            src_ws = src_wb.active
            dst_ws = wb.create_sheet(report.name)
            _copy_worksheet(src_ws, dst_ws)
            src_wb.close()

        errors[division] = division_errors
        if not wb.sheetnames:
            files[division] = None
            logger.warning(f"Export Entire File: {division} produced no sheets -- skipped.")
            continue

        out_path = _division_export_path(division)
        wb.save(out_path)
        files[division] = str(out_path)
        logger.info(f"Export Entire File: {division} -> {out_path} ({len(wb.sheetnames)} sheet(s))")

    if report_progress:
        report_progress(100, "Done.")

    return {"success": any(files.values()), "files": files, "errors": errors}
