"""Parses an uploaded Work Distribution (RGD coverage) report into one
doctor record per row. Parsing only -- no KPI thresholds, no findings, no
database writes (see app/work_distribution_service.py for those).

This is a real business report, not a normalized database export: header
row position is not assumed (row 0 is NOT guaranteed to be the header --
title/banner rows may precede it), and the exact header text is confirmed
against a real uploaded file (Guardians MSL.xlsx, 2026-08-04) to be "Dr.
Code" / "Dr. Name" (abbreviated), not "Doctor Code" / "Doctor Name" --
COLUMN_SYNONYMS below tolerates both spellings (plus a few other likely
variants) so a differently-worded month's file doesn't break the same way
again. Matching is case/whitespace-insensitive (see `_normalize_header`);
merged header cells need no special handling -- openpyxl always assigns a
merged range's value to its top-left ("anchor") cell, which is enough on
its own for header matching (same reasoning as app/excel_validation.py's
own module docstring). A blank row's cells all normalize to "", which
never matches any real column synonym, so it's skipped automatically, same
as any other non-header row.

Supports .xlsx/.xls/.xlsm/.csv, the same set every other upload page
supports (app/excel_validation.SUPPORTED_EXTENSIONS).

The two visit columns are detected by PATTERN (starts with "bm visit" /
"abm visit", case/whitespace-insensitive), never hardcoded to a specific
month -- confirmed against the real file's "BM Visit Jul-2026"/"ABM Visit
Jul-2026" -- so a new month's report needs no code change. Whichever text
follows "BM Visit"/"ABM Visit" in the real header is carried through as
`period_label`, for display only.

Visit counting: a visit cell holds a comma-separated list of day-of-month
values ("04,12,25" = 3 visits, "25" = 1 visit, blank = 0 visits) -- see
`_count_visits`. Confirmed against the real file: a single-day cell is
often stored as a bare number (25), not a string -- `_count_visits` treats
both the same way.

Diagnostics (requirement: never a bare "missing columns" message): when no
sheet has a row satisfying every required column, `parse_work_distribution_report`
reports the BEST partial match found across every sheet scanned --
`sheet_name`, `header_row_number` (1-indexed, matching how the row would be
counted in Excel itself), `matched_columns`, `missing_columns`, and
`detected_columns` (every non-blank cell actually seen in that row) -- so a
genuinely different report shape is debuggable immediately instead of a
guessing game.
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd
from loguru import logger

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".csv"}
MAX_HEADER_SCAN_ROWS = 100

# canonical name -> acceptable normalized header text(s). Confirmed against
# a real uploaded report (Guardians MSL.xlsx) that "Dr. Code"/"Dr. Name" is
# the actual real-world spelling, not "Doctor Code"/"Doctor Name" -- both
# are accepted so a future file using either wording still parses.
COLUMN_SYNONYMS = {
    "Division": ["division"],
    "Dr. Code": ["dr. code", "dr code", "doctor code"],
    "Dr. Name": ["dr. name", "dr name", "doctor name"],
    "Speciality": ["speciality", "specialty"],
    "Category": ["category"],
    "ABM RGD": ["abm rgd"],
    "City": ["city"],
    "HQ": ["hq", "h.q."],
    "Region": ["region"],
    "BM": ["bm"],
    "ABM": ["abm"],
}

_BM_VISIT_PATTERN = re.compile(r"^bm visit\b")
_ABM_VISIT_PATTERN = re.compile(r"^abm visit\b")
_VISIT_PREFIX_STRIP_PATTERN = re.compile(r"^(bm|abm)\s+visit\s*[:\-]?\s*", re.IGNORECASE)


def _normalize_header(value) -> str:
    """Lowercase, NBSP-safe, whitespace-collapsed form of a header cell,
    used for matching only -- never for display. Mirrors
    app/excel_validation.py's own `_normalize_header`."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = text.strip().lower()
    return re.sub(r"\s+", " ", text)


def _clean(value) -> str:
    """Blank-safe, whitespace-trimmed string form of a data cell -- never
    the literal "nan" pandas produces for a blank cell."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return "" if text.lower() == "nan" else text


def _count_visits(value) -> int:
    """"25" -> 1, "04,21" -> 2, "04,12,25" -> 3, blank -> 0. A visit cell
    is a comma-separated list of day-of-month values; this only ever
    counts entries, it never interprets them as real calendar dates.
    Handles a single-day cell stored as a bare number (confirmed against
    the real report: e.g. 25, not "25") the same as a string cell."""
    text = _clean(value)
    if not text:
        return 0
    tokens = [t.strip() for t in text.split(",") if t.strip()]
    return len(tokens)


def _iter_candidate_sheets(file_path: str, extension: str):
    """Yields (sheet_name, grid) pairs to header-scan -- one per real sheet
    for Excel formats, one pseudo-sheet (sheet_name=None) for CSV. Mirrors
    app/excel_validation.py's own `_iter_candidate_sheets`."""
    if extension == ".csv":
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                df = pd.read_csv(
                    file_path, header=None, nrows=MAX_HEADER_SCAN_ROWS,
                    dtype=str, keep_default_na=False, encoding=encoding,
                )
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")
        grid = [[None if str(v) == "" else v for v in row] for row in df.itertuples(index=False, name=None)]
        yield None, grid
    elif extension == ".xls":
        book = xlrd.open_workbook(file_path)
        for sheet in book.sheets():
            max_row = min(sheet.nrows, MAX_HEADER_SCAN_ROWS)
            grid = []
            for r in range(max_row):
                row = [
                    None if sheet.cell(r, c).ctype == xlrd.XL_CELL_EMPTY else sheet.cell(r, c).value
                    for c in range(sheet.ncols)
                ]
                grid.append(row)
            yield sheet.name, grid
    else:  # .xlsx / .xlsm
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                yield sheet_name, [list(row) for row in ws.iter_rows(max_row=MAX_HEADER_SCAN_ROWS, values_only=True)]
        finally:
            wb.close()


def _match_fixed_columns(normalized_row: list) -> dict:
    """{canonical_name: column_index} for every COLUMN_SYNONYMS entry found
    anywhere in `normalized_row` -- order-independent, tolerant of any
    accepted synonym spelling."""
    synonym_sets = {name: set(_normalize_header(s) for s in syns) for name, syns in COLUMN_SYNONYMS.items()}
    column_map = {}
    for name, synonym_set in synonym_sets.items():
        for i, norm_v in enumerate(normalized_row):
            if norm_v and norm_v in synonym_set:
                column_map[name] = i
                break
    return column_map


def _find_header_row(grid: list, best_partial: dict):
    """Scans `grid` (one sheet) for a row where every COLUMN_SYNONYMS
    column is found AND at least one BM Visit / ABM Visit column is
    detected by pattern -- ignores title/blank/formatting rows simply by
    not matching them (a non-header row matches few or none of the
    required columns, so scanning continues to the next row).

    Returns (row_idx, column_map, bm_visit_idx, abm_visit_idx,
    bm_visit_header_text, abm_visit_header_text) on a full match, else
    None. Updates `best_partial` (shared across every sheet scanned) in
    place with whichever row matched the most columns, for a genuinely
    useful error if nothing anywhere fully matches."""
    for row_idx, row in enumerate(grid):
        normalized_row = [_normalize_header(v) for v in row]
        column_map = _match_fixed_columns(normalized_row)

        bm_visit_idx = next((i for i, v in enumerate(normalized_row) if _BM_VISIT_PATTERN.match(v)), None)
        abm_visit_idx = next((i for i, v in enumerate(normalized_row) if _ABM_VISIT_PATTERN.match(v)), None)

        matched_count = len(column_map) + (1 if bm_visit_idx is not None else 0) + (1 if abm_visit_idx is not None else 0)
        if matched_count > best_partial.get("matched_count", -1):
            missing = [n for n in COLUMN_SYNONYMS if n not in column_map]
            if bm_visit_idx is None:
                missing.append("BM Visit <Month>")
            if abm_visit_idx is None:
                missing.append("ABM Visit <Month>")
            best_partial["matched_count"] = matched_count
            best_partial["row_idx"] = row_idx
            best_partial["matched_columns"] = sorted(column_map.keys()) + (
                ["BM Visit"] if bm_visit_idx is not None else []
            ) + (["ABM Visit"] if abm_visit_idx is not None else [])
            best_partial["missing_columns"] = missing
            best_partial["detected_columns"] = [v for v in row if v is not None and str(v).strip() != ""]

        if len(column_map) != len(COLUMN_SYNONYMS) or bm_visit_idx is None or abm_visit_idx is None:
            continue

        return row_idx, column_map, bm_visit_idx, abm_visit_idx, row[bm_visit_idx], row[abm_visit_idx]

    return None


def _period_label(bm_visit_header, abm_visit_header) -> str:
    for header_text in (bm_visit_header, abm_visit_header):
        text = _clean(header_text)
        if not text:
            continue
        stripped = _VISIT_PREFIX_STRIP_PATTERN.sub("", text).strip()
        if stripped:
            return stripped
    return ""


def parse_work_distribution_report(file_path: str, progress_callback=None) -> dict:
    """Validate and parse an uploaded Work Distribution report.

    Returns: {success, doctors, period_label, sheet_name, missing_columns,
    error, debug}. `doctors` is a list of dicts: division, doctor_code,
    doctor_name, speciality, category, abm_rgd, city, hq, region, bm, abm,
    bm_visit_count, abm_visit_count, period_label.

    On failure to find a full header match anywhere, `debug` is populated
    with {sheet_name, header_row_number (1-indexed), matched_columns,
    missing_columns, detected_columns} describing the BEST partial match
    found across every sheet scanned -- never just a generic "missing
    columns" message with no context. Never raises -- every failure mode
    is reported via the returned dict, matching every other upload page's
    own validation contract.
    """

    def report(percent: float, message: str) -> None:
        if progress_callback:
            progress_callback(percent, message)

    logger.info(f"Parsing Work Distribution report: {file_path}")
    report(5, "Initializing...")

    extension = Path(file_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        error = f"Unsupported file type: '{extension}'"
        logger.error(f"Work Distribution report '{file_path}': {error}")
        return {
            "success": False, "doctors": [], "period_label": "", "sheet_name": None,
            "missing_columns": [], "error": error, "debug": None,
        }

    best_partial = {"matched_count": -1, "sheet_name": None}
    try:
        report(15, "Reading workbook...")
        found = None
        for sheet_name, grid in _iter_candidate_sheets(file_path, extension):
            sheet_best = {"matched_count": -1}
            match = _find_header_row(grid, sheet_best)
            if sheet_best.get("matched_count", -1) > best_partial["matched_count"]:
                best_partial = {**sheet_best, "sheet_name": sheet_name}
            if match is not None:
                found = (sheet_name, *match)
                break
    except Exception as exc:
        logger.error(f"Failed to open Work Distribution report '{file_path}': {exc}")
        return {
            "success": False, "doctors": [], "period_label": "", "sheet_name": None,
            "missing_columns": [], "error": str(exc), "debug": None,
        }

    if found is None:
        debug = {
            "sheet_name": best_partial.get("sheet_name"),
            "header_row_number": (best_partial["row_idx"] + 1) if best_partial.get("row_idx") is not None else None,
            "matched_columns": best_partial.get("matched_columns", []),
            "missing_columns": best_partial.get("missing_columns", list(COLUMN_SYNONYMS)),
            "detected_columns": best_partial.get("detected_columns", []),
        }
        logger.warning(
            f"Work Distribution report '{file_path}': required header row not found on any sheet -- "
            f"best match: sheet={debug['sheet_name']!r}, row={debug['header_row_number']}, "
            f"matched={debug['matched_columns']}, missing={debug['missing_columns']}, "
            f"detected_columns={debug['detected_columns']}"
        )
        return {
            "success": False, "doctors": [], "period_label": "", "sheet_name": debug["sheet_name"],
            "missing_columns": debug["missing_columns"], "error": None, "debug": debug,
        }

    sheet_name, row_idx, column_map, bm_visit_idx, abm_visit_idx, bm_visit_header, abm_visit_header = found
    period_label = _period_label(bm_visit_header, abm_visit_header)

    try:
        report(35, "Processing data...")
        if extension == ".csv":
            df = None
            for encoding in ("utf-8-sig", "latin-1"):
                try:
                    df = pd.read_csv(file_path, header=row_idx, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")
        elif extension == ".xls":
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=row_idx, engine="xlrd")
        else:  # .xlsx / .xlsm
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=row_idx, engine="openpyxl")
    except Exception as exc:
        logger.error(f"Failed to load Work Distribution report '{file_path}' after header detection: {exc}")
        return {
            "success": False, "doctors": [], "period_label": period_label, "sheet_name": sheet_name,
            "missing_columns": [], "error": str(exc), "debug": None,
        }

    report(60, "Renaming columns...")
    rename_map = {df.columns[col_idx]: canonical_name for canonical_name, col_idx in column_map.items()}
    rename_map[df.columns[bm_visit_idx]] = "BM Visit"
    rename_map[df.columns[abm_visit_idx]] = "ABM Visit"
    df = df.rename(columns=rename_map)

    report(75, "Counting visits...")
    doctors = []
    for _, row in df.iterrows():
        doctor_code = _clean(row.get("Dr. Code"))
        doctor_name = _clean(row.get("Dr. Name"))
        if not doctor_code and not doctor_name:
            continue  # a fully blank trailing row -- not a real doctor
        doctors.append({
            "division": _clean(row.get("Division")),
            "doctor_code": doctor_code,
            "doctor_name": doctor_name,
            "speciality": _clean(row.get("Speciality")),
            "category": _clean(row.get("Category")),
            "abm_rgd": _clean(row.get("ABM RGD")),
            "city": _clean(row.get("City")),
            "hq": _clean(row.get("HQ")),
            "region": _clean(row.get("Region")),
            "bm": _clean(row.get("BM")),
            "abm": _clean(row.get("ABM")),
            "bm_visit_count": _count_visits(row.get("BM Visit")),
            "abm_visit_count": _count_visits(row.get("ABM Visit")),
            "period_label": period_label,
        })

    report(95, "Finalizing...")
    logger.info(
        f"Work Distribution report parsed: sheet '{sheet_name}', header row {row_idx + 1}, "
        f"period '{period_label}', {len(doctors)} doctor row(s)"
    )
    report(100, "Done")
    return {
        "success": True, "doctors": doctors, "period_label": period_label, "sheet_name": sheet_name,
        "missing_columns": [], "error": None, "debug": None,
    }
