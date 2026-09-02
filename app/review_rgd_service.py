"""RGD Visit and Support -- a filtered, UNMODIFIED replica of the Visits &
Support raw upload (Category != "General" only) for one division at a
time. Deliberately the simplest report in Review System: no calculation,
no aggregation, no rounding, no HQ Distribution applicability filter --
per explicit spec (2026-08-19): "This output is essentially a filtered
replica of the raw Visits and Support data" / "Do not apply any
additional filtering unless explicitly required elsewhere by the
application" -- HQ Distribution's NOT APPLICABLE/DATA MISSING concept is
Opus/Coverage-specific business logic, not part of this report's own
two-step filter (BM Code implicit via "every row already belongs to one
BM" + Category != General).

Reuses app.review_coverage_service._load_visits_support verbatim -- same
source file (coverage_visits_support_{division}), same sheet-selection,
same "BM code"/"BM Code" cross-division casing normalization -- never a
second parser for the same file.

Column mapping, verified against the real files (2026-08-19): both the
support-value columns (Feb/Mar/Apr/May/Jun 2026) and "BM Visit Count
{Month}-2026" columns are datetime-typed / string-named per month and
genuinely DIFFER in which months exist per division (Xandra has no
Feb/Mar visit-count columns; Onyx does) -- so every month is looked up
individually by number, never assumed present. The spec's requested
columns run one month further in each direction than any real file
currently has (SUPPORT asks for Jul-26, VISITS asks for Aug) -- those
columns are still emitted (so the output shape matches spec exactly) but
are always blank for now, never fabricated, exactly like Coverage
Summary's own no-fabricated-July precedent.

"B-RGD/A-RGD" (output column 8) is the raw Category value under a
different header -- not a separate source column and not a transformed
one; every row here already has Category != "General" by construction,
so its value is always "B-RGD" or "B-RGD/A-RGD" verbatim.
"""

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import REVIEW_UPLOADS_DIR
from app.review_coverage_service import _load_visits_support, _visits_support_slot_id
from app.review_upload_service import get_slot_state

DIVISIONS = ("Xandra", "Onyx", "Guardians")

# (display header, output row key) -- identity columns 1-9.
IDENTITY_COLUMNS = (
    ("Region", "region"), ("HQ", "hq"), ("BM Code", "bm_code"), ("BM Name", "bm_name"),
    ("Dr Code", "dr_code"), ("Dr Name", "dr_name"), ("Town", "town"),
    ("B-RGD/A-RGD", "category"), ("Speciality", "speciality"),
)
_IDENTITY_SOURCE_COLUMN = {
    "region": "Region", "hq": "HQ", "bm_code": "BM code", "bm_name": "BM Name",
    "dr_code": "Dr. Code", "dr_name": "Dr. Name", "town": "Town",
    "category": "Category", "speciality": "Speciality",
}

# (display header, output row key, month number) -- SUPPORT columns 10-15.
SUPPORT_COLUMNS = (
    ("Feb-26", "support_feb", 2), ("Mar-26", "support_mar", 3), ("Apr-26", "support_apr", 4),
    ("May-26", "support_may", 5), ("Jun-26", "support_jun", 6), ("Jul-26", "support_jul", 7),
)

# (display header, output row key, month number) -- VISITS columns 16-20.
VISIT_COLUMNS = (
    ("Apr", "visit_apr", 4), ("May", "visit_may", 5), ("Jun", "visit_jun", 6),
    ("Jul", "visit_jul", 7), ("Aug", "visit_aug", 8),
)

OUTPUT_HEADERS = tuple(h for h, _ in IDENTITY_COLUMNS) + tuple(h for h, _, _ in SUPPORT_COLUMNS) + tuple(h for h, _, _ in VISIT_COLUMNS)
IDENTITY_KEYS = tuple(k for _, k in IDENTITY_COLUMNS)
SUPPORT_KEYS = tuple(k for _, k, _ in SUPPORT_COLUMNS)
VISIT_KEYS = tuple(k for _, k, _ in VISIT_COLUMNS)


def _generated_output_dir() -> Path:
    out = REVIEW_UPLOADS_DIR / "generated_reports"
    out.mkdir(parents=True, exist_ok=True)
    return out


def generated_rgd_path(division: str) -> Path:
    return _generated_output_dir() / f"rgd_visit_and_support_{division.strip().lower()}.xlsx"


def generated_rgd_preview_path(division: str) -> Path:
    return _generated_output_dir() / f"rgd_visit_and_support_{division.strip().lower()}.preview.json"


def rgd_prerequisites_ready(division: str) -> tuple:
    """(ready: bool, missing: [str]) -- only the Visits & Support upload
    itself; unlike Coverage Summary this report never reads Avg. & Calls."""
    slot_id = _visits_support_slot_id(division)
    state = get_slot_state(slot_id)
    missing = [] if (state["uploaded"] and state["valid"]) else [slot_id]
    return (len(missing) == 0, missing)


def _support_columns_by_month(df: pd.DataFrame) -> dict:
    """{month_num: column_key} for the datetime-typed monthly support-value
    columns -- located by `.month`, per app.review_coverage_service's own
    established pattern for this exact file shape."""
    by_month = {}
    for col in df.columns:
        if isinstance(col, (datetime, date)):
            by_month[col.month] = col
    return by_month


def _visit_columns_by_month(df: pd.DataFrame) -> dict:
    """{month_num: column_name} for "BM Visit Count {Month}-2026" columns
    -- located by name since (unlike support columns) these are always
    plain strings, never datetime-typed."""
    by_month = {}
    for col in df.columns:
        if isinstance(col, str) and col.startswith("BM Visit Count "):
            month_abbr = col[len("BM Visit Count "):len("BM Visit Count ") + 3]
            try:
                month_num = datetime.strptime(month_abbr, "%b").month
            except ValueError:
                continue
            by_month[month_num] = col
    return by_month


def _raw(value):
    """Preserves the source value exactly -- NaN (blank cell) becomes ""
    for display, everything else (a float, a literal "-" string, a code
    string) passes through completely untouched. Never rounds, never
    coerces a blank to 0."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return value


def _load_rgd_rows(division: str) -> list:
    """The full two-step filter -- Category != "General" is the only
    selection applied (every remaining row already carries its own BM
    Code, so there is no separate "per employee" pass to make when
    building the whole-division dataset -- see module docstring)."""
    df = _load_visits_support(division)
    rgd_df = df[df["Category"].apply(lambda v: str(v).strip().upper() if v is not None else "") != "GENERAL"]

    support_cols = _support_columns_by_month(rgd_df)
    visit_cols = _visit_columns_by_month(rgd_df)

    rows = []
    for _, r in rgd_df.iterrows():
        row = {key: _raw(r.get(_IDENTITY_SOURCE_COLUMN[key])) for key in IDENTITY_KEYS}
        for _header, key, month_num in SUPPORT_COLUMNS:
            col = support_cols.get(month_num)
            row[key] = _raw(r.get(col)) if col is not None else ""
        for _header, key, month_num in VISIT_COLUMNS:
            col = visit_cols.get(month_num)
            row[key] = _raw(r.get(col)) if col is not None else ""
        rows.append(row)
    return rows


# --- Workbook writing ---------------------------------------------------------

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(name="Calibri", size=10, bold=True)
_BODY_FONT = Font(name="Calibri", size=10, bold=False)
_CENTER = Alignment(horizontal="center")


def _write_sheet(ws, rows: list) -> None:
    for col_idx, text in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 23.25
    ws.freeze_panes = "A2"
    _widths = (16, 14, 12, 22, 12, 22, 14, 12, 16, 9, 9, 9, 9, 9, 9, 8, 8, 8, 8, 8)
    for col_idx, width in enumerate(_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    all_keys = IDENTITY_KEYS + SUPPORT_KEYS + VISIT_KEYS
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(all_keys, start=1):
            # Deliberately NO number_format override -- a raw value's own
            # type (float vs the literal string "-") controls its Excel
            # display; forcing a format here would be a transformation the
            # spec explicitly forbids ("do not round", "do not convert").
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font = _BODY_FONT
            cell.border = _BORDER


def _write_workbook(rows: list, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RGD VISIT AND SUPPORT"
    _write_sheet(ws, rows)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _format_cell(value) -> str:
    """Display string for the JSON preview -- exact passthrough, never
    rounded (see app.review_opus_service._format_cell for contrast: that
    one rounds for display, this one deliberately does not). The one
    exception is a whole-number float's trailing ".0" (e.g. a visit count
    read back as 4.0 because pandas upcasts a column with any blank cell
    to float64) -- stripped only in this preview STRING, matching how
    Excel's own General format already displays such a value in the real
    exported .xlsx (still the literal float 4.0 there, untouched)."""
    if value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _build_preview_rows(rows: list) -> list:
    all_keys = IDENTITY_KEYS + SUPPORT_KEYS + VISIT_KEYS
    return [{key: _format_cell(row[key]) for key in all_keys} | {"months": []} for row in rows]


# --- Top-level entry point -----------------------------------------------------

def generate_rgd_summary(division: str, report_progress=None) -> dict:
    """Generates the RGD Visit and Support workbook for `division`. Never
    raises for expected failure modes -- same contract as
    generate_opus_summary/generate_coverage_summary.

    Returns:
        {"success", "division", "file_path", "generated_at", "row_count", "errors"}
    """
    if report_progress:
        report_progress(0, f"Checking {division} prerequisites...")

    if division not in DIVISIONS:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "row_count": 0, "errors": [f"Unknown division {division!r}."]}

    ready, missing = rgd_prerequisites_ready(division)
    if not ready:
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "row_count": 0, "errors": [f"Required source file(s) not uploaded/valid yet: {', '.join(missing)}"]}

    try:
        if report_progress:
            report_progress(30, "Loading Visits & Support...")
        rows = _load_rgd_rows(division)

        if report_progress:
            report_progress(80, "Writing workbook...")
        out_path = generated_rgd_path(division)
        _write_workbook(rows, out_path)

        preview_path = generated_rgd_preview_path(division)
        with open(preview_path, "w", encoding="utf-8") as f:
            json.dump({"columns": list(OUTPUT_HEADERS), "rows": _build_preview_rows(rows)}, f)

    except Exception as exc:
        logger.exception(f"RGD Visit and Support generation failed for {division}")
        return {"success": False, "division": division, "file_path": None, "generated_at": None,
                "row_count": 0, "errors": [f"Generation failed: {exc!r}"]}

    if report_progress:
        report_progress(100, "Done.")

    logger.info(f"RGD Visit and Support generated for {division}: {len(rows)} row(s) -> {out_path}")
    return {
        "success": True, "division": division, "file_path": str(out_path),
        "generated_at": datetime.now(), "row_count": len(rows), "errors": [],
    }
