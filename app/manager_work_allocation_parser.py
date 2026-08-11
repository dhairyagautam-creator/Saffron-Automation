"""Parses an uploaded Manager Work Allocation report into one joint-working
record per row. Parsing only -- no ABM/BM filtering, no duplicate merging,
no findings, no database writes (see app/manager_work_allocation_service.py
for those).

Same real-business-report assumptions as app/work_distribution_parser.py
(header row position is NOT assumed to be row 1; matching is
case/whitespace-insensitive; a blank row's cells never match any column
synonym, so it's skipped automatically) -- deliberately duplicated here
rather than imported, mirroring how work_distribution_parser.py itself
duplicates rather than shares app/excel_validation.py's identical-purpose
helpers, so each report format's parser stays independently readable. This
duplication is deliberate: fixing THIS parser's header tolerance (see
below) has zero effect on work_distribution_parser.py or any other
module's parser.

Each row represents one manager working jointly with one subordinate
during one month -- fixed columns, no dynamic per-month column name (unlike
Work Distribution's "BM Visit <Month>").

Header tolerance (fixed against a real uploaded Xandra ABM report,
2026-08-04, that abbreviates several headers): REQUIRED_COLUMN_SYNONYMS
below is genuinely alias-based, not just case/whitespace-insensitive --
`_normalize_header` also strips periods before comparing, so "Emp. Code"
and "Emp Code" normalize identically with no extra synonym entry needed,
and abbreviation-only headers ("Desig." for Emp Designation, "Team Emp.
Desig." for Team Emp Designation) are matched via an explicit bare-word
synonym ("desig" / "team emp desig") once periods are stripped. Every
required column must still be found for the header row to match -- this
loosens WHICH SPELLINGS count, not WHETHER a column is required.

OPTIONAL_COLUMN_SYNONYMS covers columns the module's own spec says to
"also import and store" even though this phase's ABM calculation doesn't
use them (Rep HQ, Zone, Region, Team Emp HQ, Total Visits Done in Joint,
Dates Spent in Joint, General, B-RGD, Total Dr., Covered Dr., General
Covered, B-RGD Covered) -- their ABSENCE from an uploaded file never fails
validation (only REQUIRED_COLUMN_SYNONYMS gates success); their PRESENCE
is captured into the parsed record dict for
app.manager_work_allocation_service to persist.

Supports .xlsx/.xls/.xlsm/.csv, the same set every other upload page
supports.

Diagnostics (never a bare "missing columns" message): when no sheet has a
row satisfying every REQUIRED column, `parse_manager_work_allocation_report`
reports the BEST partial match found across every sheet scanned --
`sheet_name`, `header_row_number` (1-indexed), `matched_columns`,
`missing_columns`, and `detected_columns` -- same contract as
app.work_distribution_parser.parse_work_distribution_report. A file
missing a genuinely required column (e.g. no Month column anywhere) is
still correctly rejected -- alias tolerance only widens accepted
spellings, it does not make any required column optional.
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd
from loguru import logger

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".csv"}
MAX_HEADER_SCAN_ROWS = 100

# canonical name -> acceptable normalized header text(s), ALL of which must
# be found in a row for it to count as the header row. Synonym strings are
# written in already-period-stripped form for readability -- `_normalize_header`
# strips periods from the REAL header cell before comparing, so "Emp. Code"
# and "Emp Code" both normalize to "emp code" and match the single "emp
# code" entry below; no separate dotted variant is needed. "desig"/"team
# emp desig" are the bare-abbreviation forms confirmed against a real
# uploaded Xandra ABM report that used "Desig."/"Team Emp. Desig." instead
# of the fully-spelled-out column names.
REQUIRED_COLUMN_SYNONYMS = {
    "Division": ["division"],
    "Emp Code": ["emp code", "employee code"],
    "Emp Name": ["emp name", "employee name"],
    "Emp Designation": ["emp designation", "employee designation", "desig"],
    "Month": ["month"],
    "Team Emp Code": ["team emp code", "team employee code"],
    "Team Emp Name": ["team emp name", "team employee name"],
    "Team Emp Designation": [
        "team emp designation", "team employee designation", "team emp desig", "team desig",
    ],
    "# Days Spent In Joint": [
        "# days spent in joint", "days spent in joint", "no of days spent in joint",
        "days spent in joint working", "# days spent in joint working",
    ],
}

# Imported and stored if present, but their absence never fails validation
# -- see module docstring. Canonical names double as the record dict keys
# used below (snake_case via `_OPTIONAL_RECORD_KEYS`).
OPTIONAL_COLUMN_SYNONYMS = {
    "Rep HQ": ["rep hq"],
    "Zone": ["zone"],
    "Region": ["region"],
    "Team Emp HQ": ["team emp hq", "team employee hq"],
    "Total Visits Done in Joint": ["total visits done in joint", "total visits in joint"],
    "Dates Spent in Joint": ["dates spent in joint"],
    "General": ["general"],
    "B-RGD": ["b-rgd", "b rgd", "brgd"],
    "Total Dr.": ["total dr", "total doctor", "total doctors"],
    "Covered Dr.": ["covered dr", "covered doctor", "covered doctors"],
    "General Covered": ["general covered"],
    "B-RGD Covered": ["b-rgd covered", "b rgd covered", "brgd covered"],
}

# canonical OPTIONAL_COLUMN_SYNONYMS name -> record dict key.
_OPTIONAL_RECORD_KEYS = {
    "Rep HQ": "rep_hq",
    "Zone": "zone",
    "Region": "region",
    "Team Emp HQ": "team_emp_hq",
    "Total Visits Done in Joint": "total_visits_done_in_joint",
    "Dates Spent in Joint": "dates_spent_in_joint",
    "General": "general",
    "B-RGD": "b_rgd",
    "Total Dr.": "total_dr",
    "Covered Dr.": "covered_dr",
    "General Covered": "general_covered",
    "B-RGD Covered": "b_rgd_covered",
}

# Optional columns whose value is a genuine count, parsed the same
# blank-safe way as "# Days Spent In Joint" -- every other optional column
# is stored as raw cleaned text (see module docstring: no confirmed
# business rule yet for General/B-RGD/Total Dr./Covered Dr.'s exact
# numeric semantics, so nothing is assumed here beyond what's certain).
_OPTIONAL_COUNT_COLUMNS = {"Total Visits Done in Joint"}


def _normalize_header(value) -> str:
    """Lowercase, NBSP-safe, whitespace-collapsed, PERIOD-stripped form of
    a header cell, used for matching only -- never for display. The
    period-stripping is the key difference from
    app.work_distribution_parser's own `_normalize_header`: it's what lets
    "Emp. Code"/"Desig."/"Team Emp. Desig." match the same synonym entries
    as their fully-spelled, unabbreviated equivalents without listing every
    dotted variant by hand."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = text.replace(".", "")
    text = text.strip().lower()
    return re.sub(r"\s+", " ", text)


def _clean(value) -> str:
    """Blank-safe, whitespace-trimmed string form of a data cell -- never
    the literal "nan" pandas produces for a blank cell."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return "" if text.lower() == "nan" else text


def _parse_count(value) -> int:
    """Blank -> 0; otherwise the cell's numeric value rounded to the
    nearest whole number -- handles both a bare number (4) and a numeric
    string ("4" or "4.0") the same way. Used for "# Days Spent In Joint"
    and the optional "Total Visits Done in Joint" column."""
    text = _clean(value)
    if not text:
        return 0
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def _iter_candidate_sheets(file_path: str, extension: str):
    """Yields (sheet_name, grid) pairs to header-scan -- one per real sheet
    for Excel formats, one pseudo-sheet (sheet_name=None) for CSV. Mirrors
    app/work_distribution_parser.py's own `_iter_candidate_sheets`."""
    if extension == ".csv":
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                df = pd.read_csv(
                    file_path, header=None, nrows=MAX_HEADER_SCAN_ROWS,
                    dtype=str, keep_default_na=False, encoding=encoding,
                )
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")
        grid = [[None if str(v) == "" else v for v in row] for row in df.itertuples(index=False, name=None)]
        yield None, grid
    elif extension == ".xls":
        book = xlrd.open_workbook(file_path)
        for sheet in book.sheets():
            max_row = min(sheet.nrows, MAX_HEADER_SCAN_ROWS)
            grid = []
            for r in range(max_row):
                row = [
                    None if sheet.cell(r, c).ctype == xlrd.XL_CELL_EMPTY else sheet.cell(r, c).value
                    for c in range(sheet.ncols)
                ]
                grid.append(row)
            yield sheet.name, grid
    else:  # .xlsx / .xlsm
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                yield sheet_name, [list(row) for row in ws.iter_rows(max_row=MAX_HEADER_SCAN_ROWS, values_only=True)]
        finally:
            wb.close()


def _match_columns(normalized_row: list, synonyms: dict) -> dict:
    """{canonical_name: column_index} for every entry in `synonyms` found
    anywhere in `normalized_row` -- order-independent, tolerant of any
    accepted synonym spelling. Generic over which synonym dict is passed,
    so the same matcher serves both REQUIRED_COLUMN_SYNONYMS and
    OPTIONAL_COLUMN_SYNONYMS."""
    synonym_sets = {name: set(_normalize_header(s) for s in syns) for name, syns in synonyms.items()}
    column_map = {}
    for name, synonym_set in synonym_sets.items():
        for i, norm_v in enumerate(normalized_row):
            if norm_v and norm_v in synonym_set:
                column_map[name] = i
                break
    return column_map


def _find_header_row(grid: list, best_partial: dict):
    """Scans `grid` (one sheet) for a row where every REQUIRED_COLUMN_SYNONYMS
    column is found -- ignores title/blank/formatting rows simply by not
    matching them. Returns (row_idx, column_map) on a full match (column_map
    covers BOTH required and any optional columns found in that row), else
    None. Updates `best_partial` (shared across every sheet scanned) in
    place with whichever row matched the most REQUIRED columns, for a
    genuinely useful error if nothing anywhere fully matches."""
    for row_idx, row in enumerate(grid):
        normalized_row = [_normalize_header(v) for v in row]
        required_map = _match_columns(normalized_row, REQUIRED_COLUMN_SYNONYMS)

        matched_count = len(required_map)
        if matched_count > best_partial.get("matched_count", -1):
            best_partial["matched_count"] = matched_count
            best_partial["row_idx"] = row_idx
            best_partial["matched_columns"] = sorted(required_map.keys())
            best_partial["missing_columns"] = [n for n in REQUIRED_COLUMN_SYNONYMS if n not in required_map]
            best_partial["detected_columns"] = [v for v in row if v is not None and str(v).strip() != ""]

        if len(required_map) != len(REQUIRED_COLUMN_SYNONYMS):
            continue

        optional_map = _match_columns(normalized_row, OPTIONAL_COLUMN_SYNONYMS)
        return row_idx, {**required_map, **optional_map}

    return None


def parse_manager_work_allocation_report(file_path: str, progress_callback=None) -> dict:
    """Validate and parse an uploaded Manager Work Allocation report.

    Returns: {success, records, sheet_name, missing_columns, error, debug}.
    `records` is a list of dicts: division, emp_code, emp_name,
    emp_designation, team_emp_code, team_emp_name, team_emp_designation,
    month, joint_days, plus every _OPTIONAL_RECORD_KEYS field (blank/0 if
    that optional column wasn't present in this particular file).

    On failure to find a full header match anywhere, `debug` is populated
    the same way as app.work_distribution_parser's own parse function.
    Never raises -- every failure mode is reported via the returned dict.
    """

    def report(percent: float, message: str) -> None:
        if progress_callback:
            progress_callback(percent, message)

    logger.info(f"Parsing Manager Work Allocation report: {file_path}")
    report(5, "Initializing...")

    extension = Path(file_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        error = f"Unsupported file type: '{extension}'"
        logger.error(f"Manager Work Allocation report '{file_path}': {error}")
        return {
            "success": False, "records": [], "sheet_name": None,
            "missing_columns": [], "error": error, "debug": None,
        }

    best_partial = {"matched_count": -1, "sheet_name": None}
    try:
        report(15, "Reading workbook...")
        found = None
        for sheet_name, grid in _iter_candidate_sheets(file_path, extension):
            sheet_best = {"matched_count": -1}
            match = _find_header_row(grid, sheet_best)
            if sheet_best.get("matched_count", -1) > best_partial["matched_count"]:
                best_partial = {**sheet_best, "sheet_name": sheet_name}
            if match is not None:
                found = (sheet_name, *match)
                break
    except Exception as exc:
        logger.error(f"Failed to open Manager Work Allocation report '{file_path}': {exc}")
        return {
            "success": False, "records": [], "sheet_name": None,
            "missing_columns": [], "error": str(exc), "debug": None,
        }

    if found is None:
        debug = {
            "sheet_name": best_partial.get("sheet_name"),
            "header_row_number": (best_partial["row_idx"] + 1) if best_partial.get("row_idx") is not None else None,
            "matched_columns": best_partial.get("matched_columns", []),
            "missing_columns": best_partial.get("missing_columns", list(REQUIRED_COLUMN_SYNONYMS)),
            "detected_columns": best_partial.get("detected_columns", []),
        }
        logger.warning(
            f"Manager Work Allocation report '{file_path}': required header row not found on any sheet -- "
            f"best match: sheet={debug['sheet_name']!r}, row={debug['header_row_number']}, "
            f"matched={debug['matched_columns']}, missing={debug['missing_columns']}, "
            f"detected_columns={debug['detected_columns']}"
        )
        return {
            "success": False, "records": [], "sheet_name": debug["sheet_name"],
            "missing_columns": debug["missing_columns"], "error": None, "debug": debug,
        }

    sheet_name, row_idx, column_map = found

    try:
        report(35, "Processing data...")
        if extension == ".csv":
            df = None
            for encoding in ("utf-8-sig", "latin-1"):
                try:
                    df = pd.read_csv(file_path, header=row_idx, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")
        elif extension == ".xls":
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=row_idx, engine="xlrd")
        else:  # .xlsx / .xlsm
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=row_idx, engine="openpyxl")
    except Exception as exc:
        logger.error(f"Failed to load Manager Work Allocation report '{file_path}' after header detection: {exc}")
        return {
            "success": False, "records": [], "sheet_name": sheet_name,
            "missing_columns": [], "error": str(exc), "debug": None,
        }

    report(60, "Renaming columns...")
    rename_map = {df.columns[col_idx]: canonical_name for canonical_name, col_idx in column_map.items()}
    df = df.rename(columns=rename_map)

    report(80, "Counting joint days...")
    records = []
    for _, row in df.iterrows():
        emp_code = _clean(row.get("Emp Code"))
        emp_name = _clean(row.get("Emp Name"))
        team_emp_code = _clean(row.get("Team Emp Code"))
        team_emp_name = _clean(row.get("Team Emp Name"))
        if not emp_code and not emp_name and not team_emp_code and not team_emp_name:
            continue  # a fully blank trailing row -- not a real record

        record = {
            "division": _clean(row.get("Division")),
            "emp_code": emp_code,
            "emp_name": emp_name,
            "emp_designation": _clean(row.get("Emp Designation")),
            "month": _clean(row.get("Month")),
            "team_emp_code": team_emp_code,
            "team_emp_name": team_emp_name,
            "team_emp_designation": _clean(row.get("Team Emp Designation")),
            "joint_days": _parse_count(row.get("# Days Spent In Joint")),
        }
        for canonical_name, record_key in _OPTIONAL_RECORD_KEYS.items():
            cell = row.get(canonical_name)
            record[record_key] = _parse_count(cell) if canonical_name in _OPTIONAL_COUNT_COLUMNS else _clean(cell)
        records.append(record)

    report(95, "Finalizing...")
    logger.info(
        f"Manager Work Allocation report parsed: sheet '{sheet_name}', header row {row_idx + 1}, "
        f"{len(records)} record(s)"
    )
    report(100, "Done")
    return {
        "success": True, "records": records, "sheet_name": sheet_name,
        "missing_columns": [], "error": None, "debug": None,
    }
