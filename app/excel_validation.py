"""Robust Excel validation engine for the Inventory Monitoring module.

This module validates and loads uploaded workbooks only. It does NOT
calculate thresholds, does NOT write to any database, and does NOT
compare inventory. Everything here returns an in-memory pandas DataFrame
at most.

Supports four upload formats -- .xlsx, .xls, .xlsm, .csv (see
SUPPORTED_EXTENSIONS) -- via one format-dispatch seam
(`_iter_candidate_sheets`). Every downstream behavior (header-row
scanning, required-columns matching, `_normalize_header`) is completely
format-agnostic: it only ever sees a grid of (sheet_name, rows) pairs,
regardless of which format produced them.

Two entry points, both thin wrappers around one shared `_validate_and_load`
engine (same scan/load/rename logic, only the required-columns list and
report label differ):

- `validate_inventory_report(file_path)` — requires
  INVENTORY_REPORT_REQUIRED_COLUMNS.
- `load_sales_report(file_path)` — requires SALES_REPORT_REQUIRED_COLUMNS.

Robustness notes:
- Leading/trailing/multiple/non-breaking spaces, case, and a trailing
  period are all normalized before any column-name comparison (see
  `_normalize_header`) -- the trailing-period case exists specifically
  because real Tally-style exports commonly end column headers with a
  period (e.g. "TotalQty.", "InvoiceNo.").
- A header row is found by scanning, not assumed to be row 0 — this
  handles blank/banner rows before the real header for free.
- Every sheet is scanned in order, not just the first — this handles a
  workbook whose real data isn't on the first sheet.
- Merged header cells need no special resolution: openpyxl always
  assigns a merged range's value to its top-left ("anchor") cell, and
  that anchor cell is enough on its own for header MATCHING (we only
  need the value to appear somewhere in the row, not at every column
  the merge visually spans) -- confirmed empirically, not assumed.
- Hidden rows need no special handling at all: "hidden" is a purely
  cosmetic row attribute in Excel: openpyxl/pandas read a hidden row's
  cell values exactly the same as a visible row's.

Performance: the header-row scan opens the workbook with
`read_only=True` and reads at most MAX_HEADER_SCAN_ROWS rows via
`ws.iter_rows(...)` (openpyxl's fast streaming API) -- headers are
always near the top, so nothing past that cap is ever parsed, regardless
of how many data rows the sheet actually has. The real data load
afterward (`pd.read_excel`) is unbounded and unchanged -- it already
used pandas' own fast reader, never the bottleneck. Do not switch the
header-scan back to the plain (non-read_only) `load_workbook()` + manual
`ws.cell(row, column)` grid-walk: that parses the ENTIRE workbook into
memory up front regardless of how few rows the caller's own loop
touches afterward, which is what made large uploads slow before this
was fixed.
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd
import xlrd
from loguru import logger

MAX_HEADER_SCAN_ROWS = 100  # headers are always near the top; never scan a whole large sheet to find one

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".csv"}

INVENTORY_REPORT_REQUIRED_COLUMNS = [
    "BranchLocation",
    "Item Group",
    "Item Code",
    "Item Name",
    "TotalQty",  # represents Closing Stock for the Inventory Report specifically -- see app/replenishment_service.py
    "Transit Stock",
]

SALES_REPORT_REQUIRED_COLUMNS = [
    "Division",
    "CFA",
    "Item Name",
    "Packing",
    "Sales",
]

HISTORICAL_PAYMENT_REPORT_REQUIRED_COLUMNS = [
    "Month",
    "Party Name",
    "Type",
    "Inv. No.",
    "LR Date",
    "Due Date",
    "Clear Date",
]

OUTSTANDING_REPORT_REQUIRED_COLUMNS = [
    "Party Name",
    "Division",
    "H.Q.",
    "Bill No.",
    "L.R. Date",
    "Due Date",
    "Bill Amount",
    "Month",
]


def _normalize_header(value) -> str:
    """Lowercase, NBSP-safe, whitespace-collapsed, trailing-period-stripped
    form of a header cell value, used for matching only -- never for
    anything shown to the user, which always uses the original/canonical
    spelling. The trailing-period strip is narrow (only a run of periods
    at the very end) so it never touches internal-period headers like
    "M.R.P." that aren't required-column targets anyway."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")  # explicit NBSP -> space, don't rely on regex \s semantics
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text).strip()
    return re.sub(r"\.+$", "", text)


def _is_blank_row(row: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _build_grid(ws) -> list:
    """Read at most MAX_HEADER_SCAN_ROWS rows via openpyxl's fast
    streaming API (`ws` must come from a read_only=True workbook) --
    bounded work regardless of how many rows the sheet actually has."""
    return [list(row) for row in ws.iter_rows(max_row=MAX_HEADER_SCAN_ROWS, values_only=True)]


def _iter_openpyxl_sheets(file_path: str):
    """Yields (sheet_name, grid) for every sheet in an .xlsx/.xlsm
    workbook -- openpyxl reads both formats identically, it never
    inspects the file extension. Extracted verbatim from the original
    `_scan_workbook_for_header` loop (same read_only fast-streaming
    path, same MAX_HEADER_SCAN_ROWS bound)."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            yield sheet_name, _build_grid(wb[sheet_name])
    finally:
        wb.close()


def _iter_xls_sheets(file_path: str):
    """Yields (sheet_name, grid) for every sheet in a legacy binary .xls
    workbook via xlrd -- openpyxl cannot open this format at all.
    Normalizes xlrd's empty-cell convention ('' for an empty cell) to
    None so it matches openpyxl's convention, which the rest of this
    module's blank-row/header-matching logic already assumes."""
    book = xlrd.open_workbook(file_path)
    for sheet in book.sheets():
        max_row = min(sheet.nrows, MAX_HEADER_SCAN_ROWS)
        grid = []
        for r in range(max_row):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                row.append(None if cell.ctype == xlrd.XL_CELL_EMPTY else cell.value)
            grid.append(row)
        yield sheet.name, grid


def _build_csv_grid(file_path: str) -> list:
    """Reads at most MAX_HEADER_SCAN_ROWS rows of a CSV as a plain grid
    (no header assumption -- header-row detection happens the same way
    as for Excel formats). Tries utf-8-sig first (handles the BOM Excel
    commonly adds to CSV exports), falls back to latin-1 (never raises
    on decode -- accepts any byte) for CSVs from regional Windows ERP
    exports that aren't pure UTF-8."""
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            df = pd.read_csv(
                file_path, header=None, nrows=MAX_HEADER_SCAN_ROWS,
                dtype=str, keep_default_na=False, encoding=encoding,
            )
            break
        except UnicodeDecodeError:
            continue
    else:
        raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")

    return [
        [None if str(v) == "" else v for v in row]
        for row in df.itertuples(index=False, name=None)
    ]


def _iter_candidate_sheets(file_path: str, extension: str):
    """Yields (sheet_name, grid) pairs to header-scan -- one pair per
    real sheet for Excel formats, exactly one pseudo-sheet
    (sheet_name=None) for CSV, which has no sheet concept. This is the
    one seam where format-specific logic lives; everything downstream
    (_find_header_row, _normalize_header) only ever sees a plain grid."""
    if extension == ".csv":
        yield None, _build_csv_grid(file_path)
    elif extension == ".xls":
        yield from _iter_xls_sheets(file_path)
    else:  # .xlsx / .xlsm
        yield from _iter_openpyxl_sheets(file_path)


def _find_first_nonblank_row(grid: list) -> int:
    for row_idx, row in enumerate(grid):
        if not _is_blank_row(row):
            return row_idx
    return 0


def _find_header_row(grid: list, required_columns: list, best_partial: dict) -> tuple | None:
    """Scan `grid` for a row where every required column is found
    (normalized, order-independent). Returns (row_idx, {name: col_idx})
    on a full match. Updates `best_partial` in place with whichever row
    matched the most required columns, for useful debug output if no row
    anywhere fully matches."""
    targets = {name: _normalize_header(name) for name in required_columns}

    for row_idx, row in enumerate(grid):
        normalized_row = [_normalize_header(v) for v in row]
        column_map = {}
        for name, target in targets.items():
            if target and target in normalized_row:
                column_map[name] = normalized_row.index(target)

        if len(column_map) == len(required_columns):
            return row_idx, column_map

        if len(column_map) > best_partial.get("matched_count", -1):
            best_partial["matched_count"] = len(column_map)
            best_partial["row_idx"] = row_idx
            best_partial["matched_columns"] = sorted(column_map.keys())
            best_partial["missing_columns"] = [n for n in required_columns if n not in column_map]
            best_partial["raw_row"] = [v for v in row if v is not None and str(v).strip() != ""]

    return None


def _scan_workbook_for_header(file_path: str, required_columns: list, extension: str) -> dict:
    """Scans every candidate sheet (format-dispatched via
    `_iter_candidate_sheets`, each bounded to MAX_HEADER_SCAN_ROWS) for
    the first row that satisfies every required column. Returns a dict:
    {found: bool, sheet_name, header_row_index, column_map, best_partial}."""
    best_partial = {"matched_count": -1, "sheet_name": None}
    for sheet_name, grid in _iter_candidate_sheets(file_path, extension):
        sheet_best = {}
        match = _find_header_row(grid, required_columns, sheet_best)
        if sheet_best.get("matched_count", -1) > best_partial["matched_count"]:
            best_partial = {**sheet_best, "sheet_name": sheet_name}
        if match is not None:
            row_idx, column_map = match
            return {
                "found": True,
                "sheet_name": sheet_name,
                "header_row_index": row_idx,
                "column_map": column_map,
                "best_partial": best_partial,
            }

    return {"found": False, "sheet_name": None, "header_row_index": None, "column_map": None, "best_partial": best_partial}


def _validate_and_load(file_path: str, required_columns: list, report_label: str, progress_callback=None) -> dict:
    """Shared engine behind both public entry points: scan for a header
    row satisfying `required_columns` (across every sheet), load the data,
    rename matched columns to their canonical (required_columns) spelling.
    Never raises -- every failure mode (unreadable file, missing columns)
    is reported via the returned dict.

    `progress_callback`, if given, is called as `progress_callback(percent,
    message)` at each phase -- purely cosmetic instrumentation, does not
    affect validation outcomes. Safe to omit (defaults to None).

    Returns: {success, df, missing_columns, sheet_name, header_row_index,
    detected_columns, error}
    """
    def report(percent: float, message: str) -> None:
        if progress_callback:
            progress_callback(percent, message)

    logger.info(f"Validating {report_label}: {file_path}")
    report(5, "Initializing...")

    extension = Path(file_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        # Defensive fallback only -- the UI layer already rejects
        # unsupported extensions before this engine is ever reached.
        error = f"Unsupported file type: '{extension}'"
        logger.error(f"{report_label} '{file_path}': {error}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": error,
        }

    try:
        report(15, "Reading workbook...")
        scan = _scan_workbook_for_header(file_path, required_columns, extension)
    except Exception as exc:
        logger.error(f"Failed to open {report_label} '{file_path}': {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": str(exc),
        }

    if not scan["found"]:
        best = scan["best_partial"]
        missing_columns = best.get("missing_columns", required_columns)
        detected_columns = best.get("raw_row", [])
        logger.warning(f"{report_label} missing required columns: {missing_columns}")
        logger.warning(
            f"{report_label} best-match sheet '{best.get('sheet_name')}' row {best.get('row_idx')} "
            f"detected columns: {detected_columns}"
        )
        return {
            "success": False, "df": None, "missing_columns": missing_columns, "sheet_name": best.get("sheet_name"),
            "header_row_index": best.get("row_idx"), "detected_columns": detected_columns, "error": None,
        }

    sheet_name = scan["sheet_name"]
    header_row_index = scan["header_row_index"]
    column_map = scan["column_map"]

    try:
        report(35, "Processing data...")
        if extension == ".csv":
            df = None
            for encoding in ("utf-8-sig", "latin-1"):
                try:
                    df = pd.read_csv(file_path, header=header_row_index, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")
        elif extension == ".xls":
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index, engine="xlrd")
        else:  # .xlsx / .xlsm
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index, engine="openpyxl")
    except Exception as exc:
        logger.error(f"Failed to load {report_label} '{file_path}' after header detection: {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": sheet_name,
            "header_row_index": header_row_index, "detected_columns": [], "error": str(exc),
        }

    report(75, "Validating...")
    rename_map = {df.columns[col_idx]: canonical_name for canonical_name, col_idx in column_map.items()}
    df = df.rename(columns=rename_map)

    report(95, "Finalizing...")
    logger.info(
        f"{report_label} validated successfully: sheet '{sheet_name}', header row {header_row_index}, "
        f"{len(df)} data row(s)"
    )

    report(100, "Done")
    return {
        "success": True, "df": df, "missing_columns": [], "sheet_name": sheet_name,
        "header_row_index": header_row_index, "detected_columns": df.columns.tolist(), "error": None,
    }


PIVOT_IDENTITY_LABELS = ["Alias", "Product"]
PIVOT_FIELD_LABELS = ["Net Sales", "Net Sales Value", "Closing", "Closing Value", "Transit"]


def _forward_fill_row(row: list) -> list:
    """Merged CFA-name header cells only carry a value in their leftmost
    (anchor) column -- openpyxl/xlrd/csv all read every other column in
    the merge back as None/blank. Forward-fill left-to-right so every
    column ends up associated with its group's name."""
    filled = []
    last = None
    for v in row:
        if v is not None and str(v).strip() != "":
            last = v
        filled.append(last)
    return filled


def _detect_pivot_groups(grid: list):
    """Scans `grid` for the new pivoted-by-CFA Inventory Report header:
    a row whose first two cells normalize to 'alias'/'product', followed
    immediately by a row whose columns from index 2 onward repeat
    PIVOT_FIELD_LABELS in exact order, in unbroken groups of 5 -- one
    group per CFA/branch. The CFA name for each group comes from the
    identity row (forward-filled, since it's a merged cell spanning the
    group's 5 columns).

    Three distinct outcomes, so the caller can tell "not a pivoted file"
    (safe to fall back to the old flat-row format) apart from "clearly a
    pivoted file, but malformed" (should report a pivot-specific error
    instead of silently falling back to a misleading flat-format one):
    - (row_idx, groups) with groups non-empty: a full match.
    - (row_idx, []): an Alias/Product identity row was found, but its
      field-labels row yielded zero valid CFA groups.
    - None: no Alias/Product identity row exists anywhere in the grid.
    """
    target_fields = [_normalize_header(f) for f in PIVOT_FIELD_LABELS]
    first_identity_row_idx = None
    for row_idx in range(len(grid) - 1):
        row = grid[row_idx]
        if len(row) < 2:
            continue
        if _normalize_header(row[0]) != "alias" or _normalize_header(row[1]) != "product":
            continue
        if first_identity_row_idx is None:
            first_identity_row_idx = row_idx

        field_row = grid[row_idx + 1]
        names_row = _forward_fill_row(row)
        groups = []
        col = 2
        while col + 5 <= len(field_row):
            labels = [_normalize_header(v) for v in field_row[col:col + 5]]
            if labels != target_fields:
                break
            cfa_name = names_row[col] if col < len(names_row) else None
            if cfa_name is None or str(cfa_name).strip() == "":
                break
            groups.append({"name": str(cfa_name).strip(), "closing_col": col + 2, "transit_col": col + 4})
            col += 5

        if groups:
            return row_idx, groups

    if first_identity_row_idx is not None:
        return first_identity_row_idx, []
    return None


def _load_pivoted_grid_rows(file_path: str, extension: str, sheet_name, skiprows: int):
    """Loads every data row (no header, raw values) below the two-row
    pivoted header -- mirrors `_validate_and_load`'s own per-extension
    engine selection, so this works across all SUPPORTED_EXTENSIONS."""
    if extension == ".csv":
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                return pd.read_csv(file_path, header=None, skiprows=skiprows, dtype=object, encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise UnicodeDecodeError("csv", b"", 0, 1, "unable to decode CSV with utf-8-sig or latin-1")

    engine = "xlrd" if extension == ".xls" else "openpyxl"
    return pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=skiprows, engine=engine, dtype=object)


def validate_pivoted_inventory_report(file_path: str, progress_callback=None) -> dict:
    """Validates and unpivots the new pivoted-by-CFA Inventory Report
    format: two identity columns (Alias, Product), then one repeating
    5-column group per CFA/branch -- Net Sales, Net Sales Value, Closing,
    Closing Value, Transit -- with the CFA name as a merged header cell
    spanning its group. CFA groups are detected from the header, never
    hardcoded, so adding/removing branches needs no code change.

    Net Sales / Net Sales Value / Closing Value are located but
    intentionally never used: this report is the DAILY/current-month
    upload, and the replenishment threshold is decided exclusively by
    the separate Previous Month Sales Report upload (see
    app/threshold_service.py) -- never by this report's own sales
    figures, even though it has a same-named "Net Sales" column.

    Unpivots into the same flat shape the replenishment engine already
    expects: one row per (CFA, product) with BranchLocation, Item Code,
    Item Name, TotalQty (= Closing), Transit Stock (= Transit). Note
    "Item Group" is not produced here -- this format doesn't carry it at
    all -- app/replenishment_service.py sources it from the matched
    threshold record instead, which is guaranteed present by the time
    it's needed.

    Never raises -- every failure mode is reported via the returned dict.
    Returns the same shape as `_validate_and_load`: {success, df,
    missing_columns, sheet_name, header_row_index, detected_columns,
    error}, so the upload page's existing success/failure branches work
    unmodified.
    """
    def report(percent: float, message: str) -> None:
        if progress_callback:
            progress_callback(percent, message)

    logger.info(f"Validating pivoted Inventory Report: {file_path}")
    report(5, "Initializing...")

    extension = Path(file_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        error = f"Unsupported file type: '{extension}'"
        logger.error(f"Inventory Report '{file_path}': {error}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": error,
        }

    try:
        report(15, "Reading workbook...")
        found = None
        identity_seen = False  # an Alias/Product row was found on some sheet, even if no valid CFA group followed it
        malformed_sheet = None
        sheets_scanned = 0
        for sheet_name, grid in _iter_candidate_sheets(file_path, extension):
            sheets_scanned += 1
            detection = _detect_pivot_groups(grid)
            if detection is None:
                continue
            row_idx, groups = detection
            if groups:
                found = (sheet_name, row_idx, groups)
                break
            identity_seen = True
            malformed_sheet = sheet_name
    except Exception as exc:
        logger.error(f"Failed to open Inventory Report '{file_path}': {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": str(exc),
        }

    if found is None:
        if identity_seen:
            # Clearly an attempt at the pivoted format (Alias/Product identity
            # columns present) but no valid CFA group followed -- a real,
            # pivot-specific error, not "this isn't a pivoted file" (which
            # would incorrectly trigger the old flat-format fallback and show
            # a misleading missing-columns message instead).
            error = (
                "Could not detect any CFA/branch column groups. Each branch must have exactly 5 "
                "columns in this order: Net Sales, Net Sales Value, Closing, Closing Value, Transit."
            )
            logger.warning(f"Inventory Report '{file_path}': {error} (sheet '{malformed_sheet}')")
            return {
                "success": False, "df": None, "missing_columns": [], "sheet_name": malformed_sheet,
                "header_row_index": None, "detected_columns": [], "error": error,
            }
        logger.info(f"Inventory Report '{file_path}': no pivoted CFA header found on any sheet")
        return {
            "success": False, "df": None,
            "missing_columns": PIVOT_IDENTITY_LABELS + PIVOT_FIELD_LABELS,
            "sheet_name": None, "header_row_index": None, "detected_columns": [], "error": None,
        }

    sheet_name, identity_row_idx, groups = found
    field_row_idx = identity_row_idx + 1

    try:
        report(35, "Processing data...")
        raw = _load_pivoted_grid_rows(file_path, extension, sheet_name, skiprows=field_row_idx + 1)
    except Exception as exc:
        logger.error(f"Failed to load Inventory Report '{file_path}' after header detection: {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": sheet_name,
            "header_row_index": field_row_idx, "detected_columns": [], "error": str(exc),
        }

    report(65, "Unpivoting CFA groups...")
    records = []
    blank_alias_skipped = 0
    for _, row in raw.iterrows():
        item_code = row[0]
        if pd.isna(item_code) or str(item_code).strip() == "":
            blank_alias_skipped += 1
            continue
        item_name = row[1]
        for group in groups:
            closing = row[group["closing_col"]] if group["closing_col"] < len(row) else None
            transit = row[group["transit_col"]] if group["transit_col"] < len(row) else None
            records.append(
                {
                    "BranchLocation": group["name"],
                    "Item Code": item_code,
                    "Item Name": item_name,
                    "TotalQty": closing,
                    "Transit Stock": transit,
                }
            )

    df = pd.DataFrame.from_records(
        records, columns=["BranchLocation", "Item Code", "Item Name", "TotalQty", "Transit Stock"]
    )

    if not records:
        if len(raw) == 0:
            reason = "the data section below the header was completely empty (0 rows)"
        elif blank_alias_skipped == len(raw):
            reason = (
                f"every one of the {len(raw)} raw row(s) had a blank/missing Alias (Item Code) value, "
                "so all were skipped before any record could be generated"
            )
        else:
            reason = "an unexpected combination -- groups were detected but no records were produced"
        logger.warning(f"Inventory Report '{file_path}': 0 records created after unpivoting -- {reason}.")

    report(95, "Finalizing...")
    logger.info(
        f"Inventory Report validated successfully: sheet '{sheet_name}', {len(groups)} CFA group(s) detected, "
        f"{len(raw)} product row(s), {len(df)} unpivoted record(s)"
    )
    report(100, "Done")
    return {
        "success": True, "df": df, "missing_columns": [], "sheet_name": sheet_name,
        "header_row_index": field_row_idx, "detected_columns": df.columns.tolist(), "error": None,
    }


_NEW_FORMAT_CWH_GROUP_PATTERN = re.compile(r"\bcwh\b|\bcentral\s+warehouse\b")

# Synthesized BranchLocation for every row this parser produces from the
# new format's CWH column group. Deliberately chosen so
# app.threshold_service.is_ahmedabad_cwh_branch (which matches "ahmedabad"
# followed by "cwh"/"central warehouse"/"head office") keeps recognizing
# these translated rows exactly like it already recognizes the flat/
# pivoted formats' own real "AHMEDABAD (HEAD OFFICE)"-style CWH rows --
# app.replenishment_service.evaluate_replenishment excludes them, and
# app.cwh_service.evaluate_cwh_stock reads them, with ZERO changes to
# either function. CWH is not itself a CFA and this label is never shown
# to a user (BranchLocation only reaches the UI for real CFA rows, which
# this label never is) -- it exists purely as the internal signal the
# existing exclusion logic already keys off.
NEW_FORMAT_CWH_BRANCH_LABEL = "AHMEDABAD (CWH)"


def _is_new_format_cwh_group(name) -> bool:
    """True if a new-format warehouse column group's own header name (not
    a full BranchLocation string -- just "CWH", "Central Warehouse", etc.)
    refers to the Head Warehouse itself, as opposed to Ahmedabad (which
    also has a single Closing-only column in this format but is a normal
    CFA, not CWH -- see `_detect_new_format_groups`'s docstring)."""
    return bool(_NEW_FORMAT_CWH_GROUP_PATTERN.search(_normalize_header(name)))


def _detect_new_format_groups(grid: list):
    """Scans `grid` for the new one-row-per-product Inventory Report
    header (Version 2.0, Milestone 57): a row whose first cell normalizes
    to 'product' -- NOT the two-column 'Alias'/'Product' identity row of
    the older pivoted format (see `_detect_pivot_groups`), which never
    matches here since its own first cell normalizes to 'alias' -- followed
    immediately by a row of Closing/Transit sub-labels.

    Unlike `_detect_pivot_groups`'s fixed 5-column groups, each warehouse's
    column span here is variable and read directly from how many
    consecutive columns share the same (forward-filled) group name in the
    identity row: CWH and Ahmedabad each carry exactly one column
    (Closing only), every other CFA carries two (Closing, then Transit).
    Distinguishing "this 1-column group is CWH" from "this 1-column group
    is Ahmedabad" is done by name (`_is_new_format_cwh_group`), not width
    -- both are structurally identical single-Closing-column groups.

    Same three-outcome contract as `_detect_pivot_groups`: (row_idx,
    groups) on a full match, (row_idx, []) if a 'Product' identity row was
    found but no valid group followed it (a real, format-specific error --
    the caller must not silently fall back to a different format's
    missing-columns message), or None if no such identity row exists
    anywhere on this sheet."""
    for row_idx in range(len(grid) - 1):
        row = grid[row_idx]
        if not row or _normalize_header(row[0]) != "product":
            continue

        field_row_idx = row_idx + 1
        if field_row_idx >= len(grid):
            return row_idx, []
        field_row = grid[field_row_idx]
        names_row = _forward_fill_row(row)

        groups = []
        col = 1
        malformed = False
        while col < len(field_row):
            name = names_row[col] if col < len(names_row) else None
            if name is None or str(name).strip() == "":
                break

            span_end = col
            while span_end < len(names_row) and names_row[span_end] == name:
                span_end += 1
            width = span_end - col
            labels = [_normalize_header(v) for v in field_row[col:span_end]]

            if width == 1 and labels == ["closing"]:
                groups.append({"name": str(name).strip(), "closing_col": col, "transit_col": None})
            elif width == 2 and labels == ["closing", "transit"]:
                groups.append({"name": str(name).strip(), "closing_col": col, "transit_col": col + 1})
            else:
                malformed = True
                break
            col = span_end

        if malformed or not groups:
            return row_idx, []
        return row_idx, groups

    return None


def validate_new_format_inventory_report(file_path: str, progress_callback=None) -> dict:
    """Validates and unpivots the new one-row-per-product Inventory Report
    format (Version 2.0, Milestone 57 -- Inventory Format Migration): a
    single 'Product' identity column, followed by one column group per
    warehouse. CWH and Ahmedabad each carry a single 'Closing' column;
    every other CFA carries 'Closing' then 'Transit' (see
    `_detect_new_format_groups`).

    This function is a pure translation layer -- it does not change, and
    is not used by, any business logic. CWH's column is translated into a
    row whose BranchLocation is `NEW_FORMAT_CWH_BRANCH_LABEL`
    ("AHMEDABAD (CWH)"), so app.replenishment_service.evaluate_replenishment
    (excludes it, being the Head Warehouse, not a CFA) and
    app.cwh_service.evaluate_cwh_stock (reads it) keep behaving EXACTLY as
    they already do for the flat/pivoted formats' own CWH rows -- neither
    function required any change. Ahmedabad's own column passes through
    with its real name unchanged, so it is evaluated as an ordinary CFA;
    its Transit Stock is always 0 (this format carries no Transit column
    for it at all -- there is nothing to read). Every other CFA's Transit
    Stock is its real value from the file. Effective Available Stock
    (Closing + Transit) is computed downstream, unchanged, exactly as for
    every other Inventory Report format.

    Unpivots into the exact same flat shape every other Inventory Report
    format already produces: one row per (warehouse, product) with
    BranchLocation, Item Code (always None -- this format has no item
    code/alias column at all), Item Name, TotalQty (= Closing), Transit
    Stock (= 0 for single-column groups, the real value otherwise).
    evaluate_replenishment()/evaluate_cwh_stock() see this DataFrame and
    cannot tell it apart from one produced by any older format.

    Never raises. Returns the same {success, df, missing_columns,
    sheet_name, header_row_index, detected_columns, error} shape as
    `validate_pivoted_inventory_report` / `_validate_and_load`.
    """
    def report(percent: float, message: str) -> None:
        if progress_callback:
            progress_callback(percent, message)

    logger.info(f"Validating new-format Inventory Report: {file_path}")
    report(5, "Initializing...")

    extension = Path(file_path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        error = f"Unsupported file type: '{extension}'"
        logger.error(f"Inventory Report '{file_path}': {error}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": error,
        }

    try:
        report(15, "Reading workbook...")
        found = None
        identity_seen = False
        malformed_sheet = None
        for sheet_name, grid in _iter_candidate_sheets(file_path, extension):
            detection = _detect_new_format_groups(grid)
            if detection is None:
                continue
            row_idx, groups = detection
            if groups:
                found = (sheet_name, row_idx, groups)
                break
            identity_seen = True
            malformed_sheet = sheet_name
    except Exception as exc:
        logger.error(f"Failed to open Inventory Report '{file_path}': {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": None,
            "header_row_index": None, "detected_columns": [], "error": str(exc),
        }

    if found is None:
        if identity_seen:
            # Clearly an attempt at the new format ('Product' identity
            # column present) but no valid warehouse group followed it --
            # a real, format-specific error, not "this isn't the new
            # format" (which would incorrectly fall through to the older
            # pivoted/flat formats and show a misleading error instead).
            error = (
                "Could not detect any warehouse column groups under the 'Product' header. Each "
                "warehouse must have either a single 'Closing' column (CWH, Ahmedabad) or "
                "'Closing' followed by 'Transit' (every other CFA)."
            )
            logger.warning(f"Inventory Report '{file_path}': {error} (sheet '{malformed_sheet}')")
            return {
                "success": False, "df": None, "missing_columns": [], "sheet_name": malformed_sheet,
                "header_row_index": None, "detected_columns": [], "error": error,
            }
        logger.info(f"Inventory Report '{file_path}': no new-format 'Product' header found on any sheet")
        return {
            "success": False, "df": None, "missing_columns": ["Product"],
            "sheet_name": None, "header_row_index": None, "detected_columns": [], "error": None,
        }

    sheet_name, identity_row_idx, groups = found
    field_row_idx = identity_row_idx + 1

    try:
        report(35, "Processing data...")
        raw = _load_pivoted_grid_rows(file_path, extension, sheet_name, skiprows=field_row_idx + 1)
    except Exception as exc:
        logger.error(f"Failed to load Inventory Report '{file_path}' after header detection: {exc}")
        return {
            "success": False, "df": None, "missing_columns": [], "sheet_name": sheet_name,
            "header_row_index": field_row_idx, "detected_columns": [], "error": str(exc),
        }

    report(65, "Unpivoting warehouse groups...")
    records = []
    blank_product_skipped = 0
    for _, row in raw.iterrows():
        item_name = row[0]
        if pd.isna(item_name) or str(item_name).strip() == "":
            blank_product_skipped += 1
            continue
        for group in groups:
            closing = row[group["closing_col"]] if group["closing_col"] < len(row) else None
            transit_col = group["transit_col"]
            transit = row[transit_col] if transit_col is not None and transit_col < len(row) else 0
            branch_location = (
                NEW_FORMAT_CWH_BRANCH_LABEL if _is_new_format_cwh_group(group["name"]) else group["name"]
            )
            records.append(
                {
                    "BranchLocation": branch_location,
                    "Item Code": None,
                    "Item Name": item_name,
                    "TotalQty": closing,
                    "Transit Stock": transit,
                }
            )

    df = pd.DataFrame.from_records(
        records, columns=["BranchLocation", "Item Code", "Item Name", "TotalQty", "Transit Stock"]
    )

    if not records:
        if len(raw) == 0:
            reason = "the data section below the header was completely empty (0 rows)"
        elif blank_product_skipped == len(raw):
            reason = (
                f"every one of the {len(raw)} raw row(s) had a blank/missing Product value, "
                "so all were skipped before any record could be generated"
            )
        else:
            reason = "an unexpected combination -- groups were detected but no records were produced"
        logger.warning(f"Inventory Report '{file_path}': 0 records created after unpivoting -- {reason}.")

    report(95, "Finalizing...")
    logger.info(
        f"Inventory Report validated successfully (new format): sheet '{sheet_name}', "
        f"{len(groups)} warehouse group(s) detected, {len(raw)} product row(s), {len(df)} unpivoted record(s)"
    )
    report(100, "Done")
    return {
        "success": True, "df": df, "missing_columns": [], "sheet_name": sheet_name,
        "header_row_index": field_row_idx, "detected_columns": df.columns.tolist(), "error": None,
    }


def validate_inventory_report(file_path: str, progress_callback=None) -> dict:
    """Validate an uploaded Inventory Report workbook.

    Tries formats newest-first, each a pure translation layer into the
    same flat shape (BranchLocation, Item Code, Item Name, TotalQty,
    Transit Stock) that app.replenishment_service.evaluate_replenishment
    and app.cwh_service.evaluate_cwh_stock already expect -- neither of
    those functions, nor anything else downstream, changes based on which
    format actually produced that shape:

    1. The new one-row-per-product format (see
       `validate_new_format_inventory_report`) -- the current ERP export
       shape as of Version 2.0 Milestone 57.
    2. If not found, the older pivoted-by-CFA format (see
       `validate_pivoted_inventory_report`) -- one repeating 5-column
       group per branch.
    3. If neither is found, the original flat-row format
       (INVENTORY_REPORT_REQUIRED_COLUMNS, via `_validate_and_load`).

    This ordering means older-style exports (or a transition period where
    multiple shapes are still in circulation) keep working without any
    user-visible change -- each tier's own real, format-specific error
    (as opposed to "this isn't that format") is returned immediately
    without falling through to a different tier's misleading message.
    """
    new_format_result = validate_new_format_inventory_report(file_path, progress_callback)
    if new_format_result["success"]:
        return new_format_result
    if new_format_result["error"] is not None:
        return new_format_result

    pivoted_result = validate_pivoted_inventory_report(file_path, progress_callback)
    if pivoted_result["success"]:
        return pivoted_result
    if pivoted_result["error"] is not None:
        # A real, pivot-specific error (not just "no pivoted header found")
        # -- never fall back to the flat-row format here, since that would
        # misreport a genuinely broken pivoted file as a missing-columns
        # error in the old format instead of surfacing the real problem.
        return pivoted_result

    # Falling back to the old flat-row format: if the uploaded file is
    # actually one of the newer formats but malformed in some way not
    # caught above, this fallback will also fail (it requires completely
    # different column headers) -- that combined failure is what used to
    # produce a silent "0 Products Evaluated" if the UI didn't surface the
    # flat-format failure message clearly.
    logger.info(f"No pivoted CFA header found in '{file_path}', falling back to flat-row format")
    return _validate_and_load(file_path, INVENTORY_REPORT_REQUIRED_COLUMNS, "Inventory Report", progress_callback)


def load_sales_report(file_path: str, progress_callback=None) -> dict:
    """Validate an uploaded Previous Month Sales Report workbook against
    SALES_REPORT_REQUIRED_COLUMNS. See `_validate_and_load` for the exact
    behavior and returned dict shape."""
    return _validate_and_load(file_path, SALES_REPORT_REQUIRED_COLUMNS, "Previous Month Sales Report", progress_callback)


def load_historical_payment_report(file_path: str, progress_callback=None) -> dict:
    """Validate an uploaded Historical Payment Report workbook against
    HISTORICAL_PAYMENT_REPORT_REQUIRED_COLUMNS. See `_validate_and_load` for
    the exact behavior and returned dict shape -- this is upload/validation
    only, no payment calculation, customer classification, or rolling-window
    logic happens here."""
    return _validate_and_load(
        file_path, HISTORICAL_PAYMENT_REPORT_REQUIRED_COLUMNS, "Historical Payment Report", progress_callback
    )


def load_monthly_payment_report(file_path: str, progress_callback=None) -> dict:
    """Validate an uploaded Monthly Payment Report workbook -- same
    required columns as the Historical Payment Report (same source
    system, same shape), just a distinct report_label for clearer logs.
    See `_validate_and_load` for the exact behavior and returned dict
    shape; rolling-window logic (which month this is, whether it can be
    accepted) happens downstream in app/payment_analytics_service.py."""
    return _validate_and_load(
        file_path, HISTORICAL_PAYMENT_REPORT_REQUIRED_COLUMNS, "Monthly Payment Report", progress_callback
    )


def load_outstanding_report(file_path: str, progress_callback=None) -> dict:
    """Validate an uploaded Outstanding Report workbook against
    OUTSTANDING_REPORT_REQUIRED_COLUMNS -- the Collections Action
    Center's daily working-set upload. Column names match the real
    production report exactly (Division, H.Q., Bill No., L.R. Date, Bill
    Amount, Month) -- see `_normalize_header` for how case/spacing
    differences are still tolerated on top of that. See
    `_validate_and_load` for the exact behavior and returned dict shape;
    row-level parsing and Days Outstanding/Days Over Due/Status
    computation happen downstream in app/collections_service.py."""
    return _validate_and_load(
        file_path, OUTSTANDING_REPORT_REQUIRED_COLUMNS, "Outstanding Report", progress_callback
    )
