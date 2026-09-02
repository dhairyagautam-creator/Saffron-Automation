"""Parses the Organization Data workbooks (one per division: Onyx,
Guardians, Xandra) into the normalized `employee_hierarchy` table.

Each workbook is a flat list of employees — one row per employee, in
whatever order the division keeps its org chart — with columns for
Employee Code, Employee Name, Designation, Mobile Number, and Email
Address already built in. There is no separate email directory anymore:
every employee's own email lives on their own row here.

The reporting hierarchy is inferred purely from ROW ORDER, top to bottom,
per sheet:

- An RBM row starts a new reporting section: it becomes the "current RBM",
  and resets the "current ABM" (so any BM appearing before that section's
  first ABM is correctly left without an ABM rather than inheriting one
  from the previous RBM's section).
- An ABM row belongs to the current RBM, and itself becomes the "current
  ABM" for any BM rows that follow.
- A BM row belongs to both the current ABM and the current RBM.
- Any other designation (AGM, SM, GM, ...) is imported as a row but does
  not participate in — or change — the current RBM/ABM tracking.

Rows whose employee name contains "Vacant" (case-insensitive) are skipped
entirely: no employee record, no hierarchy relationship, no email mapping.
A vacant RBM/ABM row DOES reset the corresponding current-RBM/current-ABM
tracker to "nobody" -- so a BM/ABM appearing under a vacant position is
correctly left without that rung (and falls through to the next rung up via
app.hierarchy_service) rather than silently inheriting whichever ABM/RBM
happened to hold that slot in the previous section.

The final "Senior" every employee reports to -- after applying the fixed
per-designation fallback chain (e.g. BM -> RBM -> Senior RBM -> SM -> AGM ->
GM, deliberately skipping ABM -- see app.hierarchy_service.FALLBACK_CHAINS,
redesigned 2026-07-28; each rung skipped if vacant/missing/no-email) -- is
computed once per refresh by app.hierarchy_service.compute_seniors() and
stored as senior_code/senior_name/senior_email/senior_designation on every
row. Nothing else in the app re-derives this independently; see that
module's docstring for the full rule.
"""

import re
import warnings

import openpyxl
import pandas as pd
from loguru import logger
from sqlalchemy import inspect, text

from app.hierarchy_service import compute_seniors
from app.workbook_connections import WORKBOOK_NAMES, get_connections, get_status
from database.connection import get_data_engine

HIERARCHY_TABLE = "employee_hierarchy"

HIERARCHY_COLUMNS = [
    "employee_code",
    "employee_name",
    "designation",
    "mobile",
    "email",
    # Date of Joining, normalized to an ISO 'YYYY-MM-DD' string (or "" if
    # blank/unparseable) at parse time -- see _doj_cell() below. The
    # canonical source app.doj_eligibility_service reads (via
    # get_all_doj()/get_doj_by_name() further down this file) to decide
    # whether Work Distribution (RGD Coverage + Manager Work Allocation)
    # should treat this employee as active during a given month.
    "doj",
    # Raw, per-employee direct assignments -- internal inputs to
    # app.hierarchy_service's fallback walk. Not displayed anywhere
    # (superseded by senior_* below); ABM/RBM are the only two rungs
    # carried this way, see app.hierarchy_service's module docstring.
    "abm_code",
    "abm_name",
    "rbm_code",
    "rbm_name",
    # The final resolved reporting senior after the fixed per-designation
    # fallback chain (e.g. BM -> RBM -> Senior RBM -> SM -> AGM -> GM,
    # never ABM -- see app.hierarchy_service.FALLBACK_CHAINS), computed
    # once by app.hierarchy_service.compute_seniors() in
    # refresh_hierarchy() below.
    # This -- not abm_name/rbm_name above -- is what the Organization Data
    # table displays and what the email notification system reads; both
    # read this exact column, so they can never disagree.
    "senior_code",
    "senior_name",
    "senior_email",
    "senior_designation",
    "division",
    "source_sheet",
]

# Designations that participate in the current-RBM/current-ABM tracking as
# section-starters. Everything else (AGM, SM, GM, ...) is imported but
# doesn't change — or inherit from — the tracked chain.
RBM_DESIGNATIONS = {"RBM"}
ABM_DESIGNATIONS = {"ABM"}
BM_DESIGNATIONS = {"BM"}

# Only these designations are ever eligible to be flagged by a rule — see
# rules/same_location.py's get_all_designations()-based filter.
ANALYZABLE_DESIGNATIONS = {"BM", "ABM"}

COLUMN_CANDIDATES = {
    "employee_code": ["Emp Code", "Employee Code", "Employee Code.", "Code"],
    "employee_name": ["Name", "Employee Name", "Emp Name"],
    "designation": ["Designation", "Desig.", "Desig"],
    "mobile": ["Mobile", "Mobile Number", "Phone", "Contact Number"],
    "email": ["Email-Id", "Email Id", "Email", "Email Address", "E-mail", "E-mail Address"],
    "doj": ["DOJ", "Date of Joining", "Date Of Joining", "Joining Date", "DOJ Date"],
}
REQUIRED_HEADER_FIELDS = ("employee_code", "employee_name", "designation")


def _normalize_header(value) -> str:
    """Lowercase and collapse underscores/hyphens/periods to spaces, so
    'Emp Code', 'emp_code', and 'Emp-Code' all match."""
    normalized = re.sub(r"[_\-.]+", " ", str(value).strip().lower()) if value is not None else ""
    return re.sub(r"\s+", " ", normalized).strip()


def _find_column(header_row: list, candidates: list) -> int | None:
    normalized = [_normalize_header(v) for v in header_row]
    for candidate in candidates:
        target = _normalize_header(candidate)
        if target in normalized:
            return normalized.index(target)
    return None


def _find_header(grid: list) -> tuple | None:
    """Scan `grid` for the header row, returning (row_index, {field: col_index})."""
    for row_idx, row in enumerate(grid):
        column_map = {}
        for field, candidates in COLUMN_CANDIDATES.items():
            col_idx = _find_column(row, candidates)
            if col_idx is not None:
                column_map[field] = col_idx

        if all(field in column_map for field in REQUIRED_HEADER_FIELDS):
            return row_idx, column_map

    return None


def _is_blank_row(row: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _cell(row: list, column_map: dict, field: str):
    idx = column_map.get(field)
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    return str(value).strip() if value is not None else None


def _is_vacant(employee_name: str | None) -> bool:
    return bool(employee_name) and "vacant" in employee_name.lower()


def _doj_cell(row: list, column_map: dict) -> str:
    """Like _cell(), but for the "doj" column specifically -- normalizes
    to an ISO 'YYYY-MM-DD' string instead of the generic str(value) every
    other field gets, since app.doj_eligibility_service needs to parse
    this back into a date on every read. An Excel date cell arrives here
    as a python date/datetime (openpyxl, data_only=True); a plain string
    cell (e.g. "15-04-2026", "15/04/2026") is tolerated too via pandas.
    dayfirst=True assumes the DD-MM-YYYY convention this company's other
    data already uses (see e.g. app.manager_work_allocation_shared's own
    "Feb-26" real-report precedent) when the day/month can't otherwise be
    told apart. "" if blank or unparseable -- never guessed at; that's
    treated identically to "no DOJ on file" downstream (see
    app.doj_eligibility_service's own docstring for why)."""
    idx = column_map.get("doj")
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    with warnings.catch_warnings():
        # Silences pandas' "dayfirst=True was specified" notice for the
        # (common, unambiguous) ISO-format cells -- dayfirst only ever
        # matters for the genuinely ambiguous DD/MM vs MM/DD case anyway.
        warnings.simplefilter("ignore", UserWarning)
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return "" if pd.isna(parsed) else parsed.date().isoformat()


def _parse_sheet(grid: list, workbook_name: str, sheet_name: str) -> tuple[list, dict]:
    """Walk one sheet top to bottom, tracking the current RBM/ABM as
    described in the module docstring. Returns (records, stats) where
    stats has `vacant_ignored`."""
    header = _find_header(grid)
    if header is None:
        return [], {"vacant_ignored": 0, "header_found": False}

    header_row_idx, column_map = header

    records = []
    vacant_ignored = 0
    current_rbm: dict | None = None
    current_abm: dict | None = None

    for row in grid[header_row_idx + 1 :]:
        if _is_blank_row(row):
            continue

        employee_code = _cell(row, column_map, "employee_code")
        employee_name = _cell(row, column_map, "employee_name")
        designation_raw = _cell(row, column_map, "designation")

        if not employee_code and not employee_name:
            continue  # stray formatting row, not data

        if _is_vacant(employee_name):
            vacant_ignored += 1
            vacant_designation = (designation_raw or "").strip().upper()
            # A vacant RBM/ABM row still resets the corresponding tracker
            # to "nobody" -- so anyone below it correctly falls through to
            # the next rung up (via app.hierarchy_service) instead of
            # silently inheriting whichever ABM/RBM held that slot in the
            # previous section. No record is created either way.
            if vacant_designation in RBM_DESIGNATIONS:
                current_rbm = None
                current_abm = None
            elif vacant_designation in ABM_DESIGNATIONS:
                current_abm = None
            continue

        designation = (designation_raw or "").strip().upper()
        mobile = _cell(row, column_map, "mobile")
        email = _cell(row, column_map, "email")
        doj = _doj_cell(row, column_map)

        if designation in RBM_DESIGNATIONS:
            current_rbm = {"code": employee_code, "name": employee_name}
            current_abm = None
            abm_code = abm_name = None
            rbm_code = rbm_name = None
        elif designation in ABM_DESIGNATIONS:
            rbm_code = current_rbm["code"] if current_rbm else None
            rbm_name = current_rbm["name"] if current_rbm else None
            abm_code = abm_name = None
            current_abm = {"code": employee_code, "name": employee_name}
        elif designation in BM_DESIGNATIONS:
            abm_code = current_abm["code"] if current_abm else None
            abm_name = current_abm["name"] if current_abm else None
            rbm_code = current_rbm["code"] if current_rbm else None
            rbm_name = current_rbm["name"] if current_rbm else None
        else:
            # AGM/SM/GM/etc. -- imported, but not part of the tracked chain.
            abm_code = abm_name = None
            rbm_code = rbm_name = None

        records.append(
            {
                "employee_code": employee_code,
                "employee_name": employee_name,
                "designation": designation,
                "mobile": mobile,
                "email": email,
                "doj": doj,
                "abm_code": abm_code,
                "abm_name": abm_name,
                "rbm_code": rbm_code,
                "rbm_name": rbm_name,
                "division": workbook_name,
                "source_sheet": sheet_name,
            }
        )

    return records, {"vacant_ignored": vacant_ignored, "header_found": True}


def parse_workbook(file_path: str, workbook_name: str) -> tuple:
    """Parse every sheet of one division's Organization Data workbook.

    Returns (rows, stats) where stats has `sheets_processed`,
    `sheets_skipped` (sheet_name -> reason), and `vacant_ignored`.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    all_rows = []
    sheets_processed = 0
    sheets_skipped = {}
    vacant_ignored = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]

        rows, stats = _parse_sheet(grid, workbook_name, sheet_name)
        vacant_ignored += stats["vacant_ignored"]

        if not stats["header_found"]:
            sheets_skipped[sheet_name] = "No recognizable Organization Data header found"
            continue
        if not rows:
            sheets_skipped[sheet_name] = "No data rows found"
            continue

        all_rows.extend(rows)
        sheets_processed += 1

    return all_rows, {
        "sheets_processed": sheets_processed,
        "sheets_skipped": sheets_skipped,
        "vacant_ignored": vacant_ignored,
    }


def _ensure_indexes() -> None:
    with get_data_engine().begin() as conn:
        conn.execute(
            text(f"CREATE INDEX IF NOT EXISTS idx_{HIERARCHY_TABLE}_code ON {HIERARCHY_TABLE}(employee_code)")
        )
        conn.execute(
            text(f"CREATE INDEX IF NOT EXISTS idx_{HIERARCHY_TABLE}_name ON {HIERARCHY_TABLE}(employee_name)")
        )


def refresh_hierarchy() -> dict:
    """Read every Connected Organization Data workbook, parse it, and
    rebuild the employee_hierarchy table. Workbooks that aren't configured
    (or whose file can't be found) are skipped silently — that's expected,
    not an error.

    Returns a dict with: `employees_loaded`, `total_bm`, `total_abm`,
    `total_rbm`, `vacant_ignored`, `emails_loaded`,
    `hierarchy_relationships`, `worksheets_processed`, `workbooks_read`,
    and `workbooks_skipped` (name -> reason, only for workbooks that were
    connected but failed to parse).
    """
    connections = get_connections(WORKBOOK_NAMES)

    all_rows = []
    workbooks_read = []
    workbooks_skipped = {}
    total_sheets_processed = 0
    vacant_ignored = 0

    for workbook_name, file_path in connections.items():
        if get_status(file_path) != "Connected":
            continue  # not configured yet — nothing to report

        try:
            rows, stats = parse_workbook(file_path, workbook_name)
        except Exception as exc:
            logger.error(f"Failed to parse Organization Data workbook '{workbook_name}' ({file_path}): {exc}")
            workbooks_skipped[workbook_name] = f"Error: {exc}"
            continue

        all_rows.extend(rows)
        workbooks_read.append(workbook_name)
        total_sheets_processed += stats["sheets_processed"]
        vacant_ignored += stats["vacant_ignored"]
        if stats["sheets_skipped"]:
            logger.warning(f"'{workbook_name}': skipped sheet(s) {stats['sheets_skipped']}")

    # Resolve every employee's final "Senior" (each designation's fixed
    # fallback chain -- e.g. BM -> RBM -> Senior RBM -> SM -> AGM -> GM,
    # never ABM) in one bulk, in-memory pass -- see
    # app.hierarchy_service.FALLBACK_CHAINS and its module docstring. This is the
    # ONE place the whole app computes this; everything else (the
    # Organization Data table, the email notification system) reads the
    # resulting senior_code/senior_name/senior_email/senior_designation
    # columns instead of re-deriving anything.
    compute_seniors(all_rows)

    hierarchy_df = (
        pd.DataFrame(all_rows)[HIERARCHY_COLUMNS] if all_rows else pd.DataFrame(columns=HIERARCHY_COLUMNS)
    )
    # Never let a missing value round-trip through SQLite/pandas as SQL
    # NULL -- read_sql_table() turns that into float NaN on the way back
    # out, which is exactly how "NaN" ended up displayed in the
    # Organization Data table (see ui/organization_data_page.py).
    hierarchy_df = hierarchy_df.fillna("")
    hierarchy_df.to_sql(HIERARCHY_TABLE, con=get_data_engine(), if_exists="replace", index=False)
    _ensure_indexes()

    total_bm = int((hierarchy_df["designation"] == "BM").sum()) if len(hierarchy_df) else 0
    total_abm = int((hierarchy_df["designation"] == "ABM").sum()) if len(hierarchy_df) else 0
    total_rbm = int((hierarchy_df["designation"] == "RBM").sum()) if len(hierarchy_df) else 0

    if len(hierarchy_df):
        has_email = hierarchy_df["email"].notna() & (hierarchy_df["email"].astype(str).str.strip() != "")
        emails_loaded = int(has_email.sum())

        bm_rows = hierarchy_df[hierarchy_df["designation"] == "BM"]
        abm_rows = hierarchy_df[hierarchy_df["designation"] == "ABM"]
        relationships = (
            int((bm_rows["abm_code"].notna() & (bm_rows["abm_code"].astype(str).str.strip() != "")).sum())
            + int((bm_rows["rbm_code"].notna() & (bm_rows["rbm_code"].astype(str).str.strip() != "")).sum())
            + int((abm_rows["rbm_code"].notna() & (abm_rows["rbm_code"].astype(str).str.strip() != "")).sum())
        )
    else:
        emails_loaded = 0
        relationships = 0

    logger.info(
        f"Refreshed '{HIERARCHY_TABLE}': {len(hierarchy_df)} employee(s) from "
        f"{len(workbooks_read)} workbook(s) across {total_sheets_processed} sheet(s); "
        f"{total_bm} BM, {total_abm} ABM, {total_rbm} RBM, {vacant_ignored} vacant ignored, "
        f"{emails_loaded} email(s) loaded, {relationships} hierarchy relationship(s)"
    )

    return {
        "employees_loaded": len(hierarchy_df),
        "total_bm": total_bm,
        "total_abm": total_abm,
        "total_rbm": total_rbm,
        "vacant_ignored": vacant_ignored,
        "emails_loaded": emails_loaded,
        "hierarchy_relationships": relationships,
        "worksheets_processed": total_sheets_processed,
        "workbooks_read": workbooks_read,
        "workbooks_skipped": workbooks_skipped,
    }


def find_by_employee_code(employee_code: str) -> dict | None:
    """Look up a single hierarchy entry by employee_code."""
    if not employee_code or not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return None
    with get_data_engine().connect() as conn:
        row = conn.execute(
            text(f"SELECT * FROM {HIERARCHY_TABLE} WHERE employee_code = :code LIMIT 1"),
            {"code": employee_code},
        ).mappings().first()
    return dict(row) if row else None


def find_by_employee_name(employee_name: str) -> list:
    """Look up hierarchy entries by employee_name (case/whitespace-insensitive;
    may return more than one row, e.g. the same manager appearing in several sheets)."""
    if not employee_name or not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return []
    with get_data_engine().connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM {HIERARCHY_TABLE} WHERE TRIM(LOWER(employee_name)) = :name"),
            {"name": employee_name.strip().lower()},
        ).mappings().all()
    return [dict(r) for r in rows]


def find_by_designation(division: str | None, source_sheet: str | None, designation: str) -> dict | None:
    """Look up the (first) hierarchy row with the given designation within
    the same division + source_sheet -- the org chart's own zone/region
    partition (the same scope the RBM/ABM chain itself resets at per sheet,
    see module docstring). Used by app.hierarchy_service's fallback walk to
    find the next rung above an unavailable RBM: SRRBM/SM/AGM/GM aren't
    individually assigned per BM/ABM the way RBM is, so this is a
    designation lookup scoped to the same org unit, not a direct field on
    the flagged employee's own row.

    Returns None if division/source_sheet aren't given, the table doesn't
    exist, or no row matches -- callers treat all of these as "this rung is
    unavailable", not an error."""
    if not division or not source_sheet or not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return None
    with get_data_engine().connect() as conn:
        row = conn.execute(
            text(
                f"SELECT * FROM {HIERARCHY_TABLE} WHERE division = :division AND source_sheet = :sheet "
                f"AND UPPER(designation) = :designation LIMIT 1"
            ),
            {"division": division, "sheet": source_sheet, "designation": designation.upper()},
        ).mappings().first()
    return dict(row) if row else None


def get_all_designations() -> dict[str, str]:
    """Return {employee_code: designation} for every hierarchy row, for the
    rule engine's bulk BM/ABM-only eligibility filter (see
    rules/same_location.py) — one query instead of one per employee/day
    group."""
    if not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return {}
    with get_data_engine().connect() as conn:
        rows = conn.execute(
            text(f"SELECT employee_code, designation FROM {HIERARCHY_TABLE} WHERE employee_code IS NOT NULL")
        ).all()
    return {code: designation for code, designation in rows}


def get_all_doj() -> dict[str, str]:
    """Return {employee_code: doj} (the ISO 'YYYY-MM-DD' string stored by
    _doj_cell() above) for every hierarchy row with both an employee_code
    and a non-blank doj -- bulk, one query, mirroring
    get_all_designations()'s own "one query instead of one per employee"
    reasoning. See app.doj_eligibility_service for how this is actually
    parsed and applied."""
    if not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return {}
    with get_data_engine().connect() as conn:
        rows = conn.execute(
            text(
                f"SELECT employee_code, doj FROM {HIERARCHY_TABLE} "
                "WHERE employee_code IS NOT NULL AND employee_code != '' "
                "AND doj IS NOT NULL AND doj != ''"
            )
        ).all()
    return {code: doj for code, doj in rows}


def get_doj_by_name() -> dict[str, str]:
    """Return {TRIM(LOWER(employee_name)): doj} for every hierarchy row
    with a non-blank doj -- name-keyed, for callers (RGD Coverage) whose
    own uploaded rows only ever carry an employee's NAME, never a code.
    First-match-wins per name, same convention as
    app.hierarchy_service.build_lookup_maps."""
    if not inspect(get_data_engine()).has_table(HIERARCHY_TABLE):
        return {}
    with get_data_engine().connect() as conn:
        rows = conn.execute(
            text(
                f"SELECT employee_name, doj FROM {HIERARCHY_TABLE} "
                "WHERE employee_name IS NOT NULL AND employee_name != '' "
                "AND doj IS NOT NULL AND doj != ''"
            )
        ).all()
    result: dict[str, str] = {}
    for name, doj in rows:
        key = name.strip().lower()
        result.setdefault(key, doj)
    return result
