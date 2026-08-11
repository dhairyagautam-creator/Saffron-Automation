"""Version 2.0's Table Export Framework -- the ONE reusable "export
exactly what's on screen" skeleton for the whole application.

Reference implementation: ui/inventory_replenishment_page.py's Export
button. Every future page's export is a THIN wrapper around
export_rows_with_ui() below -- the page's own code only ever supplies:

  - rows: the CURRENT, already-filtered/searched/sorted list of dicts
    exactly as it appears on screen right now. This must be the SAME list
    the page's own render function just used to populate its Treeview --
    never re-queried, never re-filtered here. That single rule is what
    guarantees the exported workbook is byte-for-byte what the user is
    currently looking at: if they've filtered to one CFA, searched one
    SKU, or applied any sort/dropdown the page supports, the exported
    rows already reflect it, because they ARE the rows the table is
    rendering, nothing else. If a page later adds a search box or
    column-click sorting, the export needs zero changes -- it will
    automatically capture the new filter/sort state the moment the page
    starts including it in the list it hands the framework.
  - columns: the ordered tuple of dict keys to include, in display order
    -- normally the exact same COLUMNS tuple the page's own Treeview
    already uses.
  - headings: {column_key: display_header_text} -- normally the page's
    own HEADINGS dict, unchanged.
  - a suggested export filename.

Everything else -- workbook creation, header styling, column
autosizing, the Save dialog, running the write on a background thread so
a large table never freezes the UI, and (optionally) mirroring a row's
on-screen highlight color into the exported cell fill -- lives here,
once, so no future page ever re-implements any of it.

Adopting this for a new page (worked example, Thresholds):

    from app.table_export_service import default_export_filename, export_rows_with_ui

    def _on_export_clicked(self) -> None:
        export_rows_with_ui(
            self,
            rows=self._current_display_rows,   # exactly what _render_table() just drew
            columns=COLUMNS,
            headings=HEADINGS,
            suggested_filename=default_export_filename("Thresholds"),
        )

That's the whole integration. No new dialog code, no new threading code,
no new Excel-writing code -- ever.

Two layers, so both a human clicking "Export" and a future automated
pipeline get the identical output:

  - write_rows_to_excel() -- pure, synchronous, no UI at all. Takes a
    plain file path and writes the workbook. This is the ONE function
    that actually builds a .xlsx file anywhere in the application; every
    other entry point below is a thin UI wrapper around it.
  - export_rows_with_ui() -- the button-handler convenience wrapper:
    prompts for a save location, runs write_rows_to_excel() on a
    background thread (see ui/background_task.run_in_background) behind
    a LoadingOverlay so a large table never freezes the window, then
    reports success/failure.

Future-workflow compatibility (explicitly not built yet, only kept
possible): the eventual automated pipeline

    Inventory Upload -> Replenishment Generated -> Export Framework -> Excel File -> Email Attachment

will call write_rows_to_excel() directly, with no dialog and no
threading (there's no user present to click Save, and the caller already
controls its own thread) -- passing it the exact same `rows`/`columns`/
`headings` the Replenishment page itself would use, at whatever moment
the automation runs. Because it is the SAME function the Export button
calls, the emailed workbook and the one a user would download by
clicking Export can never drift apart into two different
implementations.
"""

import os
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from tkinter import filedialog, messagebox
from typing import Callable, Optional

import openpyxl
from loguru import logger
from openpyxl.styles import Alignment, Font as ExcelFont, PatternFill
from openpyxl.utils import get_column_letter

from app.config import REPORTS_DIR
from ui.background_task import run_in_background
from ui.loading_overlay import LoadingOverlay
from ui.theme import Color

# Brand-consistent header styling -- sampled from the same tokens every
# page's UI already uses (ui/theme.py), so an exported workbook visually
# matches the app rather than looking like a generic dump.
_HEADER_HEX = Color.PRIMARY.lstrip("#")
_HEADER_FILL = PatternFill(start_color=_HEADER_HEX, end_color=_HEADER_HEX, fill_type="solid")
_HEADER_FONT = ExcelFont(bold=True, color="FFFFFF")
_HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center")

_MAX_COLUMN_WIDTH = 60
_MIN_COLUMN_WIDTH = 8


@dataclass(frozen=True)
class RowStyle:
    """Optional per-row visual style for an exported row -- the export
    equivalent of a Treeview `tag_configure()` background/foreground.
    Every field is optional; omit whichever a given table doesn't need.

    fill_color/font_color are hex strings WITHOUT a leading '#' (openpyxl's
    own convention, e.g. "FFC7CE") -- pass a ui.theme.Color constant with
    `.lstrip("#")` applied, exactly as this module does for its own header
    styling above, so an exported row's colors are pixel-identical to the
    on-screen highlight it mirrors."""

    fill_color: Optional[str] = None
    font_color: Optional[str] = None
    bold: bool = False


@dataclass
class ExportResult:
    """Outcome of one export -- returned by write_rows_to_excel() and
    passed to export_rows_with_ui()'s optional on_done callback."""

    success: bool
    file_path: Optional[str] = None
    rows_written: int = 0
    error_message: Optional[str] = None


def default_export_filename(base_name: str, *, suffix: str = "") -> str:
    """A sensible default filename: "{base_name}{_suffix if any}_{today's
    date}.xlsx", with anything unsafe for a filename stripped out of
    both `base_name` and `suffix` (e.g. a CFA name containing '/' or
    other punctuation from real data). Callers with active filters worth
    reflecting in the filename (e.g. a selected CFA) can pass it as
    `suffix`; callers with no such distinction just omit it."""
    parts = [base_name] + ([suffix] if suffix else [])
    safe = "_".join(re.sub(r"[^A-Za-z0-9]+", "", p) or "" for p in parts if p)
    date_stamp = datetime.now().strftime("%Y-%m-%d")
    return f"{safe}_{date_stamp}.xlsx"


def write_rows_to_excel(
    rows: list[dict],
    columns: tuple,
    headings: dict,
    file_path: str,
    *,
    sheet_title: str = "Sheet1",
    row_style_fn: Optional[Callable[[dict], Optional[RowStyle]]] = None,
    progress_callback: Optional[Callable[[int, str], None]] = None,
) -> ExportResult:
    """The ONE canonical writer -- pure and synchronous, no UI, no
    threading, no dialog. Builds a single-sheet workbook: one header row
    (bold, brand-colored, per `headings`) followed by one row per item in
    `rows`, reading each row's values in `columns` order. Column widths
    are set from the longest value actually present in that column
    (capped, so one huge outlier cell can't make a column absurdly wide).

    `row_style_fn(row) -> RowStyle | None`, if given, is called once per
    row; a returned RowStyle's fill/font color is applied to every cell
    in that row -- this is how a page mirrors an on-screen row highlight
    (e.g. Replenishment's CWH-shortage red) into the exported file. A
    table with no such concept simply omits this argument.

    `progress_callback(percent, message)`, if given, is called a handful
    of times while writing -- useful for a future very large table
    driving a LoadingOverlay's progress bar; harmless to omit for a
    table small enough to write instantly.

    Never raises for an ordinary write failure (permission denied, disk
    full, the file open elsewhere) -- returns ExportResult(success=False,
    error_message=...) instead, exactly like this codebase's other
    file-producing operations (see app/sync_service.py's RowsSyncResult
    convention) report failure as data, not an exception the caller must
    remember to catch."""
    try:
        if progress_callback:
            progress_callback(5, "Building workbook...")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_title[:31] or "Sheet1"  # Excel's own 31-char sheet-name limit

        column_widths = [len(str(headings.get(col, col))) for col in columns]

        for col_index, col_key in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_index, value=headings.get(col_key, col_key))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
        sheet.freeze_panes = "A2"

        total = len(rows)
        for row_index, row in enumerate(rows, start=2):
            style = row_style_fn(row) if row_style_fn else None
            for col_index, col_key in enumerate(columns, start=1):
                value = row.get(col_key, "")
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                if style is not None:
                    if style.fill_color:
                        cell.fill = PatternFill(
                            start_color=style.fill_color, end_color=style.fill_color, fill_type="solid"
                        )
                    if style.font_color or style.bold:
                        cell.font = ExcelFont(color=style.font_color or "000000", bold=style.bold)
                column_widths[col_index - 1] = max(column_widths[col_index - 1], len(str(value)))

            if progress_callback and total and (row_index % 200 == 0):
                percent = 5 + int(85 * (row_index - 1) / total)
                progress_callback(percent, f"Writing row {row_index - 1} of {total}...")

        for col_index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = max(
                _MIN_COLUMN_WIDTH, min(_MAX_COLUMN_WIDTH, width + 2)
            )

        if progress_callback:
            progress_callback(95, "Saving file...")

        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        workbook.save(file_path)

        if progress_callback:
            progress_callback(100, "Done")

        logger.info(f"Export: wrote {total} row(s) to {file_path!r}")
        return ExportResult(success=True, file_path=file_path, rows_written=total)
    except Exception as exc:
        logger.error(f"Export: failed to write {file_path!r}: {exc!r}")
        return ExportResult(success=False, file_path=file_path, error_message=str(exc))


def prompt_save_path(suggested_filename: str, *, parent=None) -> Optional[str]:
    """Native Save dialog seeded with REPORTS_DIR (app/config.py -- an
    always-writable, per-user folder already created at startup for
    exactly this purpose, previously unused) and the suggested filename.
    Returns None if the user cancels. Must be called on the main thread
    (Tkinter dialogs are not thread-safe) -- always call this BEFORE
    handing off to a background thread, never from inside one."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return filedialog.asksaveasfilename(
        parent=parent,
        title="Export to Excel",
        initialdir=str(REPORTS_DIR),
        initialfile=suggested_filename,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
    ) or None


def export_rows_with_ui(
    widget,
    rows: list[dict],
    columns: tuple,
    headings: dict,
    suggested_filename: str,
    *,
    sheet_title: str = "Sheet1",
    row_style_fn: Optional[Callable[[dict], Optional[RowStyle]]] = None,
    on_done: Optional[Callable[[ExportResult], None]] = None,
) -> None:
    """The ONE call a page's Export button needs -- see this module's
    docstring for the full worked example. Prompts for a save location
    synchronously (a no-op, instant return if the user cancels: nothing
    else happens, no thread started, no dialog shown), then writes the
    workbook on a background thread behind a LoadingOverlay parented to
    `widget` -- exactly the run_in_background + LoadingOverlay pattern
    ui/inventory_upload_page.py and ui/sales_upload_page.py already use
    for uploads, applied here to the export direction. Shows a success or
    error messagebox when done; `on_done(ExportResult)`, if given, fires
    after that so the caller can react further (e.g. update a status
    label) without needing its own success/error UI."""
    if not rows:
        messagebox.showinfo("Nothing to Export", "There is nothing to export for the current view.")
        return

    file_path = prompt_save_path(suggested_filename, parent=widget)
    if not file_path:
        return  # user cancelled -- no thread, no write, no message

    overlay = LoadingOverlay(widget)
    overlay.show()

    def work(report_progress):
        return write_rows_to_excel(
            rows, columns, headings, file_path,
            sheet_title=sheet_title, row_style_fn=row_style_fn, progress_callback=report_progress,
        )

    def on_progress(percent, message):
        overlay.update_progress(percent, message)

    def on_thread_done(result: ExportResult, error):
        if error is not None:
            overlay.hide()
            logger.error(f"Export: unexpected error exporting to {file_path!r}: {error}")
            messagebox.showerror("Export Failed", f"Could not export the table.\n\n{error}")
            if on_done:
                on_done(ExportResult(success=False, file_path=file_path, error_message=str(error)))
            return

        widget.after(400, overlay.hide)
        if result.success:
            messagebox.showinfo(
                "Export Complete", f"Exported {result.rows_written} row(s) to:\n{result.file_path}"
            )
        else:
            messagebox.showerror("Export Failed", f"Could not export the table.\n\n{result.error_message}")
        if on_done:
            on_done(result)

    run_in_background(widget, work, on_progress=on_progress, on_done=on_thread_done)
